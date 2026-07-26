"""Phase 13q: Channel Partner Meetings - Round E follow-up.

The Meetings list and its export must surface *who at the Channel Partner*
is being met (operational_metadata.contact_name), not just the internal
Sales Manager/RM owners - previously there was no way to tell from the
Meetings screen who the actual meeting is with.
"""


def _seed_capabilities(db, tenant_id):
    from app.models.organisation import BusinessRole, PermissionDefinition, RolePermission

    role = BusinessRole.query.filter_by(tenant_id=tenant_id, key='SALES_MANAGER').first()
    if not role:
        role = BusinessRole(tenant_id=tenant_id, key='SALES_MANAGER', display_name='Sales Manager', is_system=True)
        db.session.add(role)
        db.session.flush()
    for key in ('visits.view', 'visits.manage'):
        perm = PermissionDefinition.query.filter_by(key=key).first()
        if not perm:
            perm = PermissionDefinition(key=key, module='visits', action='VIEW' if key.endswith('view') else 'MANAGE')
            db.session.add(perm)
            db.session.flush()
        db.session.add(RolePermission(
            tenant_id=tenant_id, business_role_id=role.id, permission_id=perm.id,
            scope_type='TENANT', effect='ALLOW',
        ))
    db.session.flush()
    return role


def _bootstrap(app):
    from app.models.base import db
    from app.models.channel_partner import ChannelPartner
    from app.models.organisation import UserBusinessRole
    from app.models.product import Product, TenantProduct
    from app.models.tenant import Tenant
    from app.models.user import User
    from app.models.visit import Visit, VisitParticipant
    from app.utils.jwt import create_token

    db.create_all()
    tenant = Tenant(name='Meetings Round E Tenant', slug='meetings-round-e-tenant')
    db.session.add(tenant)
    db.session.flush()
    product = Product.query.filter_by(slug='lms').first()
    assert product
    db.session.add(TenantProduct(tenant_id=tenant.id, product_id=product.id, status='active'))
    role = _seed_capabilities(db, tenant.id)

    manager = User(
        name='Manager', email='meetingsrounde-mgr@example.invalid',
        password_hash='x', role='sales_manager', tenant_id=tenant.id, is_active=True,
    )
    db.session.add(manager)
    db.session.flush()
    db.session.add(UserBusinessRole(
        tenant_id=tenant.id, user_id=manager.id, business_role_id=role.id, is_primary=True,
    ))

    partner = ChannelPartner(
        tenant_id=tenant.id, code='CPRE', partner_type='INDIVIDUAL', name='Round E Partner',
        created_by=manager.id, updated_by=manager.id,
    )
    db.session.add(partner)
    db.session.flush()

    from app.models.visit import VisitStatusConfiguration, VisitTypeConfiguration
    db.session.add(VisitTypeConfiguration(
        tenant_id=tenant.id, internal_key='MEETING', display_name='Meeting', is_active=True,
    ))
    for status_key in ('SCHEDULED', 'COMPLETED'):
        db.session.add(VisitStatusConfiguration(
            tenant_id=tenant.id, internal_key=status_key, display_name=status_key.title(), is_active=True,
        ))
    db.session.flush()

    visit = Visit(
        tenant_id=tenant.id, visit_type_key='MEETING', status_key='SCHEDULED',
        assigned_user_id=manager.id, purpose='Contact-tracked meeting',
        operational_metadata={'contact_name': 'Priya Sharma'},
        created_by=manager.id, updated_by=manager.id,
    )
    db.session.add(visit)
    db.session.flush()
    db.session.add(VisitParticipant(
        tenant_id=tenant.id, visit_id=visit.id, participant_type='CHANNEL_PARTNER',
        reference_id=partner.id, display_name=partner.name, is_primary=True,
    ))
    db.session.commit()

    token = create_token(str(manager.id), 'sales_manager', tenant.id, login_context='tenant')
    headers = {'Authorization': f'Bearer {token}', 'X-Product-Slug': 'lms'}
    return tenant, visit, headers


def test_list_visits_surfaces_contact_name():
    from app import create_app

    app = create_app('testing')
    with app.app_context():
        tenant, visit, headers = _bootstrap(app)
        client = app.test_client()

        resp = client.get('/api/visits?participant_type=CHANNEL_PARTNER', headers=headers)
        assert resp.status_code == 200, resp.get_data(as_text=True)
        rows = resp.get_json()['visits']
        assert len(rows) == 1
        assert rows[0]['contact_name'] == 'Priya Sharma'


def test_export_includes_contact_name_column():
    import openpyxl
    from io import BytesIO

    from app import create_app

    app = create_app('testing')
    with app.app_context():
        tenant, visit, headers = _bootstrap(app)
        client = app.test_client()

        resp = client.get('/api/visits/export?participant_type=CHANNEL_PARTNER', headers=headers)
        assert resp.status_code == 200, resp.get_data(as_text=True)
        wb = openpyxl.load_workbook(BytesIO(resp.data))
        ws = wb.active
        header_row = [c.value for c in ws[1]]
        assert 'Contact' in header_row
        contact_col = header_row.index('Contact') + 1
        assert ws.cell(row=2, column=contact_col).value == 'Priya Sharma'


def test_inline_status_update_via_put():
    from app import create_app

    app = create_app('testing')
    with app.app_context():
        tenant, visit, headers = _bootstrap(app)
        client = app.test_client()
        json_headers = dict(headers)
        json_headers['Content-Type'] = 'application/json'

        resp = client.put(
            f'/api/visits/{visit.id}', headers=json_headers, json={'status_key': 'COMPLETED'},
        )
        assert resp.status_code == 200, resp.get_data(as_text=True)

        follow_up = client.get('/api/visits?participant_type=CHANNEL_PARTNER&status=COMPLETED', headers=headers)
        rows = follow_up.get_json()['visits']
        assert len(rows) == 1
        assert rows[0]['status_key'] == 'COMPLETED'
