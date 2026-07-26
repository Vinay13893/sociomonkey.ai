"""Unified tenant-scoped Visit administration APIs."""

from datetime import datetime
import re
from uuid import uuid4

from flask import Blueprint, jsonify, request, send_file
from sqlalchemy import or_

from app.middleware import require_capability
from app.models.activity import ActivityLog
from app.models.base import db
from app.models.channel_partner import ChannelPartner
from app.models.lead import Lead
from app.models.visit import (
    Visit, VisitAttachment, VisitParticipant, VisitStatusConfiguration,
    VisitTag, VisitTypeConfiguration,
)
from app.services.channel_partner_events import notify_channel_partner_visit
from app.services.visit_builder import (
    parse_datetime as _shared_parse_datetime,
    validate_reference as _shared_validate_reference,
    validate_user as _shared_validate_user,
    validate_configuration as _shared_validate_configuration,
    validate_visit_payload as _shared_validate_visit_payload,
)
from app.utils.time_utils import to_ist_str


visits_bp = Blueprint('visits', __name__, url_prefix='/api/visits')

PARTICIPANT_TYPES = {
    'LEAD', 'CHANNEL_PARTNER', 'CUSTOMER', 'USER', 'ORGANISATION', 'OTHER',
}
PRIORITIES = {'LOW', 'NORMAL', 'HIGH', 'URGENT'}
VISIBILITY_VALUES = {'VISIBLE', 'HIDDEN'}
CONFIG_FIELDS = {
    'display_name', 'display_order', 'colour', 'is_active', 'visibility',
}


def _tenant_id():
    return request.current_user.tenant_id or getattr(request, 'current_tenant_id', None)


def _correlation_id():
    return str(request.headers.get('X-Correlation-ID') or uuid4())


def _audit(action, resource_id, old_value, new_value, correlation_id,
           resource_type='Visit'):
    db.session.add(ActivityLog(
        tenant_id=_tenant_id(), user_id=request.current_user.id,
        action=action, module='visits', resource_type=resource_type,
        resource_id=resource_id, old_value=old_value, new_value=new_value,
        correlation_id=correlation_id, ip_address=request.remote_addr,
    ))


def _visit(visit_id):
    return Visit.query.filter_by(id=visit_id, tenant_id=_tenant_id()).first()


# The functions below are thin tenant-scoped wrappers around
# app.services.visit_builder, kept under their original names so every
# existing call site in this file is unchanged. The actual validation logic
# lives in visit_builder so app.routes.pipeline and
# app.routes.gallery_operations can reuse it without importing this routes
# module.

def _parse_datetime(value, field):
    return _shared_parse_datetime(value, field)


def _validate_reference(model, value, label, required=False, active_only=False):
    return _shared_validate_reference(
        _tenant_id(), model, value, label, required=required, active_only=active_only,
    )


def _validate_user(value, label):
    return _shared_validate_user(_tenant_id(), value, label)


def _validate_configuration(model, internal_key, label, allow_inactive_key=None):
    return _shared_validate_configuration(
        _tenant_id(), model, internal_key, label, allow_inactive_key=allow_inactive_key,
    )


def _validate_visit_payload(data, current=None):
    return _shared_validate_visit_payload(_tenant_id(), data, current=current)


def _sync_participants(row, values):
    if values is None:
        return
    row.participants[:] = []
    seen = set()
    for value in values:
        participant_type = str(value.get('participant_type') or '').upper()
        if participant_type not in PARTICIPANT_TYPES:
            raise ValueError('Invalid participant type')
        reference_id = value.get('reference_id')
        display_name = str(value.get('display_name') or '').strip() or None
        if participant_type == 'LEAD' and reference_id:
            ref = _validate_reference(Lead, reference_id, 'Participant lead')
            display_name = display_name or ref.name
        elif participant_type == 'CHANNEL_PARTNER' and reference_id:
            ref = _validate_reference(
                ChannelPartner, reference_id, 'Participant Channel Partner',
                active_only=True,
            )
            display_name = display_name or ref.name
        elif participant_type == 'USER' and reference_id:
            ref = _validate_user(reference_id, 'Participant user')
            display_name = display_name or ref.name
        elif not reference_id and not display_name:
            raise ValueError('Participant name or reference is required')
        identity = (participant_type, int(reference_id) if reference_id else None, display_name)
        if identity in seen:
            continue
        seen.add(identity)
        row.participants.append(VisitParticipant(
            tenant_id=_tenant_id(), participant_type=participant_type,
            reference_id=int(reference_id) if reference_id else None,
            display_name=display_name, is_primary=bool(value.get('is_primary')),
            participant_metadata=value.get('participant_metadata') or {},
        ))


def _sync_tags(row, values):
    if values is None:
        return
    tags = []
    for value in values:
        tag = str(value or '').strip()
        if tag and tag.lower() not in {item.lower() for item in tags}:
            tags.append(tag[:80])
    row.tags[:] = [VisitTag(tenant_id=_tenant_id(), tag=tag) for tag in tags]


def _sync_attachments(row, values):
    if values is None:
        return
    row.attachments[:] = []
    for value in values:
        file_name = str(value.get('file_name') or '').strip()
        storage_reference = str(value.get('storage_reference') or '').strip()
        if not file_name or not storage_reference:
            raise ValueError('Attachment file name and storage reference are required')
        row.attachments.append(VisitAttachment(
            tenant_id=_tenant_id(), file_name=file_name,
            mime_type=value.get('mime_type'), storage_reference=storage_reference,
            attachment_metadata=value.get('attachment_metadata') or {},
            created_by=request.current_user.id,
        ))


def _apply_visit_fields(row, data, validated):
    for key, value in validated.items():
        setattr(row, key, value)
    for field in (
        'purpose', 'notes', 'source', 'operational_metadata', 'token_code',
    ):
        if field in data:
            setattr(row, field, data[field])
    _sync_participants(row, data.get('participants') if 'participants' in data else None)
    _sync_tags(row, data.get('tags') if 'tags' in data else None)
    _sync_attachments(row, data.get('attachments') if 'attachments' in data else None)


@visits_bp.get('/configuration')
@require_capability('visits.view', 'TENANT')
def list_visit_configuration():
    types = VisitTypeConfiguration.query.filter_by(tenant_id=_tenant_id()).order_by(
        VisitTypeConfiguration.display_order, VisitTypeConfiguration.id
    ).all()
    statuses = VisitStatusConfiguration.query.filter_by(tenant_id=_tenant_id()).order_by(
        VisitStatusConfiguration.display_order, VisitStatusConfiguration.id
    ).all()
    return jsonify({
        'visit_types': [row.to_dict() for row in types],
        'visit_statuses': [row.to_dict() for row in statuses],
    })


def _update_configuration(model, internal_key, resource_type):
    key = str(internal_key or '').upper()
    row = model.query.filter_by(tenant_id=_tenant_id(), internal_key=key).first()
    if not row:
        return jsonify({'error': 'Visit configuration not found'}), 404
    data = request.get_json() or {}
    if data.get('visibility') and data['visibility'] not in VISIBILITY_VALUES:
        return jsonify({'error': 'Invalid visibility'}), 400
    if 'display_name' in data and not str(data['display_name'] or '').strip():
        return jsonify({'error': 'Display name is required'}), 400
    old = row.to_dict()
    for field in CONFIG_FIELDS | ({'is_terminal'} if model is VisitStatusConfiguration else set()):
        if field in data:
            setattr(row, field, data[field])
    row.updated_by = request.current_user.id
    correlation_id = _correlation_id()
    _audit('visit_configuration_updated', row.id, old, row.to_dict(),
           correlation_id, resource_type=resource_type)
    db.session.commit()
    return jsonify({'configuration': row.to_dict(), 'correlation_id': correlation_id})


def _create_configuration(model, resource_type):
    data = request.get_json() or {}
    key = str(data.get('internal_key') or '').strip().upper()
    name = str(data.get('display_name') or '').strip()
    if not re.fullmatch(r'[A-Z][A-Z0-9_]{0,79}', key) or not name:
        return jsonify({
            'error': 'Internal key and display name are required; key must use uppercase letters, numbers and underscores',
        }), 400
    if model.query.filter_by(tenant_id=_tenant_id(), internal_key=key).first():
        return jsonify({'error': 'Visit configuration key already exists'}), 409
    visibility = str(data.get('visibility') or 'VISIBLE').upper()
    if visibility not in VISIBILITY_VALUES:
        return jsonify({'error': 'Invalid visibility'}), 400
    row = model(
        tenant_id=_tenant_id(), internal_key=key, display_name=name,
        display_order=int(data.get('display_order') or 0),
        colour=data.get('colour') or '#64748b',
        is_active=bool(data.get('is_active', True)),
        visibility=visibility,
        updated_by=request.current_user.id,
    )
    if isinstance(row, VisitStatusConfiguration):
        row.is_terminal = bool(data.get('is_terminal', False))
    db.session.add(row)
    db.session.flush()
    correlation_id = _correlation_id()
    _audit('visit_configuration_created', row.id, None, row.to_dict(),
           correlation_id, resource_type=resource_type)
    db.session.commit()
    return jsonify({'configuration': row.to_dict(), 'correlation_id': correlation_id}), 201


@visits_bp.post('/configuration/types')
@require_capability('visits.manage', 'TENANT')
def create_visit_type():
    return _create_configuration(VisitTypeConfiguration, 'VisitTypeConfiguration')


@visits_bp.post('/configuration/statuses')
@require_capability('visits.manage', 'TENANT')
def create_visit_status():
    return _create_configuration(VisitStatusConfiguration, 'VisitStatusConfiguration')


@visits_bp.put('/configuration/types/<string:internal_key>')
@require_capability('visits.manage', 'TENANT')
def update_visit_type(internal_key):
    return _update_configuration(VisitTypeConfiguration, internal_key, 'VisitTypeConfiguration')


@visits_bp.put('/configuration/statuses/<string:internal_key>')
@require_capability('visits.manage', 'TENANT')
def update_visit_status(internal_key):
    return _update_configuration(VisitStatusConfiguration, internal_key, 'VisitStatusConfiguration')


def _apply_visit_filters(query):
    """Shared filter-building for list_visits and export_visits - keeping
    this in one place means the export always matches exactly what the
    on-screen list shows for the same query params."""
    active = str(request.args.get('active', 'true')).lower()
    if active in ('true', 'false'):
        query = query.filter(Visit.is_active == (active == 'true'))
    for argument, column in (
        ('status', Visit.status_key), ('type', Visit.visit_type_key),
        ('location_id', Visit.location_id), ('project_id', Visit.project_id),
        ('assigned_user_id', Visit.assigned_user_id), ('lead_id', Visit.lead_id),
    ):
        value = request.args.get(argument)
        if value not in (None, ''):
            try:
                parsed = int(value) if argument.endswith('_id') else value.upper()
            except ValueError:
                raise ValueError(f'{argument} must be numeric')
            query = query.filter(column == parsed)
    date_from = _parse_datetime(request.args.get('date_from'), 'Date from')
    date_to = _parse_datetime(request.args.get('date_to'), 'Date to')
    if date_from:
        query = query.filter(Visit.expected_arrival >= date_from)
    if date_to:
        query = query.filter(Visit.expected_arrival <= date_to)
    search = str(request.args.get('search') or '').strip()
    if search:
        like = f'%{search}%'
        query = query.outerjoin(Lead, Lead.id == Visit.lead_id).filter(or_(
            Visit.purpose.ilike(like), Visit.source.ilike(like), Lead.name.ilike(like),
        ))

    # Opt-in participant scoping (e.g. Channel Partner "Meetings") - never
    # applied unless explicitly requested, so Reception's existing
    # unrestricted TENANT-wide visit listing stays exactly as it was.
    participant_type = str(request.args.get('participant_type') or '').upper()
    if participant_type:
        participant_visit_ids = db.session.query(VisitParticipant.visit_id).filter(
            VisitParticipant.tenant_id == _tenant_id(),
            VisitParticipant.participant_type == participant_type,
        )
        query = query.filter(Visit.id.in_(participant_visit_ids))
        visible_ids = _visible_assignee_ids(request.current_user)
        if visible_ids is not None:
            # Visible if the caller (or their reporting-line team) holds
            # either owner slot - a Sales Manager must see meetings their
            # RM is personally attending even though assigned_user_id
            # names the RM, not them.
            query = query.filter(or_(
                Visit.assigned_user_id.in_(visible_ids),
                Visit.sales_manager_id.in_(visible_ids),
            ))
    return query, participant_type


@visits_bp.get('')
@require_capability('visits.view', 'TENANT')
def list_visits():
    page = max(1, request.args.get('page', 1, type=int))
    per_page = min(100, max(1, request.args.get('per_page', 25, type=int)))
    query = Visit.query.filter_by(tenant_id=_tenant_id())
    try:
        query, participant_type = _apply_visit_filters(query)
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400

    total = query.count()
    rows = query.order_by(
        Visit.expected_arrival.desc().nullslast(), Visit.id.desc()
    ).offset((page - 1) * per_page).limit(per_page).all()
    serialized = []
    for row in rows:
        item = row.to_dict(include_details=False)
        if participant_type:
            participant = next(
                (p for p in row.participants if p.participant_type == participant_type),
                None,
            )
            item['participant_id'] = participant.reference_id if participant else None
            item['participant_name'] = participant.display_name if participant else None
        serialized.append(item)
    return jsonify({
        'visits': serialized,
        'pagination': {'page': page, 'per_page': per_page, 'total': total},
    })


@visits_bp.get('/export')
@require_capability('visits.view', 'TENANT')
def export_visits():
    """Excel export mirroring exactly whatever filters the on-screen
    list is currently using (same query-building as list_visits)."""
    query = Visit.query.filter_by(tenant_id=_tenant_id())
    try:
        query, participant_type = _apply_visit_filters(query)
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400
    rows = query.order_by(
        Visit.expected_arrival.desc().nullslast(), Visit.id.desc()
    ).limit(5000).all()

    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Meetings' if participant_type == 'CHANNEL_PARTNER' else 'Visits'

    headers = ['#', 'Channel Partner', 'Assigned To', 'Sales Manager', 'Purpose', 'Expected Arrival', 'Status', 'Location', 'Venue Note']
    header_fill = PatternFill('solid', fgColor='1E3A5F')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    alt_fill = PatternFill('solid', fgColor='EEF2F7')
    for row_idx, row in enumerate(rows, 2):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill()
        participant = next(
            (p for p in row.participants if p.participant_type == 'CHANNEL_PARTNER'),
            None,
        )
        values = [
            row.id,
            participant.display_name if participant else '',
            row.assigned_user.name if row.assigned_user else '',
            row.sales_manager.name if row.sales_manager else '',
            row.purpose or '',
            to_ist_str(row.expected_arrival) if row.expected_arrival else '',
            row.status_key,
            row.location.name if row.location else '',
            (row.operational_metadata or {}).get('venue_note', ''),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = fill

    col_widths = [6, 22, 18, 18, 30, 20, 14, 18, 26]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'meetings_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


def _visible_assignee_ids(user):
    """None means unrestricted (admin-like); otherwise the set of user ids
    whose visits are visible to *user* - mirrors the Channel Partner
    ownership visibility model (app.routes.channel_partners)."""
    if user.role in ('superadmin', 'platform_owner'):
        return None
    from app.utils.leads import _reporting_team_ids
    return _reporting_team_ids(user)


@visits_bp.post('')
@require_capability('visits.manage', 'TENANT')
def create_visit():
    data = request.get_json() or {}
    try:
        validated = _validate_visit_payload(data)
        row = Visit(
            tenant_id=_tenant_id(), created_by=request.current_user.id,
            updated_by=request.current_user.id,
        )
        _apply_visit_fields(row, data, validated)
        db.session.add(row)
        db.session.flush()
        correlation_id = _correlation_id()
        _audit('visit_created', row.id, None, row.to_dict(), correlation_id)
        if row.status_key == 'CHECKED_IN':
            notify_channel_partner_visit(
                row, 'visit_arrival', correlation_id
            )
        elif row.status_key == 'COMPLETED':
            notify_channel_partner_visit(
                row, 'visit_completed', correlation_id
            )
        db.session.commit()
        return jsonify({'visit': row.to_dict(), 'correlation_id': correlation_id}), 201
    except (TypeError, ValueError) as exc:
        db.session.rollback()
        return jsonify({'error': str(exc)}), 400


@visits_bp.get('/<int:visit_id>')
@require_capability('visits.view', 'TENANT')
def get_visit(visit_id):
    row = _visit(visit_id)
    if not row:
        return jsonify({'error': 'Visit not found'}), 404
    return jsonify({'visit': row.to_dict()})


@visits_bp.put('/<int:visit_id>')
@require_capability('visits.manage', 'TENANT')
def update_visit(visit_id):
    row = _visit(visit_id)
    if not row:
        return jsonify({'error': 'Visit not found'}), 404
    data = request.get_json() or {}
    old = row.to_dict()
    old_status = row.status_key
    try:
        validated = _validate_visit_payload(data, current=row)
        _apply_visit_fields(row, data, validated)
        row.updated_by = request.current_user.id
        db.session.flush()
        correlation_id = _correlation_id()
        action = 'visit_lifecycle_changed' if row.status_key != old_status else 'visit_updated'
        _audit(action, row.id, old, row.to_dict(), correlation_id)
        if row.status_key != old_status and row.status_key == 'CHECKED_IN':
            notify_channel_partner_visit(
                row, 'visit_arrival', correlation_id
            )
        elif row.status_key != old_status and row.status_key == 'COMPLETED':
            notify_channel_partner_visit(
                row, 'visit_completed', correlation_id
            )
        db.session.commit()
        return jsonify({'visit': row.to_dict(), 'correlation_id': correlation_id})
    except (TypeError, ValueError) as exc:
        db.session.rollback()
        return jsonify({'error': str(exc)}), 400


def _set_active(visit_id, active):
    row = _visit(visit_id)
    if not row:
        return jsonify({'error': 'Visit not found'}), 404
    old = row.to_dict()
    row.is_active = active
    row.archived_at = None if active else datetime.utcnow()
    row.updated_by = request.current_user.id
    correlation_id = _correlation_id()
    _audit('visit_restored' if active else 'visit_archived',
           row.id, old, row.to_dict(), correlation_id)
    db.session.commit()
    return jsonify({'visit': row.to_dict(), 'correlation_id': correlation_id})


@visits_bp.post('/<int:visit_id>/archive')
@require_capability('visits.manage', 'TENANT')
def archive_visit(visit_id):
    return _set_active(visit_id, False)


@visits_bp.post('/<int:visit_id>/restore')
@require_capability('visits.manage', 'TENANT')
def restore_visit(visit_id):
    return _set_active(visit_id, True)
