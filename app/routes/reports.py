from datetime import datetime, timedelta

from flask import Blueprint, jsonify, request, send_file
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy import case, func

from app import db
from app.middleware import require_auth, require_role
from app.models.activity import ActivityLog
from app.models.lead import Lead, StatusHistory, CallbackReminder
from app.models.project import Project
from app.models.user import User
from app.services.reports import ReportService
from app.utils.leads import get_user_visible_leads, apply_test_lead_filter, apply_valid_lead_capture_scope
from app.utils.time_utils import IST

reports_bp = Blueprint('reports', __name__, url_prefix='/api/reports')


def _resolve_date_window(range_key: str, date_from: str, date_to: str):
    now = datetime.utcnow()
    key = (range_key or '').strip().lower()

    if date_from or date_to:
        try:
            start = datetime.strptime(date_from, '%Y-%m-%d') if date_from else None
            end = datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1) if date_to else None
            return start, end
        except ValueError:
            return None, None

    today_start = datetime(now.year, now.month, now.day)
    if key == 'today':
        return today_start, today_start + timedelta(days=1)
    if key == 'yesterday':
        start = today_start - timedelta(days=1)
        return start, today_start
    if key == 'last_week':
        this_week_start = today_start - timedelta(days=today_start.weekday())
        start = this_week_start - timedelta(days=7)
        return start, this_week_start
    if key == 'last_30_days':
        return now - timedelta(days=30), now + timedelta(seconds=1)
    if key == 'this_month':
        start = datetime(now.year, now.month, 1)
        return start, now + timedelta(seconds=1)
    if key == 'last_month':
        this_month_start = datetime(now.year, now.month, 1)
        prev_end = this_month_start
        prev_start = datetime(prev_end.year if prev_end.month > 1 else prev_end.year - 1,
                              prev_end.month - 1 if prev_end.month > 1 else 12,
                              1)
        return prev_start, prev_end

    return None, None


def _apply_created_date_filters(query, range_key: str, date_from: str, date_to: str):
    start, end = _resolve_date_window(range_key, date_from, date_to)
    if start:
        query = query.filter(Lead.created_at >= start)
    if end:
        query = query.filter(Lead.created_at < end)
    return query


def _safe_int(value):
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _compact_activity_row(row):
    return {
        'id': row.id,
        'user_id': row.user_id,
        'user_name': row.user_name,
        'action': row.action,
        'module': row.module,
        'resource_id': row.resource_id,
        'resource_type': row.resource_type,
        'description': row.description,
        'ip_address': row.ip_address,
        'created_at': row.created_at.isoformat() if row.created_at else None,
    }


def _activity_query_for_user(user, user_id_param=None):
    query = ActivityLog.query
    if user.tenant_id is not None:
        query = query.filter(ActivityLog.tenant_id == user.tenant_id)

    requested_user_id = _safe_int(user_id_param)
    if user.role == 'team_member':
        return query.filter(ActivityLog.user_id == user.id)

    if user.role == 'sales_manager':
        team_ids = [
            r[0] for r in User.query
            .filter_by(manager_id=user.id, is_active=True, tenant_id=user.tenant_id)
            .with_entities(User.id)
            .all()
        ]
        team_ids.append(user.id)
        query = query.filter(ActivityLog.user_id.in_(team_ids))
        if requested_user_id in team_ids:
            query = query.filter(ActivityLog.user_id == requested_user_id)
        return query

    if requested_user_id:
        query = query.filter(ActivityLog.user_id == requested_user_id)
    return query


def _apply_activity_filters(query, action=None, module=None, date_from=None, date_to=None):
    if action:
        query = query.filter(ActivityLog.action == action)
    if module:
        query = query.filter(ActivityLog.module == module)
    if date_from:
        try:
            query = query.filter(ActivityLog.created_at >= datetime.strptime(date_from, '%Y-%m-%d'))
        except ValueError:
            pass
    if date_to:
        try:
            query = query.filter(ActivityLog.created_at < datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1))
        except ValueError:
            pass
    return query


def _period_metrics(user, start_dt: datetime, end_dt: datetime):
    visible_ids = get_user_visible_leads(user).with_entities(Lead.id).subquery()
    visible_count = db.session.query(func.count(visible_ids.c.id)).scalar() or 0
    if not visible_count:
        return {
            'leads_added': 0,
            'calls_done': 0,
            'follow_ups': 0,
            'lost': 0,
            'site_visits': 0,
            'negotiations': 0,
            'closures': 0,
        }

    leads_added = Lead.query.filter(
        Lead.id.in_(visible_ids),
        Lead.created_at >= start_dt,
        Lead.created_at < end_dt,
    ).count()

    # Unique leads that had at least one call in the period
    calls_done = db.session.query(ActivityLog.resource_id).filter(
        ActivityLog.module == 'leads',
        ActivityLog.resource_id.in_(visible_ids),
        ActivityLog.action.like('call_%'),
        ActivityLog.action != 'call_initiated',
        ActivityLog.created_at >= start_dt,
        ActivityLog.created_at < end_dt,
    ).distinct().count()

    follow_ups = StatusHistory.query.join(Lead, Lead.id == StatusHistory.lead_id).filter(
        Lead.id.in_(visible_ids),
        StatusHistory.new_status.in_(['follow_up', 'callback_scheduled', 'interested', 'site_visit_planned']),
        StatusHistory.changed_at >= start_dt,
        StatusHistory.changed_at < end_dt,
    ).count()

    lost = StatusHistory.query.join(Lead, Lead.id == StatusHistory.lead_id).filter(
        Lead.id.in_(visible_ids),
        StatusHistory.new_status.in_(['not_interested', 'lost', 'junk']),
        StatusHistory.changed_at >= start_dt,
        StatusHistory.changed_at < end_dt,
    ).count()

    site_visits = StatusHistory.query.join(Lead, Lead.id == StatusHistory.lead_id).filter(
        Lead.id.in_(visible_ids),
        StatusHistory.new_status == 'site_visit_done',
        StatusHistory.changed_at >= start_dt,
        StatusHistory.changed_at < end_dt,
    ).count()

    negotiations = StatusHistory.query.join(Lead, Lead.id == StatusHistory.lead_id).filter(
        Lead.id.in_(visible_ids),
        StatusHistory.new_status == 'negotiation',
        StatusHistory.changed_at >= start_dt,
        StatusHistory.changed_at < end_dt,
    ).count()

    closures = StatusHistory.query.join(Lead, Lead.id == StatusHistory.lead_id).filter(
        Lead.id.in_(visible_ids),
        StatusHistory.new_status == 'booking_done',
        StatusHistory.changed_at >= start_dt,
        StatusHistory.changed_at < end_dt,
    ).count()

    return {
        'leads_added': leads_added,
        'calls_done': calls_done,
        'follow_ups': follow_ups,
        'lost': lost,
        'site_visits': site_visits,
        'negotiations': negotiations,
        'closures': closures,
    }


def _build_comparison_payload(user):
    now = datetime.utcnow()
    this_week_start = datetime(now.year, now.month, now.day) - timedelta(days=datetime(now.year, now.month, now.day).weekday())
    last_week_start = this_week_start - timedelta(days=7)
    this_month_start = datetime(now.year, now.month, 1)
    last_month_end = this_month_start
    last_month_start = datetime(last_month_end.year if last_month_end.month > 1 else last_month_end.year - 1,
                                last_month_end.month - 1 if last_month_end.month > 1 else 12,
                                1)

    week_current = _period_metrics(user, this_week_start, now + timedelta(seconds=1))
    week_last = _period_metrics(user, last_week_start, this_week_start)
    month_current = _period_metrics(user, this_month_start, now + timedelta(seconds=1))
    month_last = _period_metrics(user, last_month_start, last_month_end)

    return {
        'week': {
            'label_current': 'This Week',
            'label_previous': 'Last Week',
            'current': week_current,
            'previous': week_last,
        },
        'month': {
            'label_current': 'This Month',
            'label_previous': 'Last Month',
            'current': month_current,
            'previous': month_last,
        },
    }


# ---------------------------------------------------------------------------
# JSON reports
# ---------------------------------------------------------------------------

@reports_bp.route('/leads', methods=['GET'])
@require_auth
def lead_report():
    user = request.current_user
    query = get_user_visible_leads(user)

    date_from = request.args.get('date_from')
    date_to   = request.args.get('date_to')
    project_id = request.args.get('project_id', type=int)
    if date_from:
        try:
            query = query.filter(Lead.created_at >= datetime.strptime(date_from, '%Y-%m-%d'))
        except ValueError:
            pass
    if date_to:
        try:
            dt_to = datetime.strptime(date_to, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
            query = query.filter(Lead.created_at <= dt_to)
        except ValueError:
            pass

    if project_id:
        query = query.filter(Lead.project_id == project_id)

    report_sq = query.with_entities(
        Lead.id.label('id'),
        Lead.status.label('status'),
        Lead.source.label('source'),
        Lead.project_id.label('project_id'),
        Lead.created_at.label('created_at'),
    ).subquery()

    total = db.session.query(func.count(report_sq.c.id)).scalar() or 0

    converted_statuses = {'booking_done', 'negotiation'}
    converted = (
        db.session.query(func.count(report_sq.c.id))
        .filter(report_sq.c.status.in_(converted_statuses))
        .scalar()
        or 0
    )
    conversion_rate = round(converted / total * 100, 2) if total else 0

    status_counts = {
        key or 'unknown': int(count or 0)
        for key, count in db.session.query(
            func.coalesce(report_sq.c.status, 'unknown'),
            func.count(report_sq.c.id),
        ).group_by(func.coalesce(report_sq.c.status, 'unknown')).all()
    }

    source_counts = {
        key or 'unknown': int(count or 0)
        for key, count in db.session.query(
            func.coalesce(report_sq.c.source, 'unknown'),
            func.count(report_sq.c.id),
        ).group_by(func.coalesce(report_sq.c.source, 'unknown')).all()
    }

    project_counts = {
        key or 'Unassigned': int(count or 0)
        for key, count in db.session.query(
            func.coalesce(Project.name, 'Unassigned'),
            func.count(report_sq.c.id),
        )
        .outerjoin(Project, Project.id == report_sq.c.project_id)
        .group_by(func.coalesce(Project.name, 'Unassigned'))
        .all()
    }

    terminal_statuses = ['booking_done', 'lost', 'junk', 'not_interested']
    stale_cutoff = datetime.utcnow() - timedelta(days=5)
    today_start = datetime(datetime.utcnow().year, datetime.utcnow().month, datetime.utcnow().day)
    unassigned_count = query.filter(Lead.assigned_to.is_(None)).count()
    pending_callbacks = db.session.query(func.count(func.distinct(CallbackReminder.lead_id))).filter(
        CallbackReminder.tenant_id == user.tenant_id,
        CallbackReminder.status == 'pending',
        CallbackReminder.lead_id.in_(query.with_entities(Lead.id)),
    ).scalar() or 0
    overdue_callbacks = db.session.query(func.count(func.distinct(CallbackReminder.lead_id))).filter(
        CallbackReminder.tenant_id == user.tenant_id,
        CallbackReminder.status == 'pending',
        CallbackReminder.callback_datetime < datetime.utcnow(),
        CallbackReminder.lead_id.in_(query.with_entities(Lead.id)),
    ).scalar() or 0
    untouched = (
        db.session.query(func.count(report_sq.c.id))
        .filter(report_sq.c.status == 'new')
        .scalar()
        or 0
    )
    stale = query.filter(Lead.status.notin_(terminal_statuses), Lead.updated_at <= stale_cutoff).count()
    carry_forward = query.filter(Lead.created_at < today_start, Lead.status.notin_(terminal_statuses)).count()

    # Leads per day — use filtered leads if date filter applied, else last 30 days
    trend_query = db.session.query(
        func.date(report_sq.c.created_at),
        func.count(report_sq.c.id),
    )
    if not (date_from or date_to):
        trend_query = trend_query.filter(report_sq.c.created_at >= datetime.utcnow() - timedelta(days=30))
    by_date = {
        str(day): int(count or 0)
        for day, count in trend_query.group_by(func.date(report_sq.c.created_at)).all()
        if day
    }

    return jsonify({
        'total_leads': total,
        'leads_by_status': status_counts,
        'leads_by_source': source_counts,
        'leads_by_project': project_counts,
        'operational_health': {
            'allocation_unassigned': int(unassigned_count or 0),
            'workload_assigned': int(total - (unassigned_count or 0)),
            'pending_callbacks': int(pending_callbacks or 0),
            'overdue_callbacks': int(overdue_callbacks or 0),
            'untouched': int(untouched or 0),
            'stale': int(stale or 0),
            'carry_forward': int(carry_forward or 0),
        },
        'conversion_rate': conversion_rate,
        'leads_by_date': by_date,
    }), 200


@reports_bp.route('/team', methods=['GET'])
@require_auth
def team_report():
    current_user = request.current_user

    date_from = request.args.get('date_from')
    date_to   = request.args.get('date_to')
    range_key = request.args.get('range')

    scoped_leads = Lead.query.filter_by(is_active=True, tenant_id=current_user.tenant_id)
    scoped_leads = apply_valid_lead_capture_scope(apply_test_lead_filter(scoped_leads), current_user.tenant_id)
    scoped_leads = _apply_created_date_filters(scoped_leads, range_key, date_from, date_to)
    if current_user.role == 'sales_manager':
        team_user_ids = [
            r[0] for r in User.query
            .filter_by(manager_id=current_user.id, is_active=True, tenant_id=current_user.tenant_id)
            .with_entities(User.id)
            .all()
        ]
        scoped_leads = scoped_leads.filter(Lead.assigned_to.in_(team_user_ids + [current_user.id]))

    lead_stats = {}
    for row in (
        scoped_leads.with_entities(
            Lead.assigned_to.label('assigned_to'),
            func.count(Lead.id).label('total'),
            func.sum(case((Lead.status == 'interested', 1), else_=0)).label('interested'),
            func.sum(case((Lead.status == 'site_visit_planned', 1), else_=0)).label('site_visit_planned'),
            func.sum(case((Lead.status == 'site_visit_done', 1), else_=0)).label('site_visit_done'),
            func.sum(case((Lead.status == 'negotiation', 1), else_=0)).label('negotiation'),
            func.sum(case((Lead.status == 'booking_done', 1), else_=0)).label('booking_done'),
        )
        .group_by(Lead.assigned_to)
        .all()
    ):
        lead_stats[row.assigned_to] = {
            'total': int(row.total or 0),
            'interested': int(row.interested or 0),
            'site_visit_planned': int(row.site_visit_planned or 0),
            'site_visit_done': int(row.site_visit_done or 0),
            'negotiation': int(row.negotiation or 0),
            'booking_done': int(row.booking_done or 0),
        }

    def get_stats(u):
        stats = lead_stats.get(u.id, {})
        total = int(stats.get('total') or 0)
        interested = int(stats.get('interested') or 0)
        site_visit_plan = int(stats.get('site_visit_planned') or 0)
        site_visit_done = int(stats.get('site_visit_done') or 0)
        negotiation = int(stats.get('negotiation') or 0)
        booking_done = int(stats.get('booking_done') or 0)
        warm_leads       = interested + site_visit_plan
        hot_leads        = site_visit_done + negotiation
        warm_rate        = round((interested + site_visit_plan) / total * 100, 2) if total else 0
        hot_rate         = round((hot_leads / total) * 100, 2) if total else 0
        return {
            'id': u.id,
            'name': u.name,
            'email': u.email,
            'role': u.role,
            'total_leads': total,
            'interested': interested,
            'site_visit_planned': site_visit_plan,
            'site_visit_done': site_visit_done,
            'negotiation': negotiation,
            'booking_done': booking_done,
            'warm_leads': warm_leads,
            'hot_leads': hot_leads,
            'warm_rate': warm_rate,
            'hot_rate': hot_rate,
            'last_login': u.last_login.isoformat() if u.last_login else None,
        }

    if current_user.role == 'sales_manager':
        # Return just this manager's group
        manager_stats = get_stats(current_user)
        members = [
            get_stats(u)
            for u in User.query.filter_by(
                manager_id=current_user.id,
                is_active=True,
                tenant_id=current_user.tenant_id,
            ).all()
        ]
        return jsonify({
            'team_groups': [{
                'manager': manager_stats,
                'members': sorted(members, key=lambda x: -x['total_leads']),
            }]
        }), 200

    # superadmin: group all team members under their manager
    tid = current_user.tenant_id
    managers = User.query.filter_by(
        role='sales_manager', is_active=True, tenant_id=tid,
    ).order_by(User.name).all()
    groups = []
    assigned_member_ids = set()
    for mgr in managers:
        members = User.query.filter_by(manager_id=mgr.id, is_active=True, tenant_id=tid).all()
        for m in members:
            assigned_member_ids.add(m.id)
        groups.append({
            'manager': get_stats(mgr),
            'members': sorted([get_stats(m) for m in members], key=lambda x: -x['total_leads']),
        })

    # Unassigned team members (no manager) — scoped to tenant
    unassigned = User.query.filter(
        User.role == 'team_member',
        User.is_active == True,
        User.tenant_id == tid,
        ~User.id.in_(assigned_member_ids) if assigned_member_ids else True,
    ).all()

    return jsonify({
        'team_groups': groups,
        'unassigned_members': sorted([get_stats(u) for u in unassigned], key=lambda x: -x['total_leads']),
    }), 200


@reports_bp.route('/comparison', methods=['GET'])
@require_auth
def comparison_report():
    user = request.current_user
    return jsonify({'comparison': _build_comparison_payload(user)}), 200


@reports_bp.route('/activity', methods=['GET'])
@require_role('superadmin')
def activity_report():
    user = request.current_user
    base = ActivityLog.query
    if user.tenant_id is not None:
        base = base.filter(ActivityLog.tenant_id == user.tenant_id)

    user_activity = {
        name or 'Unknown': int(count or 0)
        for name, count in base.with_entities(User.name, func.count(ActivityLog.id))
        .outerjoin(User, User.id == ActivityLog.user_id)
        .group_by(User.name)
        .all()
    }
    action_activity = {
        action or 'unknown': int(count or 0)
        for action, count in base.with_entities(ActivityLog.action, func.count(ActivityLog.id))
        .group_by(ActivityLog.action)
        .all()
    }
    module_activity = {
        module or 'unknown': int(count or 0)
        for module, count in base.with_entities(ActivityLog.module, func.count(ActivityLog.id))
        .group_by(ActivityLog.module)
        .all()
    }

    cutoff = datetime.utcnow() - timedelta(days=7)
    by_date = {
        str(day): int(count or 0)
        for day, count in base.with_entities(func.date(ActivityLog.created_at), func.count(ActivityLog.id))
        .filter(ActivityLog.created_at >= cutoff)
        .group_by(func.date(ActivityLog.created_at))
        .all()
        if day
    }

    return jsonify({
        'activity_by_user': user_activity,
        'activity_by_action': action_activity,
        'activity_by_module': module_activity,
        'activity_last_7_days': by_date,
        'total_activities': base.count(),
    }), 200


@reports_bp.route('/activity-logs', methods=['GET'])
@require_auth
def get_activity_logs():
    user = request.current_user
    user_id_param = request.args.get('user_id')
    action = request.args.get('action')
    module = request.args.get('module')
    limit = request.args.get('limit', 100, type=int)
    page = max(1, request.args.get('page', 1, type=int))
    per_page = min(100, max(10, request.args.get('per_page', limit or 25, type=int)))
    sort = (request.args.get('sort') or 'newest').strip().lower()
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')

    query = _activity_query_for_user(user, user_id_param)

    if user.role == 'team_member':
        # Team members see only their own activity
        query = query.filter_by(user_id=user.id)
    elif user.role == 'sales_manager':
        # Sales managers see own + their direct team members
        team_ids = [m.id for m in User.query.filter_by(manager_id=user.id).all()]
        team_ids.append(user.id)
        query = query.filter(ActivityLog.user_id.in_(team_ids))
        requested_user_id = _safe_int(user_id_param)
        if requested_user_id in team_ids:
            query = query.filter_by(user_id=requested_user_id)
    else:
        # superadmin / platform_owner — see all
        requested_user_id = _safe_int(user_id_param)
        if requested_user_id:
            query = query.filter_by(user_id=requested_user_id)

    if action:
        query = query.filter_by(action=action)
    if module:
        query = query.filter_by(module=module)
    if date_from:
        try:
            query = query.filter(ActivityLog.created_at >= datetime.strptime(date_from, '%Y-%m-%d'))
        except ValueError:
            pass
    if date_to:
        try:
            query = query.filter(ActivityLog.created_at < datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1))
        except ValueError:
            pass

    order_by = ActivityLog.created_at.asc() if sort == 'oldest' else ActivityLog.created_at.desc()
    total = query.count()
    logs = (
        query
        .with_entities(
            ActivityLog.id,
            ActivityLog.user_id,
            User.name.label('user_name'),
            ActivityLog.action,
            ActivityLog.module,
            ActivityLog.resource_id,
            ActivityLog.resource_type,
            ActivityLog.description,
            ActivityLog.ip_address,
            ActivityLog.created_at,
        )
        .outerjoin(User, User.id == ActivityLog.user_id)
        .order_by(order_by)
        .offset((page - 1) * per_page)
        .limit(per_page)
        .all()
    )
    return jsonify({
        'activity_logs': [_compact_activity_row(l) for l in logs],
        'total': total,
        'page': page,
        'per_page': per_page,
    }), 200


@reports_bp.route('/activity-logs/download', methods=['GET'])
@require_auth
def download_activity_logs():
    user = request.current_user
    user_id_param = request.args.get('user_id')
    action  = request.args.get('action')
    module  = request.args.get('module')
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')

    query = ActivityLog.query

    if user.role == 'team_member':
        query = query.filter_by(user_id=user.id)
    elif user.role == 'sales_manager':
        team_ids = [m.id for m in User.query.filter_by(manager_id=user.id).all()]
        team_ids.append(user.id)
        query = query.filter(ActivityLog.user_id.in_(team_ids))
        if user_id_param and int(user_id_param) in team_ids:
            query = query.filter_by(user_id=int(user_id_param))
    else:
        if user_id_param:
            query = query.filter_by(user_id=int(user_id_param))

    if action:
        query = query.filter_by(action=action)
    if module:
        query = query.filter_by(module=module)
    if date_from:
        try:
            query = query.filter(ActivityLog.created_at >= datetime.strptime(date_from, '%Y-%m-%d'))
        except ValueError:
            pass
    if date_to:
        try:
            query = query.filter(ActivityLog.created_at < datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1))
        except ValueError:
            pass

    logs = query.order_by(ActivityLog.created_at.desc()).all()

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Activity Logs'

    headers = ['#', 'User', 'Action', 'Module', 'Resource Type', 'Resource ID', 'Description', 'IP Address', 'Timestamp']
    header_fill = PatternFill('solid', fgColor='1E3A5F')
    header_font = Font(color='FFFFFF', bold=True, size=11)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    alt_fill = PatternFill('solid', fgColor='EEF2F7')
    for row_idx, log in enumerate(logs, 2):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill()
        values = [
            log.id,
            log.user.name if log.user else '',
            log.action,
            log.module,
            log.resource_type or '',
            log.resource_id or '',
            log.description or '',
            log.ip_address or '',
            log.created_at.astimezone(IST).strftime('%d %b %Y %H:%M IST') if log.created_at else '',
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = fill

    col_widths = [6, 22, 22, 14, 18, 12, 55, 16, 22]
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f'activity_logs_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


# ---------------------------------------------------------------------------
# Downloadable Excel reports
# ---------------------------------------------------------------------------

@reports_bp.route('/leads/download', methods=['GET'])
@require_auth
def download_lead_report():
    user = request.current_user
    leads = get_user_visible_leads(user).all()
    buf = ReportService.lead_report_excel(leads)
    filename = f'lead_report_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


@reports_bp.route('/team/download', methods=['GET'])
@require_role('superadmin', 'sales_manager')
def download_team_report():
    user = request.current_user
    if user.role == 'superadmin':
        users = User.query.all()
    else:
        users = User.query.filter(
            (User.manager_id == user.id) | (User.id == user.id)
        ).all()

    buf = ReportService.team_report_excel(users)
    filename = f'team_report_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


@reports_bp.route('/management/download', methods=['GET'])
@require_role('superadmin', 'sales_manager')
def download_management_report():
    user = request.current_user
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    range_key = request.args.get('range')

    leads_q = _apply_created_date_filters(get_user_visible_leads(user), range_key, date_from, date_to)
    leads = leads_q.all()

    # Team / manager scope
    if user.role == 'superadmin':
        manager_users = User.query.filter_by(role='sales_manager', tenant_id=user.tenant_id, is_active=True).all()
        team_users = User.query.filter_by(role='team_member', tenant_id=user.tenant_id, is_active=True).all()
    else:
        manager_users = [user]
        team_users = User.query.filter_by(role='team_member', manager_id=user.id, tenant_id=user.tenant_id, is_active=True).all()

    # Helper maps
    leads_by_assignee = {}
    for lead in leads:
        if lead.assigned_to:
            leads_by_assignee.setdefault(lead.assigned_to, []).append(lead)

    def _row_stats(row_leads):
        total = len(row_leads)
        interested = sum(1 for l in row_leads if l.status == 'interested')
        sv_plan = sum(1 for l in row_leads if l.status == 'site_visit_planned')
        sv_done = sum(1 for l in row_leads if l.status == 'site_visit_done')
        negotiation = sum(1 for l in row_leads if l.status == 'negotiation')
        closed = sum(1 for l in row_leads if l.status == 'booking_done')
        conversion = round((closed / total) * 100, 2) if total else 0
        return {
            'total': total,
            'interested': interested,
            'site_visit_planned': sv_plan,
            'site_visit_done': sv_done,
            'negotiation': negotiation,
            'closures': closed,
            'conversion_pct': conversion,
        }

    wb = openpyxl.Workbook()
    ws_combined = wb.active
    ws_combined.title = 'Manager & Team Performance'
    ws_mgr_summary = wb.create_sheet('Manager Summary')

    header_fill = PatternFill('solid', fgColor='1E3A5F')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    manager_fill_colors = ['E8EAF6', 'E3F2FD', 'E8F5E9', 'FFF8E1', 'FCE4EC', 'EDE7F6', 'E0F7FA']
    total_fill = PatternFill('solid', fgColor='F1F5F9')

    def _write_sheet_header(ws, title, headers):
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=14, color='1E3A5F')
        ws['A2'] = f'Generated: {datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")}'
        if date_from or date_to or range_key:
            ws['A3'] = f'Range: {range_key or "custom"} {date_from or ""} {date_to or ""}'.strip()
        head_row = 5
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=head_row, column=col, value=h)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal='center', vertical='center')
        return head_row + 1

    perf_headers = ['Name', 'Role', 'All Leads', 'Interested', 'Site Visit Planned',
                    'Site Visit Done', 'Negotiation', 'Booking Done', 'Warm Leads', 'Hot Leads']

    # ── Sheet 1: Manager & Team combined (like the panel) ──────────────────
    r = _write_sheet_header(ws_combined, 'Manager & Team Performance', perf_headers)

    for color_idx, mgr in enumerate(manager_users):
        mgr_leads = list(leads_by_assignee.get(mgr.id, []))
        member_ids_for_mgr = [u.id for u in team_users if u.manager_id == mgr.id]
        for uid in member_ids_for_mgr:
            mgr_leads.extend(leads_by_assignee.get(uid, []))

        mgr_fill = PatternFill('solid', fgColor=manager_fill_colors[color_idx % len(manager_fill_colors)])
        mgr_font = Font(bold=True, size=11)

        # Manager row
        ms = _row_stats(list(leads_by_assignee.get(mgr.id, [])))
        warm_m = ms['interested'] + ms['site_visit_planned']
        hot_m = ms['site_visit_done'] + ms['negotiation']
        row_vals = [f'⭐ {mgr.name}', 'Sales Manager', ms['total'], ms['interested'],
                    ms['site_visit_planned'], ms['site_visit_done'], ms['negotiation'],
                    ms['closures'], warm_m, hot_m]
        for col, v in enumerate(row_vals, 1):
            c = ws_combined.cell(row=r, column=col, value=v)
            c.fill = mgr_fill
            c.font = mgr_font
            c.alignment = Alignment(vertical='center')
        r += 1

        # Team member rows
        for tm in [u for u in team_users if u.manager_id == mgr.id]:
            ts = _row_stats(leads_by_assignee.get(tm.id, []))
            warm_t = ts['interested'] + ts['site_visit_planned']
            hot_t = ts['site_visit_done'] + ts['negotiation']
            row_vals = [f'  ↳ {tm.name}', 'Team Member', ts['total'], ts['interested'],
                        ts['site_visit_planned'], ts['site_visit_done'], ts['negotiation'],
                        ts['closures'], warm_t, hot_t]
            for col, v in enumerate(row_vals, 1):
                ws_combined.cell(row=r, column=col, value=v)
            r += 1

        # Team Total row
        ts_agg = _row_stats(mgr_leads)
        warm_agg = ts_agg['interested'] + ts_agg['site_visit_planned']
        hot_agg = ts_agg['site_visit_done'] + ts_agg['negotiation']
        total_vals = ['∑ Team Total', '', ts_agg['total'], ts_agg['interested'],
                      ts_agg['site_visit_planned'], ts_agg['site_visit_done'], ts_agg['negotiation'],
                      ts_agg['closures'], warm_agg, hot_agg]
        for col, v in enumerate(total_vals, 1):
            c = ws_combined.cell(row=r, column=col, value=v)
            c.fill = total_fill
            c.font = Font(bold=True)
        r += 1
        r += 1  # blank row between groups

    # Unassigned members
    unassigned_tms = [u for u in team_users if not any(u.manager_id == mgr.id for mgr in manager_users)]
    if unassigned_tms:
        for tm in unassigned_tms:
            ts = _row_stats(leads_by_assignee.get(tm.id, []))
            warm_t = ts['interested'] + ts['site_visit_planned']
            hot_t = ts['site_visit_done'] + ts['negotiation']
            row_vals = [f'  ↳ {tm.name}', 'Unassigned', ts['total'], ts['interested'],
                        ts['site_visit_planned'], ts['site_visit_done'], ts['negotiation'],
                        ts['closures'], warm_t, hot_t]
            for col, v in enumerate(row_vals, 1):
                ws_combined.cell(row=r, column=col, value=v)
            r += 1

    # ── Sheet 2: Manager Summary ───────────────────────────────────────────
    sum_headers = ['Manager', 'Team Members', 'All Leads', 'Interested', 'Site Visit Planned',
                   'Site Visit Done', 'Negotiation', 'Booking Done', 'Warm Leads', 'Hot Leads', 'Warm Rate %', 'Hot Rate %']
    r2 = _write_sheet_header(ws_mgr_summary, 'Manager Summary', sum_headers)

    for color_idx, mgr in enumerate(manager_users):
        all_mgr_leads = list(leads_by_assignee.get(mgr.id, []))
        member_list = [u for u in team_users if u.manager_id == mgr.id]
        for uid in [u.id for u in member_list]:
            all_mgr_leads.extend(leads_by_assignee.get(uid, []))
        s = _row_stats(all_mgr_leads)
        warm = s['interested'] + s['site_visit_planned']
        hot = s['site_visit_done'] + s['negotiation']
        warm_rate = round((warm / s['total']) * 100, 1) if s['total'] else 0
        hot_rate = round((hot / s['total']) * 100, 1) if s['total'] else 0
        vals = [mgr.name, len(member_list), s['total'], s['interested'],
                s['site_visit_planned'], s['site_visit_done'], s['negotiation'],
                s['closures'], warm, hot, warm_rate, hot_rate]
        mgr_fill = PatternFill('solid', fgColor=manager_fill_colors[color_idx % len(manager_fill_colors)])
        for col, v in enumerate(vals, 1):
            c = ws_mgr_summary.cell(row=r2, column=col, value=v)
            c.fill = mgr_fill
        r2 += 1

    # Column widths
    for ws in [ws_combined, ws_mgr_summary]:
        col_widths = [28, 16, 12, 12, 20, 18, 14, 14, 13, 11, 12, 12]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = 'A6'
        ws.auto_filter.ref = f'A5:{get_column_letter(len(perf_headers))}5'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'management_report_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


@reports_bp.route('/activity/download', methods=['GET'])
@require_role('superadmin')
def download_activity_report():
    days = request.args.get('days', 30, type=int)
    buf = ReportService.activity_report_excel(days=days)
    filename = f'activity_report_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )
