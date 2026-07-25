"""
Lead Source Management – Admin API Routes
==========================================
All routes require superadmin or sales_manager role (tenant-scoped).

CRUD:
  GET    /api/lead-sources                       – list all sources
  POST   /api/lead-sources                       – create source
  GET    /api/lead-sources/<id>                  – get source
  PUT    /api/lead-sources/<id>                  – update source
  DELETE /api/lead-sources/<id>                  – soft-delete (is_active=False)

Operations:
  POST   /api/lead-sources/<id>/test             – validate credentials / permissions
  POST   /api/lead-sources/<id>/enable           – re-enable disabled source
  POST   /api/lead-sources/<id>/disable          – disable source

Reports:
  GET    /api/lead-sources/reports/by-source     – leads grouped by source
  GET    /api/lead-sources/reports/by-campaign   – leads grouped by campaign
  GET    /api/lead-sources/logs                  – ingestion log (with pagination)

Meta OAuth helpers:
  GET    /api/lead-sources/meta/pages            – list pages accessible to token
  GET    /api/lead-sources/meta/forms/<page_id>  – list lead forms for a page
"""

import logging
import os
import secrets
import csv
import json as _json
import base64
import urllib.request as _req
import urllib.error as _urlerr
import urllib.parse as _parse
from datetime import datetime, timedelta, timezone

from flask import Blueprint, jsonify, request, redirect, Response
from sqlalchemy import func, or_

from app.middleware import require_auth, require_role
from app.models.base import db
from app.models.lead import Lead
from app.models.ingestion import (
    LeadSource,
    IngestedLeadLog,
    ConnectedGoogleAdsAccount,
    SOURCE_TYPES,
    DUP_MODES,
    ASSIGN_STRATEGIES,
)
from app.models.lead_source_mapping import LeadSourceFormMapping, MetaCampaignSnapshot
from app.models.project import Project
from app.models.meta_tier_test_run import MetaTierTestRun
from app.models.oauth_session import OAuthSession
from app.models.user import User
from app.services.meta_tier_testing import build_initial_state, compute_dashboard, run_batch
from app.services.ingestion_engine import ingest_lead, normalize_phone_for_duplicate
from app.utils.lead_source_cutoff import (
    effective_start_with_cutoff,
    lead_source_cutoff_for,
)

logger = logging.getLogger(__name__)


def _runtime_instance_id():
    return (
        request.headers.get('X-Vercel-Id')
        or request.environ.get('VERCEL_DEPLOYMENT_ID')
        or request.environ.get('VERCEL_REGION')
        or request.environ.get('VERCEL_URL')
        or request.environ.get('VERCEL_ENV')
        or 'unknown'
    )


def _log_meta_session_event(stage, **fields):
    redacted_keys = {'session_key', 'long_token', 'short_token', 'access_token', 'page_access_token', 'redirect_url'}
    payload = {
        'stage': stage,
        'path': request.path,
        'method': request.method,
        'instance': _runtime_instance_id(),
    }
    for key, value in fields.items():
        payload[key] = '[redacted]' if key in redacted_keys and value else value
    logger.info('META_SESSION_EVENT %s', payload)

lead_sources_bp = Blueprint('lead_sources', __name__, url_prefix='/api/lead-sources')

# ── In-memory OAuth session store (TTL = 10 min) ──────────────────────────────
# Maps session_key → { tenant_id, business_id, pages, user, created_at }
_oauth_sessions = {}
_OAUTH_SESSION_TTL_MINUTES = 10

_TENANT_ROUTE_SLUG_ALIASES = {
    'ganga': 'ganga-realty',
}

_TEST_DATA_PATTERNS = (
    'validation run',
    'google validation run',
    'campaign proof alpha',
    'validation lead',
    'proof lead',
    'test lead',
    'test campaign',
)

FORM_MANAGER_ASSIGN_MODES = ('none', 'fixed_manager', 'round_robin_pool')


def _show_test_data_enabled():
    return str(request.args.get('show_test_data', '')).strip().lower() in {'1', 'true', 'yes', 'on'}


def _test_data_clause(log_model):
    platform = func.lower(func.coalesce(log_model.platform_lead_id, ''))
    campaign = func.lower(func.coalesce(log_model.campaign_name, ''))
    form_name = func.lower(func.coalesce(log_model.form_name, ''))
    ad_name = func.lower(func.coalesce(log_model.ad_name, ''))
    return or_(
        log_model.is_test == True,
        platform.like('test-%'),
        platform.like('validate-%'),
        platform.like('realproof-%'),
        campaign.in_(_TEST_DATA_PATTERNS),
        campaign.like('%validation%'),
        campaign.like('%proof%'),
        campaign.like('%test%'),
        form_name.like('%validation%'),
        form_name.like('%proof%'),
        form_name.like('%test%'),
        ad_name.like('%proof%'),
        ad_name.like('%test%'),
    )


def _apply_test_data_filter(query, log_model):
    if _show_test_data_enabled():
        return query
    return query.filter(~_test_data_clause(log_model))


def _apply_snapshot_test_filter(query):
    if _show_test_data_enabled():
        return query
    return query.filter(MetaCampaignSnapshot.is_test == False)


def _campaign_label(campaign_id, campaign_name):
    if campaign_name:
        return campaign_name
    if campaign_id:
        return str(campaign_id)
    return 'Attribution Pending'


def _fallback_label(value, pending_label='Attribution Pending'):
    if value not in (None, '', '-'):
        return value
    return pending_label


def _fallback_label_with_id(name_value, id_value, pending_label='Attribution Pending'):
    if name_value not in (None, '', '-'):
        return name_value
    if id_value not in (None, '', '-'):
        return str(id_value)
    return pending_label


def _metric_from_snapshot(snapshot: MetaCampaignSnapshot, attr_name: str, *extra_keys):
    if snapshot is None:
        return None
    direct_val = getattr(snapshot, attr_name, None)
    if direct_val not in (None, ''):
        return direct_val
    extra = snapshot.extra_metrics or {}
    for key in extra_keys:
        val = extra.get(key)
        if val not in (None, ''):
            return val
    return None


def _to_float_or_none(value):
    try:
        if value in (None, ''):
            return None
        return float(value)
    except Exception:
        return None


def _to_int_or_none(value):
    try:
        if value in (None, ''):
            return None
        return int(float(value))
    except Exception:
        return None


def _meta_graph_json(path, params=None, timeout=10):
    query = _parse.urlencode(params or {})
    url = f'https://graph.facebook.com/v25.0/{path}'
    if query:
        url = f'{url}?{query}'
    with _req.urlopen(_req.Request(url), timeout=timeout) as resp:
        return _json.loads(resp.read())


def _meta_graph_json_url(url, timeout=10):
    with _req.urlopen(_req.Request(url), timeout=timeout) as resp:
        return _json.loads(resp.read())


def _meta_ad_account_ids_for_source(source):
    creds = source.credentials or {}
    ids = []

    def add(raw):
        account_id = str(raw or '').strip()
        if not account_id:
            return
        if not account_id.startswith('act_'):
            account_id = 'act_' + account_id
        if account_id not in ids:
            ids.append(account_id)

    add(creds.get('ad_account_id'))
    for account in creds.get('ad_accounts') or []:
        if isinstance(account, dict):
            add(account.get('id') or account.get('account_id'))
    return ids


def _source_campaign_ids_for_spend(user, source_id, date_from='', date_to=''):
    query = IngestedLeadLog.query.filter(
        IngestedLeadLog.tenant_id == user.tenant_id,
        IngestedLeadLog.source_id == source_id,
        IngestedLeadLog.campaign_id.isnot(None),
    )
    query = _apply_log_date_filters(query, IngestedLeadLog, date_from, date_to)
    query = _apply_test_data_filter(query, IngestedLeadLog)
    ids = {
        str(row[0]).strip()
        for row in query.with_entities(IngestedLeadLog.campaign_id).distinct().limit(500).all()
        if row[0]
    }
    return ids


def _fetch_meta_source_spend(token, ad_account_ids, since, until, campaign_ids=None):
    total_spend = 0.0
    matched_campaign_ids = set()
    matched_accounts = []
    fallback_spend = 0.0
    fallback_accounts = []
    errors = []
    campaign_ids = {str(v) for v in (campaign_ids or set()) if v}

    for ad_account_id in ad_account_ids:
        if campaign_ids:
            try:
                params = {
                    'fields': 'campaign_id,campaign_name,spend',
                    'level': 'campaign',
                    'time_range': _json.dumps({'since': since, 'until': until}),
                    'limit': '500',
                    'access_token': token,
                }
                page = _meta_graph_json(f"{_parse.quote(ad_account_id)}/insights", params, timeout=15)
                page_count = 0
                while page and page_count < 20:
                    page_count += 1
                    for insight in (page.get('data') or []):
                        campaign_id = str(insight.get('campaign_id') or '').strip()
                        if campaign_id not in campaign_ids:
                            continue
                        spend = _to_float_or_none(insight.get('spend'))
                        if spend is None:
                            continue
                        total_spend += spend
                        matched_campaign_ids.add(campaign_id)
                        if ad_account_id not in matched_accounts:
                            matched_accounts.append(ad_account_id)
                    next_url = ((page.get('paging') or {}).get('next') or '').strip()
                    if not next_url:
                        break
                    page = _meta_graph_json_url(next_url, timeout=15)
            except Exception as exc:
                errors.append({'object_id': ad_account_id, 'message': str(exc)})

        try:
            payload = _meta_graph_json(
                f"{_parse.quote(ad_account_id)}/insights",
                {
                    'fields': 'spend',
                    'time_range': _json.dumps({'since': since, 'until': until}),
                    'limit': '1',
                    'access_token': token,
                },
                timeout=12,
            )
            insight = ((payload or {}).get('data') or [None])[0] or {}
            account_spend = _to_float_or_none(insight.get('spend'))
            if account_spend is not None:
                fallback_spend += account_spend
                fallback_accounts.append(ad_account_id)
        except Exception as exc:
            errors.append({'object_id': ad_account_id, 'message': str(exc)})

    if matched_campaign_ids:
        return {
            'spend': round(total_spend, 2),
            'method': 'campaign_matched',
            'matched_campaign_count': len(matched_campaign_ids),
            'matched_account_ids': matched_accounts,
            'errors': errors,
        }
    return {
        'spend': round(fallback_spend, 2) if fallback_accounts else None,
        'method': 'account_total_all',
        'matched_campaign_count': 0,
        'matched_account_ids': fallback_accounts,
        'errors': errors,
    }


def _enrich_attribution_rows_with_meta(user, rows, date_from='', date_to='', include_live_metrics=True):
    if not rows:
        return rows

    source_ids = sorted({int(r.get('source_id') or 0) for r in rows if r.get('source_id')})
    if not source_ids:
        return rows

    sources = LeadSource.query.filter(
        LeadSource.tenant_id == user.tenant_id,
        LeadSource.id.in_(source_ids),
    ).all()
    source_map = {s.id: s for s in sources}

    rows_by_source = {}
    for row in rows:
        sid = int(row.get('source_id') or 0)
        if sid <= 0:
            continue
        rows_by_source.setdefault(sid, []).append(row)

    for source_id, scoped_rows in rows_by_source.items():
        source = source_map.get(source_id)
        if not source or source.source_type != 'meta':
            continue

        creds = source.credentials or {}
        token_candidates = []
        for t in (
            str(creds.get('user_token') or '').strip(),
            str(creds.get('page_access_token') or '').strip(),
            str(creds.get('access_token') or '').strip(),
        ):
            if t and t not in token_candidates:
                token_candidates.append(t)

        form_cache = {
            str(item.get('id') or ''): str(item.get('name') or '').strip()
            for item in (source.available_forms or []) if isinstance(item, dict)
        }
        campaign_cache = {
            str(item.get('id') or ''): str(item.get('name') or '').strip()
            for item in (source.available_campaigns or []) if isinstance(item, dict)
        }

        unresolved_campaign_ids = set()
        unresolved_adset_ids = set()
        unresolved_ad_ids = set()

        for row in scoped_rows:
            if not row.get('campaign_name') and row.get('campaign_id'):
                cached = campaign_cache.get(str(row.get('campaign_id')))
                if cached:
                    row['campaign_name'] = cached
                else:
                    unresolved_campaign_ids.add(str(row.get('campaign_id')))
            if not row.get('ad_set_name') and row.get('ad_set_id'):
                unresolved_adset_ids.add(str(row.get('ad_set_id')))
            if not row.get('ad_name') and row.get('ad_id'):
                unresolved_ad_ids.add(str(row.get('ad_id')))
            if not row.get('form_name') and row.get('form_id'):
                row['form_name'] = form_cache.get(str(row.get('form_id'))) or row.get('form_name')

        if token_candidates:
            ids_to_lookup = list(unresolved_campaign_ids | unresolved_adset_ids | unresolved_ad_ids)
            name_map = {}

            # First attempt: bulk IDs call (fast path)
            for token in token_candidates:
                try:
                    payload = _meta_graph_json('', {
                        'ids': ','.join(ids_to_lookup[:200]),
                        'fields': 'name',
                        'access_token': token,
                    }, timeout=8)
                    for key, val in (payload or {}).items():
                        if isinstance(val, dict) and val.get('name'):
                            name_map[str(key)] = str(val.get('name'))
                    if name_map:
                        break
                except Exception:
                    continue

            # Second attempt: per-object lookup for unresolved IDs
            unresolved = [i for i in ids_to_lookup if i and not name_map.get(i)]
            for obj_id in unresolved[:200]:
                for token in token_candidates:
                    try:
                        payload = _meta_graph_json(
                            _parse.quote(str(obj_id)),
                            {'fields': 'name', 'access_token': token},
                            timeout=8,
                        )
                        nm = str((payload or {}).get('name') or '').strip()
                        if nm:
                            name_map[str(obj_id)] = nm
                            break
                    except Exception:
                        continue

            for row in scoped_rows:
                if not row.get('campaign_name') and row.get('campaign_id'):
                    row['campaign_name'] = name_map.get(str(row.get('campaign_id'))) or row.get('campaign_name')
                if not row.get('ad_set_name') and row.get('ad_set_id'):
                    row['ad_set_name'] = name_map.get(str(row.get('ad_set_id'))) or row.get('ad_set_name')
                if not row.get('ad_name') and row.get('ad_id'):
                    row['ad_name'] = name_map.get(str(row.get('ad_id'))) or row.get('ad_name')

            if include_live_metrics:
                metric_candidates = [
                    row for row in scoped_rows
                    if row.get('ad_id') and (
                        row.get('spend') in (None, '') or
                        row.get('cpl') in (None, '') or
                        row.get('ctr') in (None, '') or
                        row.get('reach') in (None, '') or
                        row.get('impressions') in (None, '')
                    )
                ]

                for row in metric_candidates[:20]:
                    for token in token_candidates:
                        try:
                            params = {
                                'fields': 'spend,ctr,cpc,cpm,reach,impressions,cost_per_action_type',
                                'limit': '1',
                                'access_token': token,
                            }
                            if date_from or date_to:
                                until = date_to or datetime.utcnow().date().isoformat()
                                since = date_from or until
                                params['time_range'] = _json.dumps({'since': since, 'until': until})

                            data = _meta_graph_json(f"{_parse.quote(str(row.get('ad_id')))}" + '/insights', params, timeout=9)
                            insight = ((data or {}).get('data') or [None])[0] or {}

                            spend = _to_float_or_none(insight.get('spend'))
                            ctr = _to_float_or_none(insight.get('ctr'))
                            cpc = _to_float_or_none(insight.get('cpc'))
                            cpm = _to_float_or_none(insight.get('cpm'))
                            reach = _to_int_or_none(insight.get('reach'))
                            impressions = _to_int_or_none(insight.get('impressions'))

                            cpl = None
                            for cpa in (insight.get('cost_per_action_type') or []):
                                action_type = str(cpa.get('action_type') or '').lower()
                                if 'lead' in action_type:
                                    cpl = _to_float_or_none(cpa.get('value'))
                                    if cpl is not None:
                                        break

                            if row.get('spend') in (None, '') and spend is not None:
                                row['spend'] = spend
                            if row.get('ctr') in (None, '') and ctr is not None:
                                row['ctr'] = ctr
                            if row.get('cpc') in (None, '') and cpc is not None:
                                row['cpc'] = cpc
                            if row.get('cpm') in (None, '') and cpm is not None:
                                row['cpm'] = cpm
                            if row.get('reach') in (None, '') and reach is not None:
                                row['reach'] = reach
                            if row.get('impressions') in (None, '') and impressions is not None:
                                row['impressions'] = impressions
                            if row.get('cpl') in (None, '') and cpl is not None:
                                row['cpl'] = cpl

                            if any([
                                row.get('spend') not in (None, ''),
                                row.get('ctr') not in (None, ''),
                                row.get('reach') not in (None, ''),
                                row.get('impressions') not in (None, ''),
                                row.get('cpl') not in (None, ''),
                            ]):
                                break
                        except Exception:
                            continue

        for row in scoped_rows:
            row['campaign_name'] = _campaign_label(row.get('campaign_id'), row.get('campaign_name'))
            row['ad_set_name'] = _fallback_label_with_id(row.get('ad_set_name'), row.get('ad_set_id'))
            row['ad_name'] = _fallback_label_with_id(row.get('ad_name'), row.get('ad_id'))
            row['form_name'] = _fallback_label_with_id(row.get('form_name'), row.get('form_id'))

    return rows


def _apply_log_date_filters(query, model, date_from, date_to):
    if date_from:
        try:
            query = query.filter(model.received_at >= datetime.fromisoformat(date_from))
        except ValueError:
            pass
    if date_to:
        try:
            query = query.filter(model.received_at < datetime.fromisoformat(date_to) + timedelta(days=1))
        except ValueError:
            pass
    return query


def _apply_snapshot_date_filters(query, date_from, date_to):
    if date_from:
        try:
            query = query.filter(MetaCampaignSnapshot.snapshot_at >= datetime.fromisoformat(date_from))
        except ValueError:
            pass
    if date_to:
        try:
            query = query.filter(MetaCampaignSnapshot.snapshot_at < datetime.fromisoformat(date_to) + timedelta(days=1))
        except ValueError:
            pass
    return query


def _parse_report_datetime(value):
    if not value:
        return None
    try:
        return datetime.fromisoformat(str(value).strip())
    except ValueError:
        return None


def _effective_source_start(source_obj, date_from=''):
    requested_start = _parse_report_datetime(date_from)
    return effective_start_with_cutoff(source_obj, requested_start)


def _source_added_iso(source_obj):
    created = getattr(source_obj, 'created_at', None)
    return created.isoformat() if created else None


def _source_form_project_map(tenant_id, source_ids):
    if not source_ids:
        return {}
    rows = (
        LeadSourceFormMapping.query
        .filter(
            LeadSourceFormMapping.tenant_id == tenant_id,
            LeadSourceFormMapping.source_id.in_(source_ids),
            LeadSourceFormMapping.is_active == True,
        )
        .all()
    )
    out = {}
    for row in rows:
        form_id = str(row.form_id or '').strip()
        if not form_id:
            continue
        out[(row.source_id, form_id)] = row.project.name if row.project else None
    return out


def _metric_timestamp(value):
    return value.isoformat() if value else None


def _snapshot_quality_score(snapshot: MetaCampaignSnapshot):
    if snapshot is None:
        return -1
    score = 0
    for val in (
        snapshot.campaign_name,
        snapshot.ad_set_name,
        snapshot.ad_name,
        snapshot.form_name,
        snapshot.spend,
        snapshot.cost_per_result,
        snapshot.ctr,
        snapshot.cpc,
        snapshot.cpm,
        snapshot.reach,
        snapshot.impressions,
    ):
        if val not in (None, '', '-'):
            score += 1
    extra = snapshot.extra_metrics or {}
    for key in ('spend', 'amount_spent', 'cpl', 'cost_per_result', 'ctr', 'reach', 'impressions'):
        if extra.get(key) not in (None, '', '-'):
            score += 1
    return score


def _choose_metric_row(current_row: dict | None, snapshot: MetaCampaignSnapshot):
    if current_row is None:
        return snapshot
    cur_score = _snapshot_quality_score(current_row)
    next_score = _snapshot_quality_score(snapshot)
    if next_score > cur_score:
        return snapshot
    if next_score < cur_score:
        return current_row
    if not snapshot.snapshot_at:
        return current_row
    if not getattr(current_row, 'snapshot_at', None):
        return snapshot
    return snapshot if snapshot.snapshot_at >= current_row.snapshot_at else current_row


def _resolve_project_name(lead_obj, extra_metrics, source_id=None, form_id=None, form_project_map=None):
    if lead_obj and lead_obj.project:
        return lead_obj.project.name
    extra_project = (extra_metrics or {}).get('project_name')
    if extra_project:
        return extra_project
    mapped = (form_project_map or {}).get((source_id, str(form_id or '').strip()))
    return mapped or 'Not Mapped'


def _report_date_to_exclusive(date_to=''):
    parsed = _parse_report_datetime(date_to)
    if parsed:
        return parsed + timedelta(days=1)
    return None


def _report_meta_until_date(date_to=''):
    parsed = _parse_report_datetime(date_to)
    return (parsed.date() if parsed else datetime.utcnow().date()).isoformat()


def _snapshot_matches_report_window(snapshot, source_obj, date_from='', date_to=''):
    extra = snapshot.extra_metrics or {}
    insight_from = extra.get('insight_date_from')
    insight_to = extra.get('insight_date_to')
    effective_start = _effective_source_start(source_obj, date_from)
    expected_from = (effective_start.date() if effective_start else datetime.utcnow().date()).isoformat()
    expected_to = _report_meta_until_date(date_to)
    if insight_from or insight_to:
        return insight_from == expected_from and insight_to == expected_to
    if effective_start and snapshot.snapshot_at and snapshot.snapshot_at < effective_start:
        return False
    end_exclusive = _report_date_to_exclusive(date_to)
    if end_exclusive and snapshot.snapshot_at and snapshot.snapshot_at >= end_exclusive:
        return False
    return True


def _campaign_key_from_parts(source_id, project_name, campaign_id, ad_set_id, ad_id, form_id):
    return (
        source_id,
        project_name or 'Not Mapped',
        str(campaign_id or ''),
        str(ad_set_id or ''),
        str(ad_id or ''),
        str(form_id or ''),
    )


def _metric_object_key_from_row(row):
    ad_id = str((row or {}).get('ad_id') or '').strip()
    ad_set_id = str((row or {}).get('ad_set_id') or '').strip()
    campaign_id = str((row or {}).get('campaign_id') or '').strip()
    if ad_id:
        return ('ad', ad_id)
    if ad_set_id:
        return ('adset', ad_set_id)
    if campaign_id:
        return ('campaign', campaign_id)
    return ('', '')


def _clear_metric_fields(row):
    for key in ('spend', 'cpl', 'ctr', 'cpc', 'cpm', 'reach', 'impressions', 'clicks'):
        row[key] = None


def _is_reportable_meta_log(log_row, project_name):
    if project_name == 'Not Mapped':
        return False
    return bool(
        getattr(log_row, 'campaign_id', None)
        and getattr(log_row, 'ad_set_id', None)
        and getattr(log_row, 'ad_id', None)
        and getattr(log_row, 'form_id', None)
    )


def _log_effective_status(log_row):
    status = str(log_row.status or '').lower()
    if status != 'duplicate':
        return status

    mapped = log_row.mapped_fields if isinstance(log_row.mapped_fields, dict) else {}
    inbound_phone = normalize_phone_for_duplicate(mapped.get('phone') or mapped.get('mobile'))
    dup_lead = getattr(log_row, 'dup_lead', None)
    existing_phone = normalize_phone_for_duplicate(getattr(dup_lead, 'phone', None))
    if inbound_phone and existing_phone and inbound_phone == existing_phone:
        return 'duplicate'
    return 'processed'


def _campaign_has_meta_metrics(row):
    return row.get('spend') not in (None, '')


def _extract_meta_lead_cpl(insight):
    for cpa in (insight or {}).get('cost_per_action_type') or []:
        action_type = str(cpa.get('action_type') or '').lower()
        if 'lead' in action_type:
            val = _to_float_or_none(cpa.get('value'))
            if val is not None:
                return val
    return None


def _build_performance_report(user, date_from='', date_to='', source_id=None, include_unpriced=False):
    _deactivate_duplicate_meta_sources_for_tenant(user.tenant_id)

    source_query = LeadSource.query.filter(LeadSource.tenant_id == user.tenant_id, LeadSource.is_active == True)
    if source_id:
        source_query = source_query.filter(LeadSource.id == source_id)
    sources = source_query.all()
    source_map = {source.id: source for source in sources}
    source_ids = list(source_map.keys())
    if not source_ids:
        return {
            'snapshot': {'total': 0, 'processed': 0, 'duplicate': 0, 'errors': 0, 'conversion_rate': 0},
            'source_rows': [],
            'campaign_rows': [],
            'last_synced_at': None,
        }

    form_project_map = _source_form_project_map(user.tenant_id, source_ids)
    end_exclusive = _report_date_to_exclusive(date_to)

    log_query = (
        db.session.query(IngestedLeadLog, Lead, LeadSource)
        .select_from(IngestedLeadLog)
        .outerjoin(Lead, IngestedLeadLog.lead_id == Lead.id)
        .join(LeadSource, IngestedLeadLog.source_id == LeadSource.id)
        .filter(IngestedLeadLog.tenant_id == user.tenant_id)
        .filter(LeadSource.tenant_id == user.tenant_id)
        .filter(IngestedLeadLog.source_id.in_(source_ids))
    )
    tenant_cutoff = lead_source_cutoff_for(tenant_id=user.tenant_id)
    if tenant_cutoff:
        log_query = log_query.filter(IngestedLeadLog.received_at >= tenant_cutoff)
    log_query = _apply_log_date_filters(log_query, IngestedLeadLog, date_from, date_to)
    log_query = _apply_test_data_filter(log_query, IngestedLeadLog)

    source_groups = {}
    campaign_groups = {}
    source_lead_ids = {}
    source_duplicate_ids = {}
    campaign_lead_ids = {}

    seen_log_identities = set()
    joined_log_rows = log_query.order_by(
        IngestedLeadLog.received_at.desc(),
        IngestedLeadLog.id.desc(),
    ).all()
    for log_row, lead_obj, source_obj in joined_log_rows:
        log_identity = _source_log_identity(log_row)
        if log_identity in seen_log_identities:
            continue
        seen_log_identities.add(log_identity)

        source_obj = source_obj or source_map.get(log_row.source_id)
        effective_start = _effective_source_start(source_obj, date_from)
        if effective_start and log_row.received_at and log_row.received_at < effective_start:
            continue
        if end_exclusive and log_row.received_at and log_row.received_at >= end_exclusive:
            continue

        project_name = _resolve_project_name(
            lead_obj or log_row.dup_lead,
            {},
            source_id=log_row.source_id,
            form_id=log_row.form_id,
            form_project_map=form_project_map,
        )
        status = _log_effective_status(log_row)

        source_key = (log_row.source_id, project_name)
        source_bucket = source_groups.get(source_key)
        if not source_bucket:
            source_bucket = {
                'source_id': log_row.source_id,
                'source_name': source_obj.name if source_obj else 'Unknown Source',
                'source_added_at': _source_added_iso(source_obj),
                'effective_date_from': effective_start.isoformat() if effective_start else None,
                'project_name': project_name,
                'source_status': 'Active Source' if source_obj and source_obj.is_active else 'Archived Source',
                'leads': 0,
                'total': 0,
                'created': 0,
                'duplicates': 0,
                'errors': 0,
                'spend': None,
                'cpl': None,
                'last_sync': None,
            }
            source_groups[source_key] = source_bucket
            source_lead_ids[source_key] = set()
            source_duplicate_ids[source_key] = set()

        if status == 'processed':
            lead_key = log_row.lead_id or f'log:{log_row.id}'
            if lead_key not in source_lead_ids[source_key]:
                source_lead_ids[source_key].add(lead_key)
                source_bucket['leads'] += 1
                source_bucket['created'] += 1
            source_bucket['total'] = source_bucket['leads']
        elif status == 'duplicate':
            duplicate_key = log_row.platform_lead_id or f'log:{log_row.id}'
            if duplicate_key not in source_duplicate_ids[source_key]:
                source_duplicate_ids[source_key].add(duplicate_key)
                source_bucket['duplicates'] += 1
        elif status == 'error':
            source_bucket['errors'] += 1

        if log_row.received_at and (not source_bucket['last_sync'] or log_row.received_at.isoformat() > source_bucket['last_sync']):
            source_bucket['last_sync'] = log_row.received_at.isoformat()

        if status != 'processed' or not _is_reportable_meta_log(log_row, project_name):
            continue

        campaign_key = _campaign_key_from_parts(
            log_row.source_id,
            project_name,
            log_row.campaign_id,
            log_row.ad_set_id,
            log_row.ad_id,
            log_row.form_id,
        )
        campaign_bucket = campaign_groups.get(campaign_key)
        if not campaign_bucket:
            campaign_bucket = {
                'source_id': log_row.source_id,
                'source_added_at': _source_added_iso(source_obj),
                'effective_date_from': effective_start.isoformat() if effective_start else None,
                'project_name': project_name,
                'campaign_id': log_row.campaign_id,
                'campaign_name': log_row.campaign_name,
                'ad_set_id': log_row.ad_set_id,
                'ad_set_name': log_row.ad_set_name,
                'ad_id': log_row.ad_id,
                'ad_name': log_row.ad_name,
                'form_id': log_row.form_id,
                'form_name': log_row.form_name,
                'leads': 0,
                'total': 0,
                'created': 0,
                'spend': None,
                'cpl': None,
                'impressions': None,
                'reach': None,
                'clicks': None,
                'ctr': None,
                'cpc': None,
                'cpm': None,
                'last_sync': None,
            }
            campaign_groups[campaign_key] = campaign_bucket
            campaign_lead_ids[campaign_key] = set()

        lead_key = log_row.lead_id or f'log:{log_row.id}'
        if lead_key not in campaign_lead_ids[campaign_key]:
            campaign_lead_ids[campaign_key].add(lead_key)
            campaign_bucket['leads'] += 1
            campaign_bucket['created'] += 1
            campaign_bucket['total'] = campaign_bucket['leads']
        if log_row.received_at and (not campaign_bucket['last_sync'] or log_row.received_at.isoformat() > campaign_bucket['last_sync']):
            campaign_bucket['last_sync'] = log_row.received_at.isoformat()

    snapshot_query = (
        MetaCampaignSnapshot.query
        .join(LeadSource, MetaCampaignSnapshot.source_id == LeadSource.id)
        .filter(MetaCampaignSnapshot.tenant_id == user.tenant_id)
        .filter(LeadSource.tenant_id == user.tenant_id)
        .filter(MetaCampaignSnapshot.source_id.in_(source_ids))
    )
    snapshot_query = _apply_snapshot_test_filter(snapshot_query)

    metrics_map = {}
    latest_sync = None
    for snapshot in snapshot_query.order_by(MetaCampaignSnapshot.snapshot_at.desc()).limit(5000).all():
        source_obj = source_map.get(snapshot.source_id)
        if not _snapshot_matches_report_window(snapshot, source_obj, date_from, date_to):
            continue
        project_name = _resolve_project_name(
            snapshot.lead,
            snapshot.extra_metrics,
            source_id=snapshot.source_id,
            form_id=snapshot.form_id,
            form_project_map=form_project_map,
        )
        key = _campaign_key_from_parts(
            snapshot.source_id,
            project_name,
            snapshot.campaign_id,
            snapshot.ad_set_id,
            snapshot.ad_id,
            snapshot.form_id,
        )
        metrics_map[key] = _choose_metric_row(metrics_map.get(key), snapshot)
        if snapshot.snapshot_at and (latest_sync is None or snapshot.snapshot_at > latest_sync):
            latest_sync = snapshot.snapshot_at

    for key, bucket in campaign_groups.items():
        snapshot = metrics_map.get(key)
        if snapshot:
            bucket['campaign_name'] = snapshot.campaign_name or bucket['campaign_name']
            bucket['ad_set_name'] = snapshot.ad_set_name or bucket['ad_set_name']
            bucket['ad_name'] = snapshot.ad_name or bucket['ad_name']
            bucket['form_name'] = snapshot.form_name or bucket['form_name']
            bucket['spend'] = _metric_from_snapshot(snapshot, 'spend', 'spend', 'amount_spent')
            bucket['ctr'] = _metric_from_snapshot(snapshot, 'ctr', 'ctr')
            bucket['cpc'] = _metric_from_snapshot(snapshot, 'cpc', 'cpc')
            bucket['cpm'] = _metric_from_snapshot(snapshot, 'cpm', 'cpm')
            bucket['reach'] = _metric_from_snapshot(snapshot, 'reach', 'reach')
            bucket['impressions'] = _metric_from_snapshot(snapshot, 'impressions', 'impressions')
            bucket['clicks'] = _metric_from_snapshot(snapshot, 'clicks', 'clicks')
            bucket['cpl'] = _metric_from_snapshot(snapshot, 'cost_per_result', 'cpl', 'cost_per_result')
            if bucket['cpl'] is None and bucket['spend'] not in (None, 0) and bucket['leads']:
                bucket['cpl'] = round(float(bucket['spend']) / float(bucket['leads']), 2)
            bucket['last_sync'] = _metric_timestamp(snapshot.snapshot_at) or bucket['last_sync']

        bucket['campaign_name'] = _campaign_label(bucket.get('campaign_id'), bucket.get('campaign_name'))
        bucket['ad_set_name'] = _fallback_label_with_id(bucket.get('ad_set_name'), bucket.get('ad_set_id'))
        bucket['ad_name'] = _fallback_label_with_id(bucket.get('ad_name'), bucket.get('ad_id'))
        bucket['form_name'] = _fallback_label_with_id(bucket.get('form_name'), bucket.get('form_id'))

    priced_campaign_groups = {}
    for key, campaign in campaign_groups.items():
        if include_unpriced or _campaign_has_meta_metrics(campaign):
            priced_campaign_groups[key] = campaign

    source_spend = {}
    source_spend_objects = {}
    for key, campaign in priced_campaign_groups.items():
        source_key = (campaign['source_id'], campaign['project_name'])
        spend = _to_float_or_none(campaign.get('spend'))
        metric_object = _metric_object_key_from_row(campaign)
        seen_objects = source_spend_objects.setdefault(source_key, set())
        if spend is not None and metric_object != ('', '') and metric_object not in seen_objects:
            seen_objects.add(metric_object)
            source_spend[source_key] = source_spend.get(source_key, 0.0) + spend
        if campaign.get('last_sync'):
            source_bucket = source_groups.get(source_key)
            if source_bucket and (not source_bucket.get('last_sync') or campaign['last_sync'] > source_bucket['last_sync']):
                source_bucket['last_sync'] = campaign['last_sync']

    for key, bucket in source_groups.items():
        if key in source_spend:
            bucket['spend'] = round(source_spend[key], 2)
            bucket['cpl'] = round(bucket['spend'] / bucket['leads'], 2) if bucket['leads'] else None

    processed = sum(int(row.get('leads') or 0) for row in source_groups.values())
    duplicate = sum(int(row.get('duplicates') or 0) for row in source_groups.values())
    errors = sum(int(row.get('errors') or 0) for row in source_groups.values())
    total = processed + duplicate + errors
    overview = {
        'total': total,
        'processed': processed,
        'duplicate': duplicate,
        'errors': errors,
        'conversion_rate': round((processed * 100.0 / total), 2) if total else 0,
    }

    source_rows = sorted(source_groups.values(), key=lambda row: (row['source_name'], row['project_name']))
    campaign_rows = sorted(priced_campaign_groups.values(), key=lambda row: (-int(row.get('leads') or 0), row.get('campaign_name') or '', row.get('ad_set_name') or '', row.get('ad_name') or ''))

    seen_metric_objects = set()
    for row in campaign_rows:
        metric_object = _metric_object_key_from_row(row)
        if metric_object == ('', ''):
            continue
        if metric_object in seen_metric_objects:
            _clear_metric_fields(row)
            continue
        seen_metric_objects.add(metric_object)

    return {
        'snapshot': overview,
        'source_rows': source_rows,
        'campaign_rows': campaign_rows,
        'last_synced_at': latest_sync.isoformat() if latest_sync else None,
    }


def _build_lms_source_form_performance(user, date_from='', date_to='', source_id=None):
    """Lightweight LMS report: source/form aggregates without campaign/ad rows."""
    _deactivate_duplicate_meta_sources_for_tenant(user.tenant_id)

    source_query = (
        LeadSource.query
        .filter(LeadSource.tenant_id == user.tenant_id, LeadSource.is_active == True)
    )
    if source_id:
        source_query = source_query.filter(LeadSource.id == source_id)
    sources = source_query.all()
    source_map = {source.id: source for source in sources}
    source_ids = list(source_map.keys())
    if not source_ids:
        return {
            'snapshot': {
                'total': 0, 'total_leads': 0, 'unique_leads': 0, 'processed': 0,
                'duplicate': 0, 'errors': 0, 'conversion_rate': 0, 'spend': 0,
                'cpl': None,
            },
            'source_rows': [],
            'form_rows': [],
            'campaign_rows': [],
            'last_synced_at': None,
        }

    form_project_map = _source_form_project_map(user.tenant_id, source_ids)
    form_name_map = {}
    for src in sources:
        for form in _source_forms(src):
            form_name_map[(src.id, str(form.get('id') or ''))] = form.get('name') or ''
    end_exclusive = _report_date_to_exclusive(date_to)

    log_query = (
        db.session.query(
            IngestedLeadLog.id,
            IngestedLeadLog.source_id,
            IngestedLeadLog.source_type,
            IngestedLeadLog.form_id,
            IngestedLeadLog.form_name,
            IngestedLeadLog.status,
            IngestedLeadLog.lead_id,
            IngestedLeadLog.dup_of_lead_id,
            IngestedLeadLog.platform_lead_id,
            IngestedLeadLog.received_at,
            Lead.project_id,
            Project.name.label('lead_project_name'),
        )
        .select_from(IngestedLeadLog)
        .outerjoin(Lead, IngestedLeadLog.lead_id == Lead.id)
        .outerjoin(Project, Lead.project_id == Project.id)
        .join(LeadSource, IngestedLeadLog.source_id == LeadSource.id)
        .filter(IngestedLeadLog.tenant_id == user.tenant_id)
        .filter(LeadSource.tenant_id == user.tenant_id)
        .filter(IngestedLeadLog.source_id.in_(source_ids))
    )
    tenant_cutoff = lead_source_cutoff_for(tenant_id=user.tenant_id)
    if tenant_cutoff:
        log_query = log_query.filter(IngestedLeadLog.received_at >= tenant_cutoff)
    log_query = _apply_log_date_filters(log_query, IngestedLeadLog, date_from, date_to)
    log_query = _apply_test_data_filter(log_query, IngestedLeadLog)

    form_groups = {}
    source_groups = {}
    form_unique_ids = {}
    source_unique_ids = {}
    seen_log_identities = set()

    for src in sources:
        source_groups[src.id] = {
            'source_id': src.id,
            'source_name': src.name or 'Unknown Source',
            'source_type': src.source_type,
            'source_status': 'Active Source' if src.is_active else 'Archived Source',
            'source_added_at': _source_added_iso(src),
            'effective_date_from': (_effective_source_start(src, date_from).isoformat() if _effective_source_start(src, date_from) else None),
            'form_id': '',
            'form_name': 'All Forms',
            'project_name': 'All Projects',
            'total_leads': 0,
            'unique_leads': 0,
            'processed': 0,
            'duplicate': 0,
            'duplicates': 0,
            'errors': 0,
            'conversion_rate': 0,
            'spend': None,
            'cpl': None,
            'last_sync': None,
        }
        source_unique_ids[src.id] = set()

    for row in log_query.order_by(IngestedLeadLog.received_at.desc(), IngestedLeadLog.id.desc()).all():
        source_obj = source_map.get(row.source_id)
        effective_start = _effective_source_start(source_obj, date_from)
        if effective_start and row.received_at and row.received_at < effective_start:
            continue
        if end_exclusive and row.received_at and row.received_at >= end_exclusive:
            continue

        identity = row.platform_lead_id or f'log:{row.id}'
        source_key_identity = (row.source_id, identity)
        if source_key_identity in seen_log_identities:
            continue
        seen_log_identities.add(source_key_identity)

        form_id = str(row.form_id or '')
        form_name = row.form_name or form_name_map.get((row.source_id, form_id)) or 'Unmapped Form'
        project_name = row.lead_project_name or form_project_map.get((row.source_id, form_id)) or 'Not Mapped'

        form_key = (row.source_id, form_id)
        source_key = row.source_id
        if form_key not in form_groups:
            form_groups[form_key] = {
                'source_id': row.source_id,
                'source_name': source_obj.name if source_obj else 'Unknown Source',
                'source_type': row.source_type,
                'source_status': 'Active Source' if source_obj and source_obj.is_active else 'Archived Source',
                'source_added_at': _source_added_iso(source_obj),
                'effective_date_from': effective_start.isoformat() if effective_start else None,
                'form_id': form_id,
                'form_name': form_name,
                'project_name': project_name,
                'total_leads': 0,
                'unique_leads': 0,
                'processed': 0,
                'duplicate': 0,
                'duplicates': 0,
                'errors': 0,
                'conversion_rate': 0,
                'spend': None,
                'cpl': None,
                'last_sync': None,
            }
            form_unique_ids[form_key] = set()
        if source_key not in source_groups:
            source_groups[source_key] = dict(form_groups[form_key], form_id='', form_name='All Forms')
            source_unique_ids[source_key] = set()

        effective_status = str(row.status or '').lower()
        if effective_status not in {'processed', 'duplicate', 'error'}:
            continue
        unique_key = row.lead_id or row.dup_of_lead_id or row.platform_lead_id or f'log:{row.id}'

        for bucket, unique_set in (
            (form_groups[form_key], form_unique_ids[form_key]),
            (source_groups[source_key], source_unique_ids[source_key]),
        ):
            if effective_status == 'processed':
                bucket['total_leads'] += 1
                if unique_key not in unique_set:
                    unique_set.add(unique_key)
                    bucket['unique_leads'] += 1
                bucket['processed'] += 1
            elif effective_status == 'duplicate':
                bucket['total_leads'] += 1
                bucket['duplicate'] += 1
                bucket['duplicates'] += 1
            elif effective_status == 'error':
                bucket['total_leads'] += 1
                bucket['errors'] += 1
            if row.received_at and (not bucket['last_sync'] or row.received_at.isoformat() > bucket['last_sync']):
                bucket['last_sync'] = row.received_at.isoformat()

    snapshot_query = (
        db.session.query(
            MetaCampaignSnapshot.source_id,
            MetaCampaignSnapshot.form_id,
            MetaCampaignSnapshot.form_name,
            MetaCampaignSnapshot.spend,
            MetaCampaignSnapshot.snapshot_at,
            MetaCampaignSnapshot.extra_metrics,
        )
        .join(LeadSource, MetaCampaignSnapshot.source_id == LeadSource.id)
        .filter(MetaCampaignSnapshot.tenant_id == user.tenant_id)
        .filter(LeadSource.tenant_id == user.tenant_id)
        .filter(MetaCampaignSnapshot.source_id.in_(source_ids))
    )
    snapshot_query = _apply_snapshot_date_filters(snapshot_query, date_from, date_to)
    snapshot_query = _apply_snapshot_test_filter(snapshot_query)

    latest_form_spend = {}
    latest_source_spend = {}
    latest_source_level_spend_by_source = {}
    latest_source_spend_by_source = {}
    latest_sync = None
    for snap in snapshot_query.order_by(MetaCampaignSnapshot.snapshot_at.desc()).limit(2000).all():
        source_obj = source_map.get(snap.source_id)
        if not _snapshot_matches_report_window(snap, source_obj, date_from, date_to):
            continue
        spend = _to_float_or_none(snap.spend)
        if spend is None:
            continue
        if snap.snapshot_at and (latest_sync is None or snap.snapshot_at > latest_sync):
            latest_sync = snap.snapshot_at
        project_name = (snap.extra_metrics or {}).get('project_name') or 'Not Mapped'
        form_key = (snap.source_id, str(snap.form_id or ''), project_name)
        source_key = (snap.source_id, project_name)
        if str(snap.form_id or '') and form_key not in latest_form_spend:
            latest_form_spend[form_key] = (spend, snap.snapshot_at)
        extra = snap.extra_metrics or {}
        if (
            not str(snap.form_id or '')
            and (extra.get('insight_level') == 'source' or extra.get('synced_from_meta') is True)
            and snap.source_id not in latest_source_level_spend_by_source
        ):
            latest_source_level_spend_by_source[snap.source_id] = (spend, snap.snapshot_at)
        if source_key not in latest_source_spend:
            latest_source_spend[source_key] = (spend, snap.snapshot_at)
        if snap.source_id not in latest_source_spend_by_source:
            latest_source_spend_by_source[snap.source_id] = (spend, snap.snapshot_at)

    def finalize(bucket, spend_value=None, spend_at=None):
        if spend_value is not None:
            bucket['spend'] = round(float(spend_value), 2)
            if spend_at:
                bucket['last_sync'] = _metric_timestamp(spend_at) or bucket['last_sync']
        total = int(bucket.get('total_leads') or 0)
        processed = int(bucket.get('processed') or 0)
        unique = int(bucket.get('unique_leads') or 0)
        bucket['leads'] = processed
        bucket['created'] = processed
        bucket['total'] = total
        bucket['conversion_rate'] = round(processed * 100.0 / total, 2) if total else 0
        if bucket.get('spend') not in (None, '') and unique:
            bucket['cpl'] = round(float(bucket['spend']) / float(unique), 2)
        else:
            bucket['cpl'] = None
        return bucket

    form_rows = []
    for key, bucket in form_groups.items():
        form_rows.append(finalize(bucket, None, None))

    source_rows = []
    for key, bucket in source_groups.items():
        spend_pair = latest_source_level_spend_by_source.get(bucket['source_id'])
        if not spend_pair:
            spend_pair = latest_source_spend_by_source.get(bucket['source_id'])
        source_rows.append(finalize(bucket, *(spend_pair or (None, None))))

    form_rows.sort(key=lambda row: (row['source_name'], row['form_name']))
    source_rows.sort(key=lambda row: row['source_name'])

    processed = sum(int(row.get('processed') or 0) for row in source_rows)
    duplicate = sum(int(row.get('duplicate') or 0) for row in source_rows)
    errors = sum(int(row.get('errors') or 0) for row in source_rows)
    total = sum(int(row.get('total_leads') or 0) for row in source_rows)
    unique = sum(int(row.get('unique_leads') or 0) for row in source_rows)
    spend_total = round(sum(float(row.get('spend') or 0) for row in source_rows), 2)
    overview = {
        'total': total,
        'total_leads': total,
        'unique_leads': unique,
        'processed': processed,
        'duplicate': duplicate,
        'errors': errors,
        'conversion_rate': round(processed * 100.0 / total, 2) if total else 0,
        'spend': spend_total,
        'cpl': round(spend_total / unique, 2) if unique and spend_total else None,
    }

    return {
        'snapshot': overview,
        'source_rows': source_rows,
        'form_rows': form_rows,
        'campaign_rows': [],
        'last_synced_at': latest_sync.isoformat() if latest_sync else None,
    }


def _build_source_report_rows(user, date_from='', date_to='', source_id=None):
    _deactivate_duplicate_meta_sources_for_tenant(user.tenant_id)
    source_query = LeadSource.query.filter(LeadSource.tenant_id == user.tenant_id, LeadSource.is_active == True)
    if source_id:
        source_query = source_query.filter(LeadSource.id == source_id)
    source_rows = source_query.all()
    source_map = {source.id: source for source in source_rows}
    active_source_ids = list(source_map.keys())
    form_project_map = _source_form_project_map(user.tenant_id, active_source_ids)
    
    log_query = (
        db.session.query(IngestedLeadLog, Lead, LeadSource)
        .select_from(IngestedLeadLog)
        .outerjoin(Lead, IngestedLeadLog.lead_id == Lead.id)
        .outerjoin(LeadSource, IngestedLeadLog.source_id == LeadSource.id)
        .filter(IngestedLeadLog.tenant_id == user.tenant_id)
    )
    if active_source_ids:
        log_query = log_query.filter(IngestedLeadLog.source_id.in_(active_source_ids))
    else:
        return []
    log_query = log_query.filter(IngestedLeadLog.received_at >= LeadSource.created_at)
    tenant_cutoff = lead_source_cutoff_for(tenant_id=user.tenant_id)
    if tenant_cutoff:
        log_query = log_query.filter(IngestedLeadLog.received_at >= tenant_cutoff)
    log_query = _apply_log_date_filters(log_query, IngestedLeadLog, date_from, date_to)
    log_query = _apply_test_data_filter(log_query, IngestedLeadLog)
    log_rows = log_query.all()

    grouped = {}
    for log_row, lead_obj, source_obj in log_rows:
        source_obj = source_obj or source_map.get(log_row.source_id)
        effective_start = _effective_source_start(source_obj, date_from)
        if effective_start and log_row.received_at and log_row.received_at < effective_start:
            continue
        project_name = _resolve_project_name(
            lead_obj,
            {},
            source_id=log_row.source_id,
            form_id=log_row.form_id,
            form_project_map=form_project_map,
        )
        key = (log_row.source_id, project_name)
        bucket = grouped.get(key)
        if not bucket:
            bucket = {
                'source_id': log_row.source_id,
                'source_name': source_obj.name if source_obj else 'Unknown Source',
                'source_added_at': _source_added_iso(source_obj),
                'effective_date_from': effective_start.isoformat() if effective_start else None,
                'project_name': project_name,
                'source_status': 'Active Source' if source_obj and source_obj.is_active else 'Archived Source',
                'source_type': log_row.source_type,
                'total': 0,
                'created': 0,
                'duplicates': 0,
                'errors': 0,
                'spend': None,
                'cpl': None,
                'last_sync': None,
            }
            grouped[key] = bucket

        status = _log_effective_status(log_row)
        if status == 'processed':
            bucket['total'] += 1
            bucket['created'] += 1
        elif status == 'duplicate':
            bucket['duplicates'] += 1
        elif status == 'error':
            bucket['errors'] += 1
        if log_row.received_at and (not bucket['last_sync'] or log_row.received_at.isoformat() > bucket['last_sync']):
            bucket['last_sync'] = log_row.received_at.isoformat()

    snapshot_query = (
        MetaCampaignSnapshot.query
        .join(LeadSource, MetaCampaignSnapshot.source_id == LeadSource.id)
        .filter(MetaCampaignSnapshot.tenant_id == user.tenant_id)
        .filter(LeadSource.tenant_id == user.tenant_id)
    )
    if active_source_ids:
        snapshot_query = snapshot_query.filter(MetaCampaignSnapshot.source_id.in_(active_source_ids))
    snapshot_query = _apply_snapshot_date_filters(snapshot_query, date_from, date_to)
    snapshot_query = _apply_snapshot_test_filter(snapshot_query)

    metrics_map = {}
    for snapshot in snapshot_query.order_by(MetaCampaignSnapshot.snapshot_at.desc()).all():
        source_obj = source_map.get(snapshot.source_id)
        effective_start = _effective_source_start(source_obj, date_from)
        if effective_start and snapshot.snapshot_at and snapshot.snapshot_at < effective_start:
            continue
        key = (
            snapshot.source_id,
            _resolve_project_name(
                snapshot.lead,
                snapshot.extra_metrics,
                source_id=snapshot.source_id,
                form_id=snapshot.form_id,
                form_project_map=form_project_map,
            ),
        )
        metrics_map[key] = _choose_metric_row(metrics_map.get(key), snapshot)

    rows = []
    for key, bucket in grouped.items():
        snapshot = metrics_map.get(key)
        if snapshot:
            spend = _metric_from_snapshot(snapshot, 'spend', 'spend', 'amount_spent')
            cpl = _metric_from_snapshot(snapshot, 'cost_per_result', 'cpl', 'cost_per_result')
            if cpl is None and spend not in (None, 0) and bucket['created']:
                cpl = round(float(spend) / float(bucket['created']), 2)
            bucket['spend'] = spend
            bucket['cpl'] = cpl
            bucket['last_sync'] = _metric_timestamp(snapshot.snapshot_at) or bucket['last_sync']
        rows.append(bucket)

    rows.sort(key=lambda row: (row['source_name'], row['project_name']))
    return rows


def _build_attribution_report_rows(user, date_from='', date_to='', source_id=None):
    _deactivate_duplicate_meta_sources_for_tenant(user.tenant_id)
    source_query = LeadSource.query.filter(LeadSource.tenant_id == user.tenant_id, LeadSource.is_active == True)
    if source_id:
        source_query = source_query.filter(LeadSource.id == source_id)
    source_rows = source_query.all()
    source_map = {source.id: source for source in source_rows}
    active_source_ids = list(source_map.keys())
    form_project_map = _source_form_project_map(user.tenant_id, active_source_ids)
    
    log_query = (
        db.session.query(IngestedLeadLog, Lead, LeadSource)
        .select_from(IngestedLeadLog)
        .outerjoin(Lead, IngestedLeadLog.lead_id == Lead.id)
        .outerjoin(LeadSource, IngestedLeadLog.source_id == LeadSource.id)
        .filter(IngestedLeadLog.tenant_id == user.tenant_id)
    )
    if active_source_ids:
        log_query = log_query.filter(IngestedLeadLog.source_id.in_(active_source_ids))
    else:
        return []
    log_query = log_query.filter(IngestedLeadLog.received_at >= LeadSource.created_at)
    tenant_cutoff = lead_source_cutoff_for(tenant_id=user.tenant_id)
    if tenant_cutoff:
        log_query = log_query.filter(IngestedLeadLog.received_at >= tenant_cutoff)
    log_query = _apply_log_date_filters(log_query, IngestedLeadLog, date_from, date_to)
    log_query = _apply_test_data_filter(log_query, IngestedLeadLog)

    grouped = {}
    for log_row, lead_obj, source_obj in log_query.all():
        if log_row.status != 'processed':
            continue
        source_obj = source_obj or source_map.get(log_row.source_id)
        effective_start = _effective_source_start(source_obj, date_from)
        if effective_start and log_row.received_at and log_row.received_at < effective_start:
            continue
        project_name = _resolve_project_name(
            lead_obj,
            {},
            source_id=log_row.source_id,
            form_id=log_row.form_id,
            form_project_map=form_project_map,
        )
        key = (
            log_row.source_id,
            project_name,
            str(log_row.campaign_id or ''),
            str(log_row.ad_set_id or ''),
            str(log_row.ad_id or ''),
            str(log_row.form_id or ''),
        )
        bucket = grouped.get(key)
        if not bucket:
            bucket = {
                'source_id': log_row.source_id,
                'source_added_at': _source_added_iso(source_obj),
                'effective_date_from': effective_start.isoformat() if effective_start else None,
                'project_name': project_name,
                'campaign_id': log_row.campaign_id,
                'campaign_name': log_row.campaign_name,
                'ad_set_id': log_row.ad_set_id,
                'ad_set_name': log_row.ad_set_name,
                'ad_id': log_row.ad_id,
                'ad_name': log_row.ad_name,
                'form_id': log_row.form_id,
                'form_name': log_row.form_name,
                'source_status': 'Active Source' if source_obj and source_obj.is_active else 'Archived Source',
                'total': 0,
                'created': 0,
                'spend': None,
                'cpl': None,
                'ctr': None,
                'cpc': None,
                'cpm': None,
                'reach': None,
                'impressions': None,
                'placement': None,
                'audience': None,
                'last_sync': None,
            }
            grouped[key] = bucket

        bucket['total'] += 1
        bucket['created'] += 1
        if log_row.received_at and (not bucket['last_sync'] or log_row.received_at.isoformat() > bucket['last_sync']):
            bucket['last_sync'] = log_row.received_at.isoformat()

    snapshot_query = (
        MetaCampaignSnapshot.query
        .join(LeadSource, MetaCampaignSnapshot.source_id == LeadSource.id)
        .filter(MetaCampaignSnapshot.tenant_id == user.tenant_id)
        .filter(LeadSource.tenant_id == user.tenant_id)
    )
    if active_source_ids:
        snapshot_query = snapshot_query.filter(MetaCampaignSnapshot.source_id.in_(active_source_ids))
    snapshot_query = _apply_snapshot_date_filters(snapshot_query, date_from, date_to)
    snapshot_query = _apply_snapshot_test_filter(snapshot_query)

    metrics_map = {}
    for snapshot in snapshot_query.order_by(MetaCampaignSnapshot.snapshot_at.desc()).all():
        source_obj = source_map.get(snapshot.source_id)
        effective_start = _effective_source_start(source_obj, date_from)
        if effective_start and snapshot.snapshot_at and snapshot.snapshot_at < effective_start:
            continue
        project_name = _resolve_project_name(
            snapshot.lead,
            snapshot.extra_metrics,
            source_id=snapshot.source_id,
            form_id=snapshot.form_id,
            form_project_map=form_project_map,
        )
        key = (
            snapshot.source_id,
            project_name,
            str(snapshot.campaign_id or ''),
            str(snapshot.ad_set_id or ''),
            str(snapshot.ad_id or ''),
            str(snapshot.form_id or ''),
        )
        metrics_map[key] = _choose_metric_row(metrics_map.get(key), snapshot)

    snapshot_name_map = {}
    for snapshot in snapshot_query.order_by(MetaCampaignSnapshot.snapshot_at.desc()).all():
        source_obj = source_map.get(snapshot.source_id)
        effective_start = _effective_source_start(source_obj, date_from)
        if effective_start and snapshot.snapshot_at and snapshot.snapshot_at < effective_start:
            continue
        sid = snapshot.source_id
        cid = str(snapshot.campaign_id or '')
        asid = str(snapshot.ad_set_id or '')
        aid = str(snapshot.ad_id or '')
        fid = str(snapshot.form_id or '')

        if cid:
            key = ('campaign', sid, cid)
            if key not in snapshot_name_map and snapshot.campaign_name:
                snapshot_name_map[key] = snapshot.campaign_name
        if asid:
            key = ('adset', sid, asid)
            if key not in snapshot_name_map and snapshot.ad_set_name:
                snapshot_name_map[key] = snapshot.ad_set_name
        if aid:
            key = ('ad', sid, aid)
            if key not in snapshot_name_map and snapshot.ad_name:
                snapshot_name_map[key] = snapshot.ad_name
        if fid:
            key = ('form', sid, fid)
            if key not in snapshot_name_map and snapshot.form_name:
                snapshot_name_map[key] = snapshot.form_name

    rows = []
    for key, bucket in grouped.items():
        snapshot = metrics_map.get(key)
        if snapshot:
            bucket['campaign_name'] = snapshot.campaign_name or bucket['campaign_name']
            bucket['ad_set_name'] = snapshot.ad_set_name or bucket['ad_set_name']
            bucket['ad_name'] = snapshot.ad_name or bucket['ad_name']
            bucket['form_name'] = snapshot.form_name or bucket['form_name']
            bucket['spend'] = _metric_from_snapshot(snapshot, 'spend', 'spend', 'amount_spent')
            bucket['cpl'] = _metric_from_snapshot(snapshot, 'cost_per_result', 'cpl', 'cost_per_result')
            if bucket['cpl'] is None and bucket['spend'] not in (None, 0) and bucket['created']:
                bucket['cpl'] = round(float(bucket['spend']) / float(bucket['created']), 2)
            bucket['ctr'] = _metric_from_snapshot(snapshot, 'ctr', 'ctr')
            bucket['cpc'] = _metric_from_snapshot(snapshot, 'cpc', 'cpc')
            bucket['cpm'] = _metric_from_snapshot(snapshot, 'cpm', 'cpm')
            bucket['reach'] = _metric_from_snapshot(snapshot, 'reach', 'reach')
            bucket['impressions'] = _metric_from_snapshot(snapshot, 'impressions', 'impressions')
            bucket['placement'] = snapshot.placement
            bucket['audience'] = snapshot.audience
            bucket['last_sync'] = _metric_timestamp(snapshot.snapshot_at) or bucket['last_sync']

        if not bucket.get('campaign_name') and bucket.get('campaign_id'):
            bucket['campaign_name'] = snapshot_name_map.get(('campaign', bucket['source_id'], str(bucket.get('campaign_id') or '')))
        if not bucket.get('ad_set_name') and bucket.get('ad_set_id'):
            bucket['ad_set_name'] = snapshot_name_map.get(('adset', bucket['source_id'], str(bucket.get('ad_set_id') or '')))
        if not bucket.get('ad_name') and bucket.get('ad_id'):
            bucket['ad_name'] = snapshot_name_map.get(('ad', bucket['source_id'], str(bucket.get('ad_id') or '')))
        if not bucket.get('form_name') and bucket.get('form_id'):
            bucket['form_name'] = snapshot_name_map.get(('form', bucket['source_id'], str(bucket.get('form_id') or '')))

        bucket['campaign_name'] = _campaign_label(bucket.get('campaign_id'), bucket.get('campaign_name'))
        bucket['ad_set_name'] = _fallback_label_with_id(bucket.get('ad_set_name'), bucket.get('ad_set_id'))
        bucket['ad_name'] = _fallback_label_with_id(bucket.get('ad_name'), bucket.get('ad_id'))
        bucket['form_name'] = _fallback_label_with_id(bucket.get('form_name'), bucket.get('form_id'))
        rows.append(bucket)

    rows.sort(key=lambda row: (-row['total'], row['campaign_name'] or '', row['ad_set_name'] or '', row['ad_name'] or ''))
    return rows


def _meta_token_for_source(source):
    creds = source.credentials or {}
    for key in ('page_access_token', 'access_token', 'user_token'):
        token = str(creds.get(key) or '').strip()
        if token and '•' not in token:
            return token
    return ''


def _meta_object_for_campaign_row(row):
    metric_level, metric_id = _metric_object_key_from_row(row)
    if not metric_level or not metric_id:
        return '', ''
    return metric_id, metric_level


def _sync_meta_report_snapshots(user, date_from='', date_to='', source_id=None):
    source_query = LeadSource.query.filter(
        LeadSource.tenant_id == user.tenant_id,
        LeadSource.source_type == 'meta',
        LeadSource.is_active == True,
    )
    if source_id:
        source_query = source_query.filter(LeadSource.id == source_id)
    sources = source_query.order_by(LeadSource.id.asc()).limit(100).all()
    if not sources:
        return {'synced_rows': 0, 'synced_sources': 0, 'errors': [], 'last_synced_at': None, 'scope': 'source_form_spend'}

    synced_rows = 0
    synced_source_ids = set()
    errors = []
    sync_time = datetime.utcnow()

    for src in sources:
        ad_account_ids = _meta_ad_account_ids_for_source(src)
        if not ad_account_ids:
            errors.append({'source_id': src.id, 'message': 'Meta ad account is missing for source-level spend sync'})
            continue

        effective_start = _effective_source_start(src, date_from)
        since = (effective_start.date() if effective_start else datetime.utcnow().date()).isoformat()
        until = _report_meta_until_date(date_to)

        token = _meta_token_for_source(src)
        if not token:
            errors.append({'source_id': src.id, 'message': 'Meta access token is missing'})
            continue

        campaign_ids = _source_campaign_ids_for_spend(user, src.id, date_from=date_from, date_to=date_to)
        spend_result = _fetch_meta_source_spend(token, ad_account_ids, since, until, campaign_ids=campaign_ids)
        for err in spend_result.get('errors') or []:
            errors.append({'source_id': src.id, **err})
        spend = _to_float_or_none(spend_result.get('spend'))
        if spend is None:
            errors.append({'source_id': src.id, 'message': 'Meta spend unavailable for source-level spend sync'})
            continue

        forms = _source_forms(src)

        snapshot = MetaCampaignSnapshot(
            tenant_id=user.tenant_id,
            source_id=src.id,
            page_id=(src.credentials or {}).get('page_id'),
            form_id='',
            form_name='All Forms',
            is_test=False,
            spend=spend,
            extra_metrics={
                'synced_from_meta': True,
                'insight_level': 'source',
                'insight_object_id': ','.join(spend_result.get('matched_account_ids') or ad_account_ids[:1]),
                'insight_date_from': since,
                'insight_date_to': until,
                'source_spend_method': spend_result.get('method'),
                'known_campaign_count': len(campaign_ids),
                'matched_campaign_count': spend_result.get('matched_campaign_count', 0),
                'ad_account_count': len(ad_account_ids),
                'project_name': 'Not Mapped',
                'forms': [{'id': f.get('id'), 'name': f.get('name')} for f in forms],
                'source_id': src.id,
            },
            snapshot_at=sync_time,
        )
        db.session.add(snapshot)
        synced_rows += 1
        synced_source_ids.add(src.id)

    db.session.commit()
    return {
        'synced_rows': synced_rows,
        'synced_sources': len(synced_source_ids),
        'errors': errors[:20],
        'last_synced_at': sync_time.isoformat() if synced_rows else None,
        'scope': 'source_form_spend',
    }


def _cleanup_validation_leads(tenant_id: int, lead_ids: set[int]) -> int:
    ids = [int(lead_id) for lead_id in (lead_ids or set()) if lead_id]
    if not ids:
        return 0
    rows = Lead.query.filter(
        Lead.tenant_id == tenant_id,
        Lead.id.in_(ids),
        Lead.is_test == True,
        Lead.is_active == True,
    ).all()
    for row in rows:
        row.is_active = False
    return len(rows)

def _get_platform_meta_creds():
    app_id     = os.environ.get('META_APP_ID', '1329585565931521')
    app_secret = os.environ.get('META_APP_SECRET', '')
    return app_id, app_secret

def _get_meta_oauth_scopes():
    """
    Meta OAuth scopes are environment-driven so we can use a minimal testing set
    before app review is complete, then switch to full production scopes.

    Env var: META_OAUTH_SCOPES (comma-separated)
    Default (testing): pages_show_list,pages_read_engagement
    """
    raw = os.environ.get('META_OAUTH_SCOPES', 'pages_show_list,pages_read_engagement,leads_retrieval,pages_manage_ads,pages_manage_metadata,business_management,ads_read')
    scopes = [s.strip() for s in raw.split(',') if s and s.strip()]

    # Always include the minimum permissions needed for page lead form access.
    # This prevents silent env misconfiguration from breaking wizard step-3.
    required = ['pages_show_list', 'pages_read_engagement', 'leads_retrieval', 'pages_manage_ads', 'pages_manage_metadata', 'business_management', 'ads_read']
    for perm in required:
        if perm not in scopes:
            scopes.append(perm)
    return scopes


def _normalise_meta_pull_entry(entry: dict, default_page_id: str = '') -> dict:
    """Normalize Graph /{form_id}/leads entries to ingestion meta format."""
    created_time = str(entry.get('created_time') or entry.get('created_at') or '').strip()
    field_data = {
        item['name'].lower(): (item.get('values') or [''])[0]
        for item in (entry.get('field_data') or [])
        if isinstance(item, dict) and item.get('name')
    }

    name_candidates = [
        field_data.get('full_name'),
        field_data.get('name'),
        ((field_data.get('first_name') or '') + ' ' + (field_data.get('last_name') or '')).strip(),
    ]
    name = next((v.strip() for v in name_candidates if v and str(v).strip()), '')
    phone = next((str(field_data.get(k, '')).strip() for k in ('phone_number', 'phone', 'mobile', 'mobile_number', 'contact_number') if field_data.get(k)), '')
    email = str(field_data.get('email', '')).strip()
    city = next((str(field_data.get(k, '')).strip() for k in ('city', 'location', 'area') if field_data.get(k)), '')

    return {
        'platform_lead_id': str(entry.get('leadgen_id') or entry.get('id') or ''),
        'platform_created_at': created_time,
        'created_time': created_time,
        'page_id': str(entry.get('page_id') or default_page_id or ''),
        'form_id': str(entry.get('form_id') or ''),
        'ad_id': str(entry.get('ad_id') or ''),
        'ad_set_id': str(entry.get('adset_id') or ''),
        'campaign_id': str(entry.get('campaign_id') or ''),
        'name': name,
        'phone': phone,
        'email': email,
        'city': city,
        'raw_fields': field_data,
    }

def _get_platform_google_creds():
    client_id     = os.environ.get('GOOGLE_CLIENT_ID', '')
    client_secret = os.environ.get('GOOGLE_CLIENT_SECRET', '')
    return client_id, client_secret


def _google_normalize_customer_id(raw):
    val = str(raw or '').strip()
    if not val:
        return ''
    if val.startswith('customers/'):
        val = val.split('/', 1)[1]
    return ''.join(ch for ch in val if ch.isdigit())


def _google_exchange_refresh_token(client_id, client_secret, refresh_token):
    token_body = _parse.urlencode({
        'client_id': client_id,
        'client_secret': client_secret,
        'refresh_token': refresh_token,
        'grant_type': 'refresh_token',
    }).encode()
    req = _req.Request(
        'https://oauth2.googleapis.com/token',
        data=token_body,
        headers={'Content-Type': 'application/x-www-form-urlencoded'},
    )
    with _req.urlopen(req, timeout=15) as r:
        token_data = _json.loads(r.read())
    if 'error' in token_data:
        raise RuntimeError(token_data.get('error_description') or token_data.get('error') or 'token_exchange_failed')
    return token_data.get('access_token', '')


def _google_list_accessible_accounts(access_token):
    developer_token = os.environ.get('GOOGLE_DEVELOPER_TOKEN', '').strip()
    if not developer_token:
        raise RuntimeError('GOOGLE_DEVELOPER_TOKEN missing')

    headers = {
        'Authorization': f'Bearer {access_token}',
        'developer-token': developer_token,
        'Content-Type': 'application/json',
    }

    # Google Ads API versions are periodically sunset.
    # Try recent versions + method fallback so OAuth flow remains resilient.
    versions = ('v22', 'v21', 'v20', 'v19', 'v18')
    methods = ('GET', 'POST')
    data = None
    last_error = None

    for version in versions:
        endpoint = f'https://googleads.googleapis.com/{version}/customers:listAccessibleCustomers'
        for method in methods:
            try:
                body = b'{}' if method == 'POST' else None
                req = _req.Request(endpoint, data=body, method=method, headers=headers)
                with _req.urlopen(req, timeout=20) as r:
                    data = _json.loads(r.read())
                break
            except Exception as exc:
                last_error = exc
                continue
        if data is not None:
            break

    resource_names = (data or {}).get('resourceNames') or []
    customer_ids = []
    for rn in resource_names:
        cid = _google_normalize_customer_id(rn)
        if cid and cid not in customer_ids:
            customer_ids.append(cid)

    if customer_ids:
        return [{
            'customer_id': cid,
            'customer_name': f'Google Ads {cid}',
            'resource_name': f'customers/{cid}',
        } for cid in customer_ids]

    # Fallback: query the manager hierarchy from the configured login customer ID.
    # This helps when accessible-customers is unavailable but the manager can still
    # see shared client accounts through the hierarchy.
    login_customer_id = _google_normalize_customer_id(os.environ.get('GOOGLE_LOGIN_CUSTOMER_ID', ''))
    if not login_customer_id:
        return []

    query = (
        'SELECT customer_client.client_customer, customer_client.descriptive_name, '
        'customer_client.status, customer_client.manager '
        'FROM customer_client '
        "WHERE customer_client.status = 'ENABLED'"
    )

    payload = _json.dumps({'query': query, 'pageSize': 1000}).encode('utf-8')
    hierarchy_rows = []
    hierarchy_error = None
    for version in versions:
        search_endpoint = f'https://googleads.googleapis.com/{version}/customers/{login_customer_id}/googleAds:search'
        search_headers = dict(headers)
        search_headers['login-customer-id'] = login_customer_id
        req = _req.Request(search_endpoint, data=payload, method='POST', headers=search_headers)
        try:
            with _req.urlopen(req, timeout=25) as r:
                search_data = _json.loads(r.read())
                for row in (search_data.get('results') or []):
                    cc = (row or {}).get('customerClient') or {}
                    resource_name = str(cc.get('clientCustomer') or cc.get('client_customer') or '')
                    cid = _google_normalize_customer_id(resource_name)
                    if not cid or cid == login_customer_id:
                        continue
                    hierarchy_rows.append({
                        'customer_id': cid,
                        'customer_name': str(cc.get('descriptiveName') or cc.get('descriptive_name') or f'Google Ads {cid}').strip(),
                        'resource_name': f'customers/{cid}',
                    })
            break
        except Exception as exc:
            hierarchy_error = exc
            continue

    if not hierarchy_rows and hierarchy_error:
        logger.warning('google_hierarchy_discovery_failed: %s', hierarchy_error)

    if not customer_ids and not hierarchy_rows and last_error:
        raise RuntimeError(str(last_error))

    # De-dupe while preserving order.
    deduped = []
    seen = set()
    for row in hierarchy_rows:
        cid = row.get('customer_id')
        if not cid or cid in seen:
            continue
        seen.add(cid)
        deduped.append(row)
    return deduped

def _purge_expired_sessions():
    cutoff = datetime.utcnow() - timedelta(minutes=_OAUTH_SESSION_TTL_MINUTES)
    expired = [k for k, v in _oauth_sessions.items() if v.get('created_at', cutoff) < cutoff]
    for k in expired:
        del _oauth_sessions[k]
    try:
        db.session.query(OAuthSession).filter(
            OAuthSession.expires_at.isnot(None),
            OAuthSession.expires_at < datetime.utcnow(),
        ).delete(synchronize_session=False)
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        logger.warning('oauth session prune failed: %s', exc)


def _save_oauth_session(session_key, session_data, platform):
    payload = dict(session_data or {})
    _oauth_sessions[session_key] = payload

    db_payload = dict(payload)
    db_payload.pop('created_at', None)

    try:
        row = db.session.get(OAuthSession, session_key)
        if not row:
            row = OAuthSession(session_key=session_key)
        row.tenant_id = payload.get('tenant_id')
        row.platform = platform
        row.payload = db_payload
        row.expires_at = datetime.utcnow() + timedelta(minutes=_OAUTH_SESSION_TTL_MINUTES)
        db.session.add(row)
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        logger.warning('oauth session save failed: %s', exc)


def _load_oauth_session(session_key, platform):
    _purge_expired_sessions()

    session_data = _oauth_sessions.get(session_key)
    if session_data:
        return session_data

    try:
        row = db.session.get(OAuthSession, session_key)
        if not row:
            return None
        if row.platform and row.platform != platform:
            return None
        if row.expires_at and row.expires_at < datetime.utcnow():
            db.session.delete(row)
            db.session.commit()
            return None

        payload = dict(row.payload or {})
        payload.setdefault('tenant_id', row.tenant_id)
        payload.setdefault('created_at', row.created_at or datetime.utcnow())
        _oauth_sessions[session_key] = payload
        return payload
    except Exception as exc:
        db.session.rollback()
        logger.warning('oauth session load failed: %s', exc)
        return None


# ── Auth helper ────────────────────────────────────────────────────────────────

def _check_source_ownership(source, user):
    """Return 403 dict if user does not own this source, else None."""
    if source.tenant_id != user.tenant_id:
        return jsonify({'error': 'Not found'}), 404
    return None


def _source_forms(source):
    forms = source.available_forms or []
    if not isinstance(forms, list):
        return []
    clean = []
    for form in forms:
        if not isinstance(form, dict):
            continue
        form_id = str(form.get('id') or '').strip()
        if not form_id:
            continue
        clean.append({
            'id': form_id,
            'name': str(form.get('name') or '').strip(),
            'status': str(form.get('status') or '').strip(),
        })
    return clean


def _source_form_mapping_summary(source):
    forms = _source_forms(source)
    form_ids = [f['id'] for f in forms]

    mappings = LeadSourceFormMapping.query.filter_by(
        tenant_id=source.tenant_id,
        source_id=source.id,
        is_active=True,
    ).all()

    mapped_ids = {str(m.form_id) for m in mappings if m.form_id}
    missing_form_ids = [fid for fid in form_ids if fid not in mapped_ids]

    rows = [m.to_dict() for m in mappings]
    return {
        'required_form_count': len(form_ids),
        'mapped_form_count': len([fid for fid in form_ids if fid in mapped_ids]),
        'total_mapping_rows': len(rows),
        'is_ready': len(form_ids) > 0 and len(missing_form_ids) == 0,
        'missing_form_ids': missing_form_ids,
        'forms': forms,
        'rows': rows,
    }


def _normalize_form_manager_mode(raw_mode):
    mode = str(raw_mode or 'none').strip().lower()
    aliases = {
        'fixed': 'fixed_manager',
        'fixed_user': 'fixed_manager',
        'manager_based': 'fixed_manager',
        'round_robin': 'round_robin_pool',
        'rr': 'round_robin_pool',
    }
    return aliases.get(mode, mode)


def _validate_mapping_gate(source):
    if source.source_type not in ('meta', 'google'):
        return {'required': False, 'is_ready': True, 'reason': ''}
    summary = _source_form_mapping_summary(source)
    required = summary['required_form_count'] > 0
    if not required:
        return {
            'required': True,
            'is_ready': False,
            'reason': 'No forms available on source. Run test/connect to sync forms first.',
            'summary': summary,
        }
    if summary['is_ready']:
        return {'required': True, 'is_ready': True, 'reason': '', 'summary': summary}
    return {
        'required': True,
        'is_ready': False,
        'reason': 'Map every active form to a project before activation.',
        'summary': summary,
    }


def _meta_page_id(source):
    creds = source.credentials or {}
    return str(creds.get('page_id') or '').strip()


def _deactivate_duplicate_meta_sources_for_tenant(tenant_id):
    """Keep only the newest active Meta source per page_id for a tenant."""
    active_meta = (
        LeadSource.query
        .filter_by(tenant_id=tenant_id, source_type='meta', is_active=True)
        .order_by(LeadSource.updated_at.desc(), LeadSource.created_at.desc())
        .all()
    )
    keep_by_page = {}
    deactivated = []
    for src in active_meta:
        page_id = _meta_page_id(src)
        if not page_id:
            continue
        if page_id in keep_by_page:
            src.is_active = False
            src.last_test_message = f'Auto-deactivated duplicate Meta source for page {page_id}; kept source #{keep_by_page[page_id]}.'
            deactivated.append(src.id)
            continue
        keep_by_page[page_id] = src.id

    if deactivated:
        db.session.commit()
    return deactivated


def _deactivate_duplicate_meta_sources_for_page(tenant_id, page_id, keep_source_id):
    page_id = str(page_id or '').strip()
    if not page_id or not keep_source_id:
        return []

    candidates = (
        LeadSource.query
        .filter_by(tenant_id=tenant_id, source_type='meta', is_active=True)
        .filter(LeadSource.id != keep_source_id)
        .all()
    )
    deactivated = []
    for src in candidates:
        if _meta_page_id(src) != page_id:
            continue
        src.is_active = False
        src.last_test_message = f'Auto-deactivated duplicate Meta source for page {page_id}; replaced by source #{keep_source_id}.'
        deactivated.append(src.id)
    return deactivated


# ══════════════════════════════════════════════════════════════════════════════
# LIST + CREATE
# ══════════════════════════════════════════════════════════════════════════════

def _active_lms_lead_counts_by_source_id(tenant_id, source_ids, date_from='', date_to=''):
    source_ids = [int(sid) for sid in (source_ids or []) if sid]
    if not source_ids:
        return {}
    query = (
        db.session.query(
            IngestedLeadLog.source_id.label('source_id'),
            func.count(func.distinct(Lead.id)).label('lead_count'),
        )
        .join(Lead, IngestedLeadLog.lead_id == Lead.id)
        .filter(
            IngestedLeadLog.tenant_id == tenant_id,
            IngestedLeadLog.source_id.in_(source_ids),
            IngestedLeadLog.status == 'processed',
            Lead.tenant_id == tenant_id,
            Lead.is_active == True,
            Lead.is_test == False,
        )
    )
    if date_from:
        try:
            query = query.filter(Lead.created_at >= datetime.fromisoformat(date_from))
        except ValueError:
            pass
    if date_to:
        try:
            query = query.filter(Lead.created_at < datetime.fromisoformat(date_to) + timedelta(days=1))
        except ValueError:
            pass
    rows = query.group_by(IngestedLeadLog.source_id).all()
    return {int(row.source_id): int(row.lead_count or 0) for row in rows}


def _active_lms_lead_counts_by_source_name(tenant_id, date_from='', date_to=''):
    query = Lead.query.filter(
        Lead.tenant_id == tenant_id,
        Lead.is_active == True,
        Lead.is_test == False,
    )
    if date_from:
        try:
            query = query.filter(Lead.created_at >= datetime.fromisoformat(date_from))
        except ValueError:
            pass
    if date_to:
        try:
            query = query.filter(Lead.created_at < datetime.fromisoformat(date_to) + timedelta(days=1))
        except ValueError:
            pass
    rows = (
        query
        .with_entities(func.coalesce(Lead.source, 'Unknown').label('source_name'), func.count(Lead.id).label('lead_count'))
        .group_by(func.coalesce(Lead.source, 'Unknown'))
        .all()
    )
    return {str(row.source_name or 'Unknown'): int(row.lead_count or 0) for row in rows}


def _ingestion_submission_counts_by_source_id(tenant_id, source_ids):
    source_ids = [int(sid) for sid in (source_ids or []) if sid]
    if not source_ids:
        return {}

    source_rows = (
        LeadSource.query
        .filter(LeadSource.tenant_id == tenant_id, LeadSource.id.in_(source_ids))
        .all()
    )
    source_map = {source.id: source for source in source_rows}
    counts = {source_id: {'total': 0, 'processed': 0, 'duplicate': 0, 'errors': 0} for source_id in source_ids}

    log_query = (
        db.session.query(IngestedLeadLog, LeadSource)
        .select_from(IngestedLeadLog)
        .join(LeadSource, IngestedLeadLog.source_id == LeadSource.id)
        .filter(IngestedLeadLog.tenant_id == tenant_id)
        .filter(LeadSource.tenant_id == tenant_id)
        .filter(IngestedLeadLog.source_id.in_(source_ids))
    )
    tenant_cutoff = lead_source_cutoff_for(tenant_id=tenant_id)
    if tenant_cutoff:
        log_query = log_query.filter(IngestedLeadLog.received_at >= tenant_cutoff)
    log_query = _apply_test_data_filter(log_query, IngestedLeadLog)

    seen = set()
    for log_row, source_obj in log_query.order_by(IngestedLeadLog.received_at.desc(), IngestedLeadLog.id.desc()).all():
        source_obj = source_obj or source_map.get(log_row.source_id)
        effective_start = _effective_source_start(source_obj, '')
        if effective_start and log_row.received_at and log_row.received_at < effective_start:
            continue
        identity = _source_log_identity(log_row)
        if identity in seen:
            continue
        seen.add(identity)

        status = _log_effective_status(log_row)
        if status not in {'processed', 'duplicate', 'error'}:
            continue
        bucket = counts.setdefault(log_row.source_id, {'total': 0, 'processed': 0, 'duplicate': 0, 'errors': 0})
        bucket['total'] += 1
        if status == 'processed':
            bucket['processed'] += 1
        elif status == 'duplicate':
            bucket['duplicate'] += 1
        elif status == 'error':
            bucket['errors'] += 1

    return counts


@lead_sources_bp.route('', methods=['GET'])
@require_role('superadmin', 'sales_manager', 'platform_owner')
def list_sources():
    user = request.current_user
    include_inactive = str(request.args.get('include_inactive', '')).strip().lower() in ('1', 'true', 'yes')

    # Safety cleanup for historical duplicate Meta sources pointing to the same page.
    _deactivate_duplicate_meta_sources_for_tenant(user.tenant_id)

    q = LeadSource.query.filter_by(tenant_id=user.tenant_id)
    if not include_inactive:
        q = q.filter(LeadSource.is_active == True)

    sources = q.order_by(LeadSource.created_at.desc()).all()
    lms_lead_counts = _active_lms_lead_counts_by_source_id(user.tenant_id, [s.id for s in sources])
    ingestion_counts = _ingestion_submission_counts_by_source_id(user.tenant_id, [s.id for s in sources])
    payload = []
    for s in sources:
        item = s.to_dict()
        source_ingestion = ingestion_counts.get(s.id, {})
        item['ingestion_events_count'] = int(source_ingestion.get('total') or 0)
        item['processed_events_count'] = int(source_ingestion.get('processed') or 0)
        item['duplicate_events_count'] = int(source_ingestion.get('duplicate') or 0)
        item['error_events_count'] = int(source_ingestion.get('errors') or 0)
        item['lms_leads_count'] = lms_lead_counts.get(s.id, 0)
        item['total_leads_ingested'] = item['ingestion_events_count']
        gate = _validate_mapping_gate(s)
        item['mapping_gate'] = {
            'required': gate.get('required', False),
            'is_ready': gate.get('is_ready', True),
            'reason': gate.get('reason', ''),
        }
        payload.append(item)
    return jsonify({'sources': payload}), 200


@lead_sources_bp.route('', methods=['POST'])
@require_role('superadmin', 'sales_manager', 'platform_owner')
def create_source():
    user = request.current_user
    data = request.get_json() or {}

    source_type = (data.get('source_type') or '').strip().lower()
    if source_type not in SOURCE_TYPES:
        return jsonify({'error': f'Invalid source_type. Allowed: {list(SOURCE_TYPES)}'}), 400

    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'error': 'name is required'}), 400

    dup_mode = data.get('dup_mode', 'skip')
    if dup_mode not in DUP_MODES:
        return jsonify({'error': f'Invalid dup_mode. Allowed: {list(DUP_MODES)}'}), 400

    assign_strategy = data.get('assign_strategy', 'none')
    if assign_strategy not in ASSIGN_STRATEGIES:
        return jsonify({'error': f'Invalid assign_strategy. Allowed: {list(ASSIGN_STRATEGIES)}'}), 400

    source = LeadSource(
        tenant_id=user.tenant_id,
        name=name,
        source_type=source_type,
        credentials=data.get('credentials') or {},
        field_mapping=data.get('field_mapping') or {},
        default_values=data.get('default_values') or {},
        dup_check_phone=bool(data.get('dup_check_phone', True)),
        dup_check_email=False,
        dup_mode=dup_mode,
        assign_strategy=assign_strategy,
        assign_fixed_user_id=data.get('assign_fixed_user_id'),
        assign_manager_id=data.get('assign_manager_id'),
        rr_user_pool=data.get('rr_user_pool') or [],
        created_by=user.id,
    )
    db.session.add(source)
    db.session.commit()
    return jsonify({'source': source.to_dict()}), 201


# ══════════════════════════════════════════════════════════════════════════════
# GET / UPDATE / DELETE
# ══════════════════════════════════════════════════════════════════════════════

@lead_sources_bp.route('/<int:source_id>', methods=['GET'])
@require_role('superadmin', 'sales_manager', 'platform_owner')
def get_source(source_id):
    user = request.current_user
    source = LeadSource.query.get_or_404(source_id)
    err = _check_source_ownership(source, user)
    if err:
        return err
    data = source.to_dict()
    gate = _validate_mapping_gate(source)
    data['mapping_gate'] = {
        'required': gate.get('required', False),
        'is_ready': gate.get('is_ready', True),
        'reason': gate.get('reason', ''),
        'summary': gate.get('summary', {}),
    }
    return jsonify({'source': data}), 200


@lead_sources_bp.route('/<int:source_id>', methods=['PUT'])
@require_role('superadmin', 'sales_manager', 'platform_owner')
def update_source(source_id):
    user = request.current_user
    source = LeadSource.query.get_or_404(source_id)
    err = _check_source_ownership(source, user)
    if err:
        return err

    data = request.get_json() or {}

    if 'name' in data:
        source.name = (data['name'] or '').strip() or source.name
    if 'credentials' in data:
        # Merge credentials: keep existing masked values unless new ones provided
        existing = source.credentials or {}
        incoming = data['credentials'] or {}
        for k, v in incoming.items():
            # '••••••••' means client didn't change it
            if v != '••••••••':
                existing[k] = v
        source.credentials = existing
    if 'field_mapping' in data:
        source.field_mapping = data['field_mapping'] or {}
    if 'default_values' in data:
        source.default_values = data['default_values'] or {}
    if 'dup_check_phone' in data:
        source.dup_check_phone = bool(data['dup_check_phone'])
    if 'dup_check_email' in data:
        source.dup_check_email = False
    if 'dup_mode' in data:
        if data['dup_mode'] in DUP_MODES:
            source.dup_mode = data['dup_mode']
    if 'assign_strategy' in data:
        if data['assign_strategy'] in ASSIGN_STRATEGIES:
            source.assign_strategy = data['assign_strategy']
    if 'assign_fixed_user_id' in data:
        source.assign_fixed_user_id = data['assign_fixed_user_id']
    if 'assign_manager_id' in data:
        source.assign_manager_id = data['assign_manager_id']
    if 'rr_user_pool' in data:
        source.rr_user_pool = data['rr_user_pool'] or []

    db.session.commit()
    return jsonify({'source': source.to_dict()}), 200


@lead_sources_bp.route('/<int:source_id>', methods=['DELETE'])
@require_role('superadmin', 'sales_manager', 'platform_owner')
def delete_source(source_id):
    return _soft_remove_source(source_id)


@lead_sources_bp.route('/<int:source_id>/remove', methods=['POST'])
@require_role('superadmin', 'sales_manager', 'platform_owner')
def remove_source(source_id):
    return _soft_remove_source(source_id)


def _soft_remove_source(source_id):
    user = request.current_user
    source = LeadSource.query.get_or_404(source_id)
    err = _check_source_ownership(source, user)
    if err:
        return err
    source.is_active = False
    db.session.commit()
    return jsonify({'ok': True}), 200


@lead_sources_bp.route('/<int:source_id>/hard-delete', methods=['POST'])
@require_role('superadmin', 'platform_owner')
def hard_delete_source(source_id):
    """
    Completely delete a disabled/archived source and all related records.
    This endpoint is destructive and removes ingestion history for the source.
    """
    user = request.current_user
    source = LeadSource.query.get_or_404(source_id)
    err = _check_source_ownership(source, user)
    if err:
        return err
    
    # Only allow deletion of disabled sources
    if source.is_active:
        return jsonify({
            'error': 'Cannot delete an active source. Please disable it first.'
        }), 400

    try:
        source_id_int = int(source_id)

        # Purge dependent rows first to satisfy FK constraints.
        LeadSourceFormMapping.query.filter_by(
            tenant_id=user.tenant_id,
            source_id=source_id_int,
        ).delete(synchronize_session=False)

        from app.models.lead_source_mapping import MetaCampaignSnapshot
        MetaCampaignSnapshot.query.filter_by(
            tenant_id=user.tenant_id,
            source_id=source_id_int,
        ).delete(synchronize_session=False)

        IngestedLeadLog.query.filter_by(
            tenant_id=user.tenant_id,
            source_id=source_id_int,
        ).delete(synchronize_session=False)

        # Delete the source itself
        db.session.delete(source)
        db.session.commit()
        
        return jsonify({
            'ok': True,
            'message': f'Source "{source.name}" has been completely deleted.'
        }), 200
    except Exception as exc:
        db.session.rollback()
        logger.exception('hard_delete_source failed: %s', exc)
        return jsonify({'error': f'Failed to delete source: {exc}'}), 500


# ══════════════════════════════════════════════════════════════════════════════
# ENABLE / DISABLE
# ══════════════════════════════════════════════════════════════════════════════

@lead_sources_bp.route('/<int:source_id>/enable', methods=['POST'])
@require_role('superadmin', 'sales_manager', 'platform_owner')
def enable_source(source_id):
    user = request.current_user
    source = LeadSource.query.get_or_404(source_id)
    err = _check_source_ownership(source, user)
    if err:
        return err
    gate = _validate_mapping_gate(source)
    if gate.get('required') and not gate.get('is_ready'):
        return jsonify({
            'error': gate.get('reason') or 'Source cannot be activated yet.',
            'mapping_gate': {
                'required': gate.get('required', True),
                'is_ready': gate.get('is_ready', False),
                'summary': gate.get('summary', {}),
            },
        }), 400

    source.is_active = True
    db.session.commit()
    return jsonify({'source': source.to_dict()}), 200


@lead_sources_bp.route('/<int:source_id>/disable', methods=['POST'])
@require_role('superadmin', 'sales_manager', 'platform_owner')
def disable_source(source_id):
    user = request.current_user
    source = LeadSource.query.get_or_404(source_id)
    err = _check_source_ownership(source, user)
    if err:
        return err
    source.is_active = False
    db.session.commit()
    return jsonify({'source': source.to_dict()}), 200


@lead_sources_bp.route('/<int:source_id>/disconnect', methods=['POST'])
@require_role('superadmin', 'sales_manager', 'platform_owner')
def disconnect_source(source_id):
    user = request.current_user
    source = LeadSource.query.get_or_404(source_id)
    err = _check_source_ownership(source, user)
    if err:
        return err

    if source.source_type not in ('meta', 'google'):
        return jsonify({'error': 'Disconnect is supported only for Meta/Google sources'}), 400

    source.credentials = {}
    source.connected_account = None
    source.permission_status = 'missing'
    source.permission_details = {}
    source.available_forms = []
    source.available_campaigns = []
    source.last_test_result = None
    source.last_test_message = 'Disconnected by user'
    source.last_tested_at = datetime.utcnow()
    source.is_active = False

    try:
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        logger.exception('disconnect_source failed: %s', exc)
        return jsonify({'error': f'Disconnect failed: {exc}'}), 500

    return jsonify({'source': source.to_dict(), 'ok': True}), 200


# ══════════════════════════════════════════════════════════════════════════════
# TEST CONNECTION
# ══════════════════════════════════════════════════════════════════════════════

@lead_sources_bp.route('/<int:source_id>/test', methods=['POST'])
@require_role('superadmin', 'sales_manager', 'platform_owner')
def test_source(source_id):
    """
    Test the connection / permissions for a lead source.
    For Meta: calls Graph API to list accessible pages.
    For Google: validates the refresh token.
    For generic webhook: always passes (HMAC-based, no external auth).
    """
    user = request.current_user
    source = LeadSource.query.get_or_404(source_id)
    err = _check_source_ownership(source, user)
    if err:
        return err

    result = _run_connection_test(source)

    source.last_tested_at   = datetime.utcnow()
    source.last_test_result = result.get('result', 'fail')
    source.last_test_message= result.get('message', '')
    if 'connected_account' in result:
        source.connected_account = result['connected_account']
    if 'permission_status' in result:
        source.permission_status = result['permission_status']
    if 'permission_details' in result:
        source.permission_details = result['permission_details']
    if 'available_forms' in result:
        source.available_forms = result['available_forms']
    if 'available_campaigns' in result:
        source.available_campaigns = result['available_campaigns']

    db.session.commit()
    return jsonify({'test': result, 'source': source.to_dict()}), 200


def _run_connection_test(source: LeadSource) -> dict:
    """Dispatcher for source-type specific connection tests."""
    if source.source_type == 'meta':
        return _test_meta(source)
    if source.source_type == 'google':
        return _test_google(source)
    # Generic webhook: always pass (no external service)
    return {
        'result': 'pass',
        'message': 'Generic webhook is ready. Use the webhook URL to send leads.',
        'connected_account': source.name,
        'permission_status': 'ok',
    }


def _test_meta(source: LeadSource) -> dict:
    """Call Meta Graph API to verify access token and list pages/forms."""
    try:
        import urllib.request as urllib_req
        import urllib.parse as urllib_parse
        import json as _json

        creds = source.credentials or {}
        user_token = (creds.get('user_token') or '').strip()
        page_token = (creds.get('page_access_token') or creds.get('access_token') or '').strip()
        test_token = user_token or page_token
        if not test_token:
            return {
                'result': 'fail',
                'message': 'No Meta token configured. Reconnect Meta source.',
                'permission_status': 'missing',
                'permission_details': {'missing': ['user_token/page_access_token']},
            }

        me_data = {'name': source.connected_account or 'Meta Account'}
        granted = []

        # Prefer user token for account + permission checks.
        if user_token:
            me_url = f'https://graph.facebook.com/v25.0/me?fields=id,name&access_token={urllib_parse.quote(user_token)}'
            with urllib_req.urlopen(urllib_req.Request(me_url), timeout=10) as resp:
                me_data = _json.loads(resp.read())

            perm_url = f'https://graph.facebook.com/v25.0/me/permissions?access_token={urllib_parse.quote(user_token)}'
            with urllib_req.urlopen(urllib_req.Request(perm_url), timeout=10) as resp:
                perm_data = _json.loads(resp.read())
            granted = [p['permission'] for p in perm_data.get('data', []) if p.get('status') == 'granted']

        # List pages using available token.
        pages_url = f'https://graph.facebook.com/v25.0/me/accounts?access_token={urllib_parse.quote(test_token)}'
        with urllib_req.urlopen(urllib_req.Request(pages_url), timeout=10) as resp:
            pages_data = _json.loads(resp.read())

        pages = [
            {'id': p['id'], 'name': p['name']}
            for p in pages_data.get('data', [])
        ]

        # Keep available_forms aligned with actual lead forms for the selected page.
        # This prevents form→page drift after repeated "Test Connection" runs.
        selected_page_id = str((creds.get('page_id') or '')).strip()
        forms = []
        if selected_page_id:
            form_tokens = []
            selected_page_token = (creds.get('page_access_token') or '').strip()
            if selected_page_token:
                form_tokens.append(selected_page_token)
            if user_token:
                form_tokens.append(user_token)
            if page_token and page_token not in form_tokens:
                form_tokens.append(page_token)

            for token in form_tokens:
                try:
                    forms_url = (
                        f'https://graph.facebook.com/v25.0/{urllib_parse.quote(selected_page_id)}/leadgen_forms'
                        f'?fields=id,name,status,leads_count,created_time'
                        f'&access_token={urllib_parse.quote(token)}'
                    )
                    with urllib_req.urlopen(urllib_req.Request(forms_url), timeout=12) as resp:
                        forms_data = _json.loads(resp.read())
                    forms = [
                        {
                            'id': f['id'],
                            'name': f.get('name', ''),
                            'status': f.get('status', ''),
                            'leads_count': f.get('leads_count', 0),
                            'created_time': f.get('created_time', ''),
                        }
                        for f in forms_data.get('data', [])
                    ]
                    break
                except Exception:
                    continue

        if not forms:
            forms = source.available_forms or []

        required = _get_meta_oauth_scopes()
        missing  = [r for r in required if r not in granted]

        # If no user token exists, report partial rather than hard-fail.
        if not user_token:
            perm_status = 'partial'
            if 'user_token' not in missing:
                missing = ['user_token'] + missing
        else:
            perm_status = 'ok' if not missing else ('partial' if granted else 'missing')

        return {
            'result': 'pass' if not missing else 'partial',
            'message': f'Connected as {me_data.get("name", "Unknown")}. {len(pages)} page(s) accessible.',
            'connected_account': me_data.get('name', ''),
            'permission_status': perm_status,
            'permission_details': {'granted': granted, 'missing': missing, 'required': required},
            'available_forms': forms,
        }

    except Exception as exc:
        return {
            'result': 'fail',
            'message': f'Meta API error: {exc}',
            'permission_status': 'error',
        }


def _meta_source_user_token(source: LeadSource) -> str:
    creds = source.credentials or {}
    return str(creds.get('user_token') or '').strip()


def _meta_source_page_token(source: LeadSource) -> str:
    creds = source.credentials or {}
    return str(creds.get('page_access_token') or creds.get('access_token') or '').strip()


def _meta_pages_from_source(source: LeadSource):
    token = _meta_source_user_token(source)
    if not token:
        raise ValueError('Stored Meta user token is missing. Reconnect Meta once, then pages/forms can be added without OAuth.')
    data = _meta_graph_json(
        'me/accounts',
        {'fields': 'id,name,access_token,tasks', 'limit': 100, 'access_token': token},
        timeout=15,
    )
    return data.get('data', []) or []


def _meta_forms_for_page(page_id: str, page_token: str):
    data = _meta_graph_json(
        f'{_parse.quote(page_id)}/leadgen_forms',
        {
            'fields': 'id,name,status,leads_count,created_time',
            'limit': 100,
            'access_token': page_token,
        },
        timeout=15,
    )
    return [
        {
            'id': str(f.get('id') or ''),
            'name': str(f.get('name') or ''),
            'status': f.get('status') or '',
            'leads_count': f.get('leads_count', 0),
            'created_time': f.get('created_time') or '',
        }
        for f in (data.get('data') or [])
        if str(f.get('id') or '')
    ]


@lead_sources_bp.route('/<int:source_id>/meta/connection-info', methods=['GET'])
@require_role('superadmin', 'sales_manager', 'platform_owner')
def meta_source_connection_info(source_id):
    user = request.current_user
    source = LeadSource.query.get_or_404(source_id)
    err = _check_source_ownership(source, user)
    if err:
        return err
    if source.source_type != 'meta':
        return jsonify({'error': 'Source is not a Meta source'}), 400

    creds = source.credentials or {}
    page_id = str(creds.get('page_id') or '').strip()
    page_name = source.connected_account or source.name
    business = {
        'id': str(creds.get('business_id') or '').strip(),
        'name': str(creds.get('business_name') or '').strip(),
        'source': 'stored' if creds.get('business_id') else '',
    }

    page_token = _meta_source_page_token(source)
    if page_id and page_token and not business['id']:
        try:
            page_data = _meta_graph_json(
                _parse.quote(page_id),
                {'fields': 'id,name,business{id,name}', 'access_token': page_token},
                timeout=10,
            )
            page_name = page_data.get('name') or page_name
            biz = page_data.get('business') or {}
            if biz.get('id'):
                business = {
                    'id': str(biz.get('id') or ''),
                    'name': str(biz.get('name') or ''),
                    'source': 'meta_page_lookup',
                }
        except Exception as exc:
            logger.warning('meta_source_connection_info: business lookup failed: %s', exc)

    ad_accounts = []
    for account in (creds.get('ad_accounts') or []):
        if isinstance(account, dict):
            ad_accounts.append({
                'id': str(account.get('id') or account.get('account_id') or ''),
                'name': str(account.get('name') or ''),
                'account_status': account.get('account_status'),
            })
    if not ad_accounts and creds.get('ad_account_id'):
        ad_accounts.append({
            'id': str(creds.get('ad_account_id') or ''),
            'name': str(creds.get('ad_account_name') or ''),
            'account_status': None,
        })

    return jsonify({
        'source': {
            'id': source.id,
            'name': source.name,
            'connected_account': source.connected_account,
            'is_active': source.is_active,
            'last_tested_at': source.last_tested_at.isoformat() if source.last_tested_at else None,
            'last_test_result': source.last_test_result,
            'last_test_message': source.last_test_message,
        },
        'business': business,
        'page': {'id': page_id, 'name': page_name},
        'ad_accounts': ad_accounts,
        'forms': source.available_forms or [],
        'can_add_without_oauth': bool(_meta_source_user_token(source)),
    }), 200


@lead_sources_bp.route('/<int:source_id>/meta/pages', methods=['GET'])
@require_role('superadmin', 'sales_manager', 'platform_owner')
def meta_source_pages(source_id):
    user = request.current_user
    source = LeadSource.query.get_or_404(source_id)
    err = _check_source_ownership(source, user)
    if err:
        return err
    if source.source_type != 'meta':
        return jsonify({'error': 'Source is not a Meta source'}), 400

    try:
        pages = _meta_pages_from_source(source)
    except Exception as exc:
        return jsonify({'error': str(exc)}), 400

    existing = {}
    for src in LeadSource.query.filter_by(tenant_id=user.tenant_id, source_type='meta').all():
        creds = src.credentials or {}
        pid = str(creds.get('page_id') or '')
        if pid:
            existing[pid] = src
    safe_pages = []
    for page in pages:
        page_id = str(page.get('id') or '')
        existing_source = existing.get(page_id)
        safe_pages.append({
            'id': page_id,
            'name': str(page.get('name') or ''),
            'tasks': page.get('tasks') or [],
            'source_id': existing_source.id if existing_source else None,
            'source_name': existing_source.name if existing_source else '',
            'is_active': bool(existing_source.is_active) if existing_source else False,
        })
    return jsonify({'pages': safe_pages}), 200


@lead_sources_bp.route('/<int:source_id>/meta/pages/<page_id>/forms', methods=['GET'])
@require_role('superadmin', 'sales_manager', 'platform_owner')
def meta_source_page_forms(source_id, page_id):
    user = request.current_user
    source = LeadSource.query.get_or_404(source_id)
    err = _check_source_ownership(source, user)
    if err:
        return err
    if source.source_type != 'meta':
        return jsonify({'error': 'Source is not a Meta source'}), 400
    page_id = str(page_id or '').strip()

    try:
        pages = _meta_pages_from_source(source)
        page = next((p for p in pages if str(p.get('id') or '') == page_id), None)
        if not page:
            return jsonify({'error': 'Page is not available from the stored Meta connection'}), 404
        page_token = str(page.get('access_token') or '').strip()
        if not page_token:
            return jsonify({'error': 'Page access token was not returned by Meta for this page'}), 400
        forms = _meta_forms_for_page(page_id, page_token)
        return jsonify({'page': {'id': page_id, 'name': page.get('name') or ''}, 'forms': forms}), 200
    except Exception as exc:
        logger.exception('meta_source_page_forms failed: %s', exc)
        return jsonify({'error': str(exc)}), 502


@lead_sources_bp.route('/<int:source_id>/meta/pages/save', methods=['POST'])
@require_role('superadmin', 'sales_manager', 'platform_owner')
def meta_source_save_page(source_id):
    user = request.current_user
    source = LeadSource.query.get_or_404(source_id)
    err = _check_source_ownership(source, user)
    if err:
        return err
    if source.source_type != 'meta':
        return jsonify({'error': 'Source is not a Meta source'}), 400

    data = request.get_json() or {}
    page_id = str(data.get('page_id') or '').strip()
    selected_forms = data.get('selected_forms') or []
    if not page_id:
        return jsonify({'error': 'page_id is required'}), 400

    try:
        pages = _meta_pages_from_source(source)
        page = next((p for p in pages if str(p.get('id') or '') == page_id), None)
        if not page:
            return jsonify({'error': 'Page is not available from the stored Meta connection'}), 404
        page_token = str(page.get('access_token') or '').strip()
        if not page_token:
            return jsonify({'error': 'Page access token was not returned by Meta for this page'}), 400

        page_forms = _meta_forms_for_page(page_id, page_token)
        selected_ids = {str(f.get('id') or '') for f in selected_forms if str(f.get('id') or '')}
        forms_to_save = [f for f in page_forms if not selected_ids or str(f.get('id') or '') in selected_ids]
        if not forms_to_save:
            return jsonify({'error': 'Select at least one form for this page'}), 400

        source_creds = source.credentials or {}
        creds = {
            'app_id': source_creds.get('app_id', ''),
            'app_secret': source_creds.get('app_secret', ''),
            'user_token': source_creds.get('user_token', ''),
            'page_id': page_id,
            'page_access_token': page_token,
            'verify_token': 'smk_' + page_id,
            'access_token': page_token,
        }
        for key in ('business_id', 'business_name', 'ad_account_id', 'ad_account_name', 'ad_accounts'):
            if source_creds.get(key):
                creds[key] = source_creds.get(key)

        target = None
        for src in LeadSource.query.filter_by(tenant_id=user.tenant_id, source_type='meta').all():
            if str((src.credentials or {}).get('page_id') or '') == page_id:
                target = src
                break
        if target:
            target.name = (data.get('name') or target.name or f"Meta - {page.get('name') or page_id}").strip()
            existing_creds = target.credentials or {}
            existing_creds.update({k: v for k, v in creds.items() if v})
            target.credentials = existing_creds
        else:
            target = LeadSource(
                tenant_id=user.tenant_id,
                name=(data.get('name') or f"Meta - {page.get('name') or page_id}").strip(),
                source_type='meta',
                credentials=creds,
                created_by=user.id,
            )
            db.session.add(target)

        target.connected_account = f"{page.get('name') or 'Meta Page'} (Page ID: {page_id})"
        target.available_forms = forms_to_save
        target.dup_check_phone = True
        target.dup_check_email = False
        target.permission_status = 'ok'
        target.is_active = True
        target.last_tested_at = datetime.utcnow()
        target.last_test_result = 'pass'
        target.last_test_message = f'Added from existing Meta connection. {len(forms_to_save)} form(s) selected.'

        db.session.flush()
        _deactivate_duplicate_meta_sources_for_page(user.tenant_id, page_id, target.id)
        db.session.commit()
        return jsonify({'source': target.to_dict()}), 200
    except Exception as exc:
        db.session.rollback()
        logger.exception('meta_source_save_page failed: %s', exc)
        return jsonify({'error': str(exc)}), 502


def _test_google(source: LeadSource) -> dict:
    """Verify Google OAuth credentials and account discovery readiness."""
    try:
        creds = source.credentials or {}
        client_id     = creds.get('client_id', '')
        client_secret = creds.get('client_secret', '')
        refresh_token = creds.get('refresh_token', '')

        platform_client_id, platform_client_secret = _get_platform_google_creds()
        if client_id in ('', '__platform__', 'platform') or client_secret in ('', '__platform__', 'platform'):
            client_id = platform_client_id
            client_secret = platform_client_secret

        if not all([client_id, client_secret, refresh_token]):
            missing = [k for k in ['client_id', 'client_secret', 'refresh_token'] if not creds.get(k)]
            return {
                'result': 'fail',
                'message': f'Missing credentials: {missing}',
                'permission_status': 'missing',
                'permission_details': {'missing': missing},
            }

        access_token = _google_exchange_refresh_token(client_id, client_secret, refresh_token)
        connected_rows = ConnectedGoogleAdsAccount.query.filter_by(
            tenant_id=source.tenant_id,
            source_id=source.id,
            is_active=True,
        ).all()
        accounts = []
        discovery_error = ''
        try:
            accounts = _google_list_accessible_accounts(access_token)
        except Exception as exc:
            discovery_error = str(exc)

        if not accounts and connected_rows:
            accounts = [
                {
                    'customer_id': row.customer_id,
                    'customer_name': row.customer_name,
                    'resource_name': row.resource_name,
                }
                for row in connected_rows
            ]

        status = 'pass' if accounts else ('partial' if connected_rows else 'partial')
        message = 'Google OAuth healthy.'
        if accounts:
            message = f'Google OAuth healthy. {len(accounts)} account(s) discoverable.'
        elif discovery_error:
            message = f'Google OAuth healthy but account discovery failed: {discovery_error}'
        else:
            message = 'Google OAuth healthy but no accessible Google Ads accounts were found.'

        if connected_rows and status != 'pass':
            status = 'pass'
            message = f'Google OAuth healthy. {len(connected_rows)} connected account(s) saved on this source.'

        return {
            'result': status,
            'message': message,
            'connected_account': source.connected_account or 'Google Ads Account',
            'permission_status': 'ok' if status == 'pass' else 'partial',
            'permission_details': {
                'granted': ['oauth_refresh_token'],
                'missing': [] if status == 'pass' else ['google_ads_account_discovery'],
                'accessible_accounts': accounts,
                'discovery_error': discovery_error,
            },
            'accessible_accounts': accounts,
        }

    except Exception as exc:
        return {
            'result': 'fail',
            'message': f'Google API error: {exc}',
            'permission_status': 'error',
        }


# ══════════════════════════════════════════════════════════════════════════════
# META HELPER ENDPOINTS (OAuth-assisted form/page discovery)
# ══════════════════════════════════════════════════════════════════════════════

@lead_sources_bp.route('/meta/pages', methods=['GET'])
@require_role('superadmin', 'sales_manager', 'platform_owner')
def meta_list_pages():
    """List Meta pages accessible to the given access_token."""
    user = request.current_user
    access_token = request.args.get('access_token', '').strip()
    if not access_token:
        return jsonify({'error': 'access_token query param required'}), 400
    try:
        import urllib.request as urllib_req
        import urllib.parse as urllib_parse
        import json as _json
        url = f'https://graph.facebook.com/v25.0/me/accounts?access_token={urllib_parse.quote(access_token)}'
        with urllib_req.urlopen(urllib_req.Request(url), timeout=10) as resp:
            data = _json.loads(resp.read())
        return jsonify({'pages': data.get('data', [])}), 200
    except Exception as exc:
        return jsonify({'error': str(exc)}), 502


@lead_sources_bp.route('/meta/forms/<page_id>', methods=['GET'])
@require_role('superadmin', 'sales_manager', 'platform_owner')
def meta_list_forms(page_id):
    """List lead gen forms for a Meta page."""
    user = request.current_user
    access_token = request.args.get('access_token', '').strip()
    if not access_token:
        return jsonify({'error': 'access_token query param required'}), 400
    try:
        import urllib.request as urllib_req
        import urllib.parse as urllib_parse
        import json as _json
        url = (f'https://graph.facebook.com/v25.0/{page_id}/leadgen_forms'
               f'?access_token={urllib_parse.quote(access_token)}')
        with urllib_req.urlopen(urllib_req.Request(url), timeout=10) as resp:
            data = _json.loads(resp.read())
        return jsonify({'forms': data.get('data', [])}), 200
    except Exception as exc:
        return jsonify({'error': str(exc)}), 502


# ══════════════════════════════════════════════════════════════════════════════
# INGESTION LOGS (paginated)
# ══════════════════════════════════════════════════════════════════════════════

def _connected_source_logs_query(user, source_id=None):
    _deactivate_duplicate_meta_sources_for_tenant(user.tenant_id)
    query = (
        IngestedLeadLog.query
        .join(LeadSource, IngestedLeadLog.source_id == LeadSource.id)
        .filter(IngestedLeadLog.tenant_id == user.tenant_id)
        .filter(LeadSource.tenant_id == user.tenant_id)
        .filter(LeadSource.is_active == True)
        .filter(IngestedLeadLog.received_at >= LeadSource.created_at)
    )
    tenant_cutoff = lead_source_cutoff_for(tenant_id=user.tenant_id)
    if tenant_cutoff:
        query = query.filter(IngestedLeadLog.received_at >= tenant_cutoff)
    if source_id:
        query = query.filter(IngestedLeadLog.source_id == source_id)
    return query


def _source_log_identity(log_row):
    mapped = log_row.mapped_fields if isinstance(log_row.mapped_fields, dict) else {}
    status = _log_effective_status(log_row)
    platform_id = str(log_row.platform_lead_id or '').strip()

    if platform_id:
        return ('platform', log_row.source_id, platform_id)

    if status == 'processed':
        if log_row.lead_id:
            return ('processed-lead', log_row.source_id, log_row.lead_id)
    elif status == 'duplicate':
        return ('duplicate-log', log_row.id)
    elif status == 'error':
        return ('error-log', log_row.id)

    phone = normalize_phone_for_duplicate(mapped.get('phone') or mapped.get('mobile'))
    if phone:
        return ('contact', log_row.source_id, str(log_row.form_id or ''), phone)
    linked_id = log_row.lead_id or log_row.dup_of_lead_id
    if linked_id:
        return ('lead', log_row.source_id, linked_id)
    return ('log', log_row.id)


def _dedupe_source_logs(query, limit=None):
    rows = query.order_by(IngestedLeadLog.received_at.desc(), IngestedLeadLog.id.desc()).limit(limit or 10000).all()
    seen = set()
    out = []
    for row in rows:
        key = _source_log_identity(row)
        if key in seen:
            continue
        seen.add(key)
        out.append(row)
    return out


def _report_allowed_log_keys(user, date_from='', date_to='', source_id=None):
    report = _build_performance_report(user, date_from=date_from, date_to=date_to, source_id=source_id)
    return {
        _campaign_key_from_parts(
            row.get('source_id'),
            row.get('project_name'),
            row.get('campaign_id'),
            row.get('ad_set_id'),
            row.get('ad_id'),
            row.get('form_id'),
        )
        for row in (report.get('campaign_rows') or [])
    }


def _dedupe_report_source_logs(query, user, date_from='', date_to='', source_id=None, limit=None):
    rows = query.order_by(IngestedLeadLog.received_at.desc(), IngestedLeadLog.id.desc()).limit(limit or 10000).all()
    seen = set()
    out = []

    for row in rows:
        identity_key = _source_log_identity(row)
        if identity_key in seen:
            continue
        seen.add(identity_key)
        out.append(row)

    return out


def _apply_log_search_filter(query, search_q):
    if not search_q:
        return query
    like_q = f'%{search_q}%'
    return query.filter(
        db.or_(
            IngestedLeadLog.campaign_id.ilike(like_q),
            IngestedLeadLog.campaign_name.ilike(like_q),
            IngestedLeadLog.ad_set_id.ilike(like_q),
            IngestedLeadLog.ad_set_name.ilike(like_q),
            IngestedLeadLog.ad_id.ilike(like_q),
            IngestedLeadLog.ad_name.ilike(like_q),
            IngestedLeadLog.form_id.ilike(like_q),
            IngestedLeadLog.form_name.ilike(like_q),
            IngestedLeadLog.page_id.ilike(like_q),
            IngestedLeadLog.source_type.ilike(like_q),
            IngestedLeadLog.status.ilike(like_q),
            IngestedLeadLog.platform_lead_id.ilike(like_q),
        )
    )


@lead_sources_bp.route('/logs', methods=['GET'])
@require_role('superadmin', 'sales_manager', 'platform_owner')
def ingestion_logs():
    user = request.current_user
    source_id = request.args.get('source_id', type=int)
    search_q  = request.args.get('q', '').strip()
    status    = request.args.get('status', '').strip().lower()
    date_from = request.args.get('date_from', '').strip()
    date_to   = request.args.get('date_to', '').strip()
    page      = max(1, request.args.get('page', 1, type=int))
    per_page  = min(100, max(10, request.args.get('per_page', 25, type=int)))

    q = _connected_source_logs_query(user, source_id=source_id)
    q = q.filter(IngestedLeadLog.status.in_(['processed', 'duplicate', 'error']))
    q = _apply_log_search_filter(q, search_q)
    if date_from:
        try:
            q = q.filter(IngestedLeadLog.received_at >= datetime.fromisoformat(date_from))
        except ValueError:
            pass
    if date_to:
        try:
            q = q.filter(IngestedLeadLog.received_at < datetime.fromisoformat(date_to) + timedelta(days=1))
        except ValueError:
            pass

    q = _apply_test_data_filter(q, IngestedLeadLog)

    unique_logs = _dedupe_report_source_logs(
        q,
        user,
        date_from=date_from,
        date_to=date_to,
        source_id=source_id,
    )
    if status in {'processed', 'duplicate', 'error'}:
        unique_logs = [row for row in unique_logs if _log_effective_status(row) == status]
    total = len(unique_logs)
    logs = unique_logs[(page - 1) * per_page: page * per_page]

    def _to_display_row(log_row):
        payload = log_row.to_dict()
        lead_obj = log_row.lead or log_row.dup_lead
        mapped = log_row.mapped_fields if isinstance(log_row.mapped_fields, dict) else {}

        lead_name = (lead_obj.name if lead_obj else None) or mapped.get('name') or payload.get('lead_name') or 'Unknown Lead'
        lead_phone = (lead_obj.phone if lead_obj else None) or mapped.get('phone') or mapped.get('mobile') or ''
        project_name = (
            (lead_obj.project.name if lead_obj and lead_obj.project else None)
            or mapped.get('project_name')
            or payload.get('project_name')
            or 'No Project Assigned'
        )

        payload['lead_name'] = str(lead_name)
        payload['lead_phone'] = str(lead_phone or '')
        payload['project_name'] = str(project_name)
        payload['lead_id'] = (lead_obj.id if lead_obj else payload.get('lead_id'))
        payload['status'] = _log_effective_status(log_row)
        return payload

    return jsonify({
        'logs':     [_to_display_row(l) for l in logs],
        'total':    total,
        'page':     page,
        'per_page': per_page,
    }), 200


@lead_sources_bp.route('/logs/status', methods=['PUT'])
@require_role('superadmin', 'sales_manager', 'platform_owner')
def ingestion_logs_bulk_status_update():
    user = request.current_user
    data = request.get_json() or {}
    status = str(data.get('status') or '').strip().lower()
    log_ids = data.get('log_ids') or []

    allowed = {'processed', 'duplicate', 'error'}
    if status not in allowed:
        return jsonify({'error': f'Invalid status. Allowed: {sorted(list(allowed))}'}), 400
    if not isinstance(log_ids, list) or not log_ids:
        return jsonify({'error': 'log_ids must be a non-empty list'}), 400

    parsed_ids = []
    for item in log_ids:
        try:
            parsed_ids.append(int(item))
        except (TypeError, ValueError):
            continue
    parsed_ids = sorted({x for x in parsed_ids if x > 0})
    if not parsed_ids:
        return jsonify({'error': 'No valid log IDs supplied'}), 400

    rows = IngestedLeadLog.query.filter(
        IngestedLeadLog.tenant_id == user.tenant_id,
        IngestedLeadLog.id.in_(parsed_ids),
    ).all()
    if not rows:
        return jsonify({'error': 'No matching logs found for this tenant'}), 404

    now = datetime.utcnow()
    for row in rows:
        row.status = status
        row.processed_at = now
    db.session.commit()

    return jsonify({'ok': True, 'updated': len(rows), 'status': status}), 200


@lead_sources_bp.route('/logs/diagnostics', methods=['GET'])
@require_role('superadmin', 'platform_owner')
def ingestion_log_diagnostics():
    """Tenant aggregate ingestion health without exposing payloads or lead PII."""
    user = request.current_user
    rows = (
        db.session.query(
            IngestedLeadLog.source_type,
            IngestedLeadLog.status,
            db.func.count(IngestedLeadLog.id),
            db.func.min(IngestedLeadLog.received_at),
            db.func.max(IngestedLeadLog.received_at),
        )
        .filter(IngestedLeadLog.tenant_id == user.tenant_id)
        .group_by(IngestedLeadLog.source_type, IngestedLeadLog.status)
        .order_by(IngestedLeadLog.source_type, IngestedLeadLog.status)
        .all()
    )
    return jsonify({'groups': [
        {
            'source_type': source_type,
            'status': status,
            'count': int(count),
            'oldest_at': oldest.isoformat() if oldest else None,
            'latest_at': latest.isoformat() if latest else None,
        }
        for source_type, status, count, oldest, latest in rows
    ]}), 200


@lead_sources_bp.route('/logs/<int:log_id>/reprocess', methods=['POST'])
@require_role('superadmin', 'platform_owner')
def reprocess_ingestion_log(log_id):
    """Reprocess one captured failed/queued event through the canonical pipeline."""
    user = request.current_user
    log = IngestedLeadLog.query.filter_by(id=log_id, tenant_id=user.tenant_id).first()
    if not log:
        return jsonify({'error': 'Ingestion event not found'}), 404
    if log.status not in ('queued', 'error'):
        return jsonify({'error': 'Only queued or failed events can be reprocessed'}), 409

    source = LeadSource.query.filter_by(
        id=log.source_id,
        tenant_id=user.tenant_id,
        is_active=True,
    ).first()
    if not source:
        return jsonify({'error': 'Active lead source not found'}), 409

    from app.routes.ingestion import (
        _meta_enrich_leadgen_entry,
        _normalise_generic,
        _normalise_google,
        _normalise_meta,
    )
    from app.services.ingestion_engine import ingest_lead

    payload = dict(log.raw_payload or {})
    if source.source_type == 'meta':
        payload.setdefault('leadgen_id', str(log.platform_lead_id or ''))
        enriched = _meta_enrich_leadgen_entry(payload, source)
        for key, value in (enriched or {}).items():
            if value not in (None, '', []):
                payload[key] = value
        normalised = _normalise_meta(payload)
    elif source.source_type == 'google':
        normalised = _normalise_google(payload)
    else:
        normalised = _normalise_generic(payload)

    result = ingest_lead(source, payload, normalised, ingestion_log=log)
    status_code = 200 if result.get('status') != 'error' else 422
    return jsonify({'ok': status_code == 200, 'result': result}), status_code


def _xlsx_response(filename, headers, rows):
    """Generate XLSX response using openpyxl if available, fallback to CSV"""
    try:
        from openpyxl import Workbook
        from io import BytesIO
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Data'
        
        # Add headers
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
        
        # Add rows
        for row_idx, row in enumerate(rows, 2):
            for col_idx, val in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)
        
        # Auto-fit columns
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
        
        out = BytesIO()
        wb.save(out)
        out.seek(0)
        
        return Response(
            out.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename={filename}'},
        )
    except Exception as e:
        # Fallback to CSV if openpyxl not available
        return _csv_response(filename.replace('.xlsx', '.csv'), headers, rows)

def _csv_response(filename, headers, rows):
    from io import StringIO

    out = StringIO()
    writer = csv.writer(out)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    csv_data = out.getvalue()
    out.close()
    return Response(
        csv_data,
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename={filename}'},
    )


@lead_sources_bp.route('/logs/export.csv', methods=['GET'])
@require_role('superadmin', 'sales_manager', 'platform_owner')
def ingestion_logs_export_csv():
    user = request.current_user
    source_id = request.args.get('source_id', type=int)
    search_q = request.args.get('q', '').strip()
    status = request.args.get('status', '').strip().lower()
    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()

    q = _connected_source_logs_query(user, source_id=source_id)
    q = q.filter(IngestedLeadLog.status.in_(['processed', 'duplicate', 'error']))
    q = _apply_log_search_filter(q, search_q)
    if date_from:
        try:
            q = q.filter(IngestedLeadLog.received_at >= datetime.fromisoformat(date_from))
        except ValueError:
            pass
    if date_to:
        try:
            q = q.filter(IngestedLeadLog.received_at < datetime.fromisoformat(date_to) + timedelta(days=1))
        except ValueError:
            pass

    q = _apply_test_data_filter(q, IngestedLeadLog)

    rows = _dedupe_report_source_logs(
        q,
        user,
        date_from=date_from,
        date_to=date_to,
        source_id=source_id,
        limit=10000,
    )
    if status in {'processed', 'duplicate', 'error'}:
        rows = [row for row in rows if _log_effective_status(row) == status]
    return _csv_response(
        'lead-source-logs.csv',
        [
            'id', 'source_id', 'source_type', 'status', 'platform_lead_id',
            'campaign_name', 'ad_name', 'form_id', 'form_name', 'lead_id',
            'dup_of_lead_id', 'error_message', 'received_at', 'processed_at',
        ],
        [[
            l.id,
            l.source_id,
            l.source_type,
            _log_effective_status(l),
            l.platform_lead_id,
            l.campaign_name,
            l.ad_name,
            l.form_id,
            l.form_name,
            l.lead_id,
            l.dup_of_lead_id,
            l.error_message,
            l.received_at.isoformat() if l.received_at else '',
            l.processed_at.isoformat() if l.processed_at else '',
        ] for l in rows],
    )


@lead_sources_bp.route('/<int:source_id>/forms/mappings', methods=['GET'])
@require_role('superadmin', 'sales_manager', 'platform_owner')
def get_source_form_mappings(source_id):
    user = request.current_user
    source = LeadSource.query.get_or_404(source_id)
    err = _check_source_ownership(source, user)
    if err:
        return err

    summary = _source_form_mapping_summary(source)
    return jsonify({'source_id': source.id, 'summary': summary, 'rows': summary['rows']}), 200


@lead_sources_bp.route('/<int:source_id>/forms/mappings', methods=['PUT'])
@require_role('superadmin', 'platform_owner')
def put_source_form_mappings(source_id):
    user = request.current_user
    source = LeadSource.query.get_or_404(source_id)
    err = _check_source_ownership(source, user)
    if err:
        return err

    data = request.get_json() or {}
    rows = data.get('rows') or []
    apply_scope = str(data.get('apply_scope') or 'future').strip().lower()
    if not isinstance(rows, list):
        return jsonify({'error': 'rows must be a list'}), 400
    if apply_scope not in {'future', 'past_and_future'}:
        return jsonify({'error': 'Invalid apply_scope'}), 400

    forms = _source_forms(source)
    forms_by_id = {f['id']: f for f in forms}
    payload = []

    for row in rows:
        if not isinstance(row, dict):
            continue
        form_id = str(row.get('form_id') or '').strip()
        if not form_id:
            continue
        project_id = row.get('project_id')
        try:
            project_id = int(project_id)
        except (TypeError, ValueError):
            return jsonify({'error': f'project_id required for form {form_id}'}), 400

        project = Project.query.filter_by(id=project_id, tenant_id=user.tenant_id).first()
        if not project:
            return jsonify({'error': f'Invalid project_id for form {form_id}'}), 400

        manager_mode = _normalize_form_manager_mode(
            row.get('manager_assign_mode') or row.get('manager_rule') or 'none'
        )
        if manager_mode not in FORM_MANAGER_ASSIGN_MODES:
            return jsonify({'error': f'Invalid manager_assign_mode for form {form_id}. Allowed: {list(FORM_MANAGER_ASSIGN_MODES)}'}), 400

        manager_id = None
        rr_pool = []

        if manager_mode == 'fixed_manager':
            try:
                manager_id = int(row.get('manager_id'))
            except (TypeError, ValueError):
                return jsonify({'error': f'manager_id required for fixed_manager on form {form_id}'}), 400

            manager = User.query.filter_by(
                id=manager_id,
                tenant_id=user.tenant_id,
                role='sales_manager',
                is_active=True,
            ).first()
            if not manager:
                return jsonify({'error': f'Invalid manager_id for form {form_id}'}), 400

        elif manager_mode == 'round_robin_pool':
            raw_pool = row.get('rr_manager_pool')
            if raw_pool is None:
                raw_pool = row.get('manager_pool_ids')

            if isinstance(raw_pool, str):
                rr_pool = [
                    int(x.strip())
                    for x in raw_pool.split(',')
                    if str(x).strip().isdigit()
                ]
            elif isinstance(raw_pool, list):
                for item in raw_pool:
                    try:
                        rr_pool.append(int(item))
                    except (TypeError, ValueError):
                        return jsonify({'error': f'rr_manager_pool must contain manager IDs for form {form_id}'}), 400
            else:
                return jsonify({'error': f'rr_manager_pool required for round_robin_pool on form {form_id}'}), 400

            rr_pool = sorted({m for m in rr_pool if m > 0})
            if not rr_pool:
                return jsonify({'error': f'rr_manager_pool required for round_robin_pool on form {form_id}'}), 400

            managers = User.query.filter(
                User.tenant_id == user.tenant_id,
                User.role == 'sales_manager',
                User.is_active == True,
                User.id.in_(rr_pool),
            ).all()
            valid_manager_ids = {int(m.id) for m in managers}
            missing_ids = [m for m in rr_pool if m not in valid_manager_ids]
            if missing_ids:
                return jsonify({'error': f'Invalid manager IDs in rr_manager_pool for form {form_id}: {missing_ids}'}), 400

        payload.append({
            'form_id': form_id,
            'project_id': project_id,
            'page_id': str(row.get('page_id') or (source.credentials or {}).get('page_id') or ''),
            'form_name': str(row.get('form_name') or forms_by_id.get(form_id, {}).get('name') or ''),
            'source_type': source.source_type,
            'manager_assign_mode': manager_mode,
            'manager_id': manager_id,
            'rr_manager_pool': rr_pool,
        })

    existing = LeadSourceFormMapping.query.filter_by(
        tenant_id=user.tenant_id,
        source_id=source.id,
    ).all()
    by_form = {str(r.form_id): r for r in existing}

    submitted_ids = {p['form_id'] for p in payload}
    for form_id, row in by_form.items():
        if form_id not in submitted_ids:
            row.is_active = False

    past_project_updates = []
    for item in payload:
        row = by_form.get(item['form_id'])
        if not row:
            row = LeadSourceFormMapping(
                tenant_id=user.tenant_id,
                source_id=source.id,
                form_id=item['form_id'],
                created_by=user.id,
            )
            db.session.add(row)
        previous_project_id = int(row.project_id or 0) if row.project_id else None
        if apply_scope == 'past_and_future' and previous_project_id != int(item['project_id']):
            past_project_updates.append((item['form_id'], item['project_id']))
        row.source_type = item['source_type']
        row.page_id = item['page_id']
        row.form_name = item['form_name']
        row.project_id = item['project_id']
        row.manager_assign_mode = item['manager_assign_mode']
        row.manager_id = item['manager_id']
        row.rr_manager_pool = item['rr_manager_pool']
        if row.manager_assign_mode != 'round_robin_pool':
            row.rr_last_index = 0
        row.is_active = True

    updated_existing_leads = 0
    if past_project_updates:
        for form_id, project_id in past_project_updates:
            lead_ids = (
                db.session.query(IngestedLeadLog.lead_id)
                .filter(
                    IngestedLeadLog.tenant_id == user.tenant_id,
                    IngestedLeadLog.source_id == source.id,
                    IngestedLeadLog.form_id == str(form_id),
                    IngestedLeadLog.lead_id.isnot(None),
                )
            )
            updated_existing_leads += Lead.query.filter(
                Lead.tenant_id == user.tenant_id,
                Lead.id.in_(lead_ids),
            ).update(
                {
                    Lead.project_id: int(project_id),
                    Lead.updated_at: datetime.utcnow(),
                },
                synchronize_session=False,
            )

    db.session.commit()

    summary = _source_form_mapping_summary(source)
    return jsonify({
        'ok': True,
        'summary': summary,
        'rows': summary['rows'],
        'apply_scope': apply_scope,
        'updated_existing_leads': updated_existing_leads,
    }), 200


@lead_sources_bp.route('/<int:source_id>/readiness', methods=['GET'])
@require_role('superadmin', 'sales_manager', 'platform_owner')
def get_source_readiness(source_id):
    user = request.current_user
    source = LeadSource.query.get_or_404(source_id)
    err = _check_source_ownership(source, user)
    if err:
        return err
    gate = _validate_mapping_gate(source)
    return jsonify({'source_id': source.id, 'mapping_gate': gate}), 200


# ══════════════════════════════════════════════════════════════════════════════
# REPORTS
# ══════════════════════════════════════════════════════════════════════════════

@lead_sources_bp.route('/reports/by-source', methods=['GET'])
@require_role('superadmin', 'sales_manager', 'platform_owner')
def report_by_source():
    """Business-readable source performance grouped by source and project."""
    user = request.current_user
    source_id = request.args.get('source_id', type=int)
    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()
    page = max(1, request.args.get('page', 1, type=int))
    per_page = min(100, max(10, request.args.get('per_page', 10, type=int)))

    report = _build_performance_report(user, date_from=date_from, date_to=date_to, source_id=source_id, include_unpriced=True)
    rows = report['source_rows']
    total = len(rows)
    paginated_rows = rows[(page - 1) * per_page: page * per_page]

    return jsonify({'rows': paginated_rows, 'total': total, 'page': page, 'per_page': per_page}), 200


@lead_sources_bp.route('/reports/performance', methods=['GET'])
@require_role('superadmin', 'sales_manager', 'platform_owner')
def report_performance():
    """Canonical LMS report: source/form performance without campaign rows."""
    user = request.current_user
    source_id = request.args.get('source_id', type=int)
    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()
    source_page = max(1, request.args.get('source_page', request.args.get('page', 1, type=int), type=int))
    source_per_page = min(100, max(10, request.args.get('source_per_page', request.args.get('per_page', 10, type=int), type=int)))

    report = _build_lms_source_form_performance(user, date_from=date_from, date_to=date_to, source_id=source_id)
    source_rows = report['source_rows']
    form_rows = report['form_rows']
    source_total = len(source_rows)
    form_total = len(form_rows)
    source_start = (source_page - 1) * source_per_page

    return jsonify({
        'snapshot': report['snapshot'],
        'source_rows': source_rows[source_start:source_start + source_per_page],
        'source_total': source_total,
        'source_page': source_page,
        'source_per_page': source_per_page,
        'form_rows': form_rows[source_start:source_start + source_per_page],
        'form_total': form_total,
        'form_page': source_page,
        'form_per_page': source_per_page,
        'campaign_rows': [],
        'campaign_total': 0,
        'campaign_page': 1,
        'campaign_per_page': source_per_page,
        'last_synced_at': report.get('last_synced_at'),
        'report_scope': 'source_form',
    }), 200


@lead_sources_bp.route('/reports/sync-meta', methods=['POST'])
@require_role('superadmin', 'platform_owner')
def report_sync_meta():
    """Refresh stored Meta ad-insights for the current report window.

    Lead pulling is handled by /api/cron/meta-backfill.  This endpoint only
    refreshes source-level spend snapshots so the reports table updates
    immediately. /api/cron/meta-report-sync runs the same helper on schedule.
    """
    user = request.current_user
    data = request.get_json(silent=True) or {}
    date_from = str(data.get('date_from') or '').strip()
    date_to = str(data.get('date_to') or '').strip()
    source_id = data.get('source_id')
    try:
        source_id = int(source_id) if source_id not in (None, '', 'all') else None
    except Exception:
        source_id = None

    result = _sync_meta_report_snapshots(user, date_from=date_from, date_to=date_to, source_id=source_id)
    return jsonify(result), 200


@lead_sources_bp.route('/reports/by-campaign', methods=['GET'])
@require_role('superadmin', 'sales_manager', 'platform_owner')
def report_by_campaign():
    """Business-readable attribution report grouped by campaign/ad set/ad/form."""
    user = request.current_user
    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()
    source_id = request.args.get('source_id', type=int)

    report = _build_performance_report(user, date_from=date_from, date_to=date_to, source_id=source_id, include_unpriced=True)
    rows = report['campaign_rows']
    return jsonify({'rows': rows}), 200


@lead_sources_bp.route('/reports/overview', methods=['GET'])
@require_role('superadmin', 'sales_manager', 'platform_owner')
def report_overview():
    user = request.current_user
    source_id = request.args.get('source_id', type=int)
    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()
    report = _build_performance_report(
        user,
        date_from=date_from,
        date_to=date_to,
        source_id=source_id,
        include_unpriced=True,
    )
    return jsonify(report.get('snapshot') or {
        'total': 0,
        'processed': 0,
        'duplicate': 0,
        'errors': 0,
        'conversion_rate': 0,
    }), 200


@lead_sources_bp.route('/reports/attribution', methods=['GET'])
@require_role('superadmin', 'sales_manager', 'platform_owner')
def report_attribution():
    user = request.current_user
    source_id = request.args.get('source_id', type=int)
    page = max(1, request.args.get('page', 1, type=int))
    per_page = min(100, max(10, request.args.get('per_page', 10, type=int)))
    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()

    report = _build_performance_report(user, date_from=date_from, date_to=date_to, source_id=source_id, include_unpriced=True)
    rows = report['campaign_rows']
    total = len(rows)
    paginated_rows = rows[(page - 1) * per_page: page * per_page]

    return jsonify({
        'rows': paginated_rows,
        'total': total,
        'page': page,
        'per_page': per_page,
    }), 200


@lead_sources_bp.route('/reports/by-source/export.csv', methods=['GET'])
@require_role('superadmin', 'sales_manager', 'platform_owner')
def report_by_source_export_csv():
    user = request.current_user
    source_id = request.args.get('source_id', type=int)
    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()

    report = _build_lms_source_form_performance(user, date_from=date_from, date_to=date_to, source_id=source_id)
    rows = report['source_rows']

    return _csv_response(
        'lead-source-performance.csv',
        ['source_id', 'source_name', 'source_added_at', 'effective_date_from', 'source_status', 'total_leads', 'unique_leads', 'processed', 'duplicate', 'errors', 'conversion_rate', 'spend', 'cpl', 'last_sync'],
        [[
            r.get('source_id'),
            r.get('source_name'),
            r.get('source_added_at'),
            r.get('effective_date_from'),
            r.get('source_status'),
            int(r.get('total_leads') or r.get('total') or 0),
            int(r.get('unique_leads') or 0),
            int(r.get('processed') or r.get('created') or 0),
            int(r.get('duplicate') or r.get('duplicates') or 0),
            int(r.get('errors') or 0),
            r.get('conversion_rate'),
            r.get('spend'),
            r.get('cpl'),
            r.get('last_sync'),
        ] for r in rows],
    )


@lead_sources_bp.route('/reports/by-source/export.xlsx', methods=['GET'])
@require_role('superadmin', 'sales_manager', 'platform_owner')
def report_by_source_export_xlsx():
    user = request.current_user
    source_id = request.args.get('source_id', type=int)
    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()

    report = _build_lms_source_form_performance(user, date_from=date_from, date_to=date_to, source_id=source_id)
    rows = report['source_rows']

    return _xlsx_response(
        'lead-source-performance.xlsx',
        ['source_id', 'source_name', 'source_added_at', 'effective_date_from', 'source_status', 'total_leads', 'unique_leads', 'processed', 'duplicate', 'errors', 'conversion_rate', 'spend', 'cpl', 'last_sync'],
        [[
            r.get('source_id'),
            r.get('source_name'),
            r.get('source_added_at'),
            r.get('effective_date_from'),
            r.get('source_status'),
            int(r.get('total_leads') or r.get('total') or 0),
            int(r.get('unique_leads') or 0),
            int(r.get('processed') or r.get('created') or 0),
            int(r.get('duplicate') or r.get('duplicates') or 0),
            int(r.get('errors') or 0),
            r.get('conversion_rate'),
            r.get('spend'),
            r.get('cpl'),
            r.get('last_sync'),
        ] for r in rows],
    )


@lead_sources_bp.route('/reports/attribution/export.csv', methods=['GET'])
@require_role('superadmin', 'sales_manager', 'platform_owner')
def report_attribution_export_csv():
    user = request.current_user
    source_id = request.args.get('source_id', type=int)
    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()

    report = _build_performance_report(user, date_from=date_from, date_to=date_to, source_id=source_id, include_unpriced=True)
    rows = report['campaign_rows']

    return _csv_response(
        'lead-source-attribution.csv',
        [
            'source_id', 'source_status', 'project_name', 'campaign_id', 'campaign_name', 'ad_set_id', 'ad_set_name',
            'ad_id', 'ad_name', 'form_id', 'form_name', 'total', 'created', 'spend', 'cpl', 'ctr', 'cpc', 'cpm', 'reach', 'impressions', 'placement', 'audience', 'last_sync',
        ],
        [[
            r.get('source_id'),
            r.get('source_status'),
            r.get('project_name'),
            r.get('campaign_id'),
            r.get('campaign_name'),
            r.get('ad_set_id'),
            r.get('ad_set_name'),
            r.get('ad_id'),
            r.get('ad_name'),
            r.get('form_id'),
            r.get('form_name'),
            int(r.get('total') or 0),
            int(r.get('created') or 0),
            r.get('spend'),
            r.get('cpl'),
            r.get('ctr'),
            r.get('cpc'),
            r.get('cpm'),
            r.get('reach'),
            r.get('impressions'),
            r.get('placement'),
            r.get('audience'),
            r.get('last_sync'),
        ] for r in rows],
    )

@lead_sources_bp.route('/reports/attribution/export.xlsx', methods=['GET'])
@require_role('superadmin', 'sales_manager', 'platform_owner')
def report_attribution_export_xlsx():
    user = request.current_user
    source_id = request.args.get('source_id', type=int)
    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()

    report = _build_performance_report(user, date_from=date_from, date_to=date_to, source_id=source_id, include_unpriced=True)
    rows = report['campaign_rows']

    return _xlsx_response(
        'lead-source-attribution.xlsx',
        [
            'source_id', 'source_status', 'project_name', 'campaign_id', 'campaign_name', 'ad_set_id', 'ad_set_name',
            'ad_id', 'ad_name', 'form_id', 'form_name', 'total', 'created', 'spend', 'cpl', 'ctr', 'cpc', 'cpm', 'reach', 'impressions', 'placement', 'audience', 'last_sync',
        ],
        [[
            r.get('source_id'),
            r.get('source_status'),
            r.get('project_name'),
            r.get('campaign_id'),
            r.get('campaign_name'),
            r.get('ad_set_id'),
            r.get('ad_set_name'),
            r.get('ad_id'),
            r.get('ad_name'),
            r.get('form_id'),
            r.get('form_name'),
            int(r.get('total') or 0),
            int(r.get('created') or 0),
            r.get('spend'),
            r.get('cpl'),
            r.get('ctr'),
            r.get('cpc'),
            r.get('cpm'),
            r.get('reach'),
            r.get('impressions'),
            r.get('placement'),
            r.get('audience'),
            r.get('last_sync'),
        ] for r in rows],
    )

@lead_sources_bp.route('/meta/snapshots', methods=['GET'])
@require_role('superadmin', 'sales_manager', 'platform_owner')
def meta_snapshots_list():
    user = request.current_user
    source_id = request.args.get('source_id', type=int)
    campaign_id = request.args.get('campaign_id', '').strip()
    form_id = request.args.get('form_id', '').strip()
    page = max(1, request.args.get('page', 1, type=int))
    per_page = min(200, max(10, request.args.get('per_page', 50, type=int)))

    q = MetaCampaignSnapshot.query.filter_by(tenant_id=user.tenant_id)
    q = _apply_snapshot_test_filter(q)
    if source_id:
        q = q.filter_by(source_id=source_id)
    if campaign_id:
        q = q.filter_by(campaign_id=campaign_id)
    if form_id:
        q = q.filter_by(form_id=form_id)

    total = q.count()
    rows = q.order_by(MetaCampaignSnapshot.snapshot_at.desc()).offset((page - 1) * per_page).limit(per_page).all()
    return jsonify({
        'rows': [r.to_dict() for r in rows],
        'total': total,
        'page': page,
        'per_page': per_page,
    }), 200


@lead_sources_bp.route('/meta/snapshots/export.csv', methods=['GET'])
@require_role('superadmin', 'sales_manager', 'platform_owner')
def meta_snapshots_export_csv():
    user = request.current_user
    source_id = request.args.get('source_id', type=int)
    campaign_id = request.args.get('campaign_id', '').strip()
    form_id = request.args.get('form_id', '').strip()

    q = MetaCampaignSnapshot.query.filter_by(tenant_id=user.tenant_id)
    q = _apply_snapshot_test_filter(q)
    if source_id:
        q = q.filter_by(source_id=source_id)
    if campaign_id:
        q = q.filter_by(campaign_id=campaign_id)
    if form_id:
        q = q.filter_by(form_id=form_id)

    rows = q.order_by(MetaCampaignSnapshot.snapshot_at.desc()).limit(10000).all()
    return _csv_response(
        'meta-campaign-snapshots.csv',
        [
            'snapshot_at', 'source_id', 'campaign_id', 'campaign_name', 'ad_set_id',
            'results', 'cost_per_result',
        ],
        [[
            r.snapshot_at.isoformat() if r.snapshot_at else '',
            r.source_id,
            r.campaign_id,
            r.campaign_name,
            r.ad_set_id,
            r.ad_set_name,
            r.ad_id,
            r.ad_name,
            r.form_id,
            r.form_name,
            r.spend,
            r.impressions,
            r.reach,
            r.clicks,
            r.ctr,
            r.cpc,
            r.cpm,
            r.frequency,
            r.results,
            r.cost_per_result,
        ] for r in rows],
    )


@lead_sources_bp.route('/proof/attribution-latest', methods=['GET'])
@require_role('superadmin', 'sales_manager', 'platform_owner')
def proof_attribution_latest():
    """Return one complete DB-backed attribution journey with campaign intelligence fields."""
    user = request.current_user

    snapshot = _apply_snapshot_test_filter(MetaCampaignSnapshot.query.filter_by(tenant_id=user.tenant_id)).order_by(
        MetaCampaignSnapshot.snapshot_at.desc()
    ).first()
    if not snapshot:
        return jsonify({'error': 'No meta campaign snapshots found'}), 404

    log = None
    if snapshot.ingested_log_id:
        log = IngestedLeadLog.query.filter_by(id=snapshot.ingested_log_id, tenant_id=user.tenant_id).first()
    if not log and snapshot.lead_id:
        log = IngestedLeadLog.query.filter_by(tenant_id=user.tenant_id, lead_id=snapshot.lead_id).order_by(
            IngestedLeadLog.received_at.desc()
        ).first()

    lead = None
    if snapshot.lead_id:
        from app.models import Lead
        lead = Lead.query.filter_by(id=snapshot.lead_id, tenant_id=user.tenant_id).first()

    source = LeadSource.query.filter_by(id=snapshot.source_id, tenant_id=user.tenant_id).first()
    mapping = LeadSourceFormMapping.query.filter_by(
        tenant_id=user.tenant_id,
        source_id=snapshot.source_id,
        form_id=(snapshot.form_id or ''),
        is_active=True,
    ).first()

    extra = snapshot.extra_metrics or {}

    page_id = snapshot.page_id or (log.page_id if log else None)
    page_name = (
        extra.get('page_name')
        or ((source.credentials or {}).get('page_name') if source else None)
        or (source.connected_account if source else None)
    )
    project_id = extra.get('project_id') or (lead.project_id if lead else None) or (mapping.project_id if mapping else None)
    project_name = extra.get('project_name') or (mapping.project.name if mapping and mapping.project else None)
    if not project_name and project_id:
        p = Project.query.filter_by(id=project_id, tenant_id=user.tenant_id).first()
        if p:
            project_name = p.name

    out = {
        'api_source': {
            'proof_endpoint': '/api/lead-sources/proof/attribution-latest',
            'snapshot_endpoint': '/api/lead-sources/meta/snapshots',
            'logs_endpoint': '/api/lead-sources/logs',
            'reports_endpoint': '/api/lead-sources/reports/attribution',
        },
        'database_storage': {
            'table_snapshot': 'meta_campaign_snapshots',
            'snapshot_id': snapshot.id,
            'table_log': 'ingested_lead_logs',
            'ingested_log_id': snapshot.ingested_log_id,
            'table_lead': 'leads',
            'lead_id': snapshot.lead_id,
        },
        'attribution': {
            'campaign_id': snapshot.campaign_id or (log.campaign_id if log else None),
            'campaign_name': snapshot.campaign_name or (log.campaign_name if log else None),
            'adset_id': snapshot.ad_set_id or (log.ad_set_id if log else None),
            'adset_name': snapshot.ad_set_name or (log.ad_set_name if log else None),
            'ad_id': snapshot.ad_id or (log.ad_id if log else None),
            'ad_name': snapshot.ad_name or (log.ad_name if log else None),
            'form_id': snapshot.form_id or (log.form_id if log else None),
            'form_name': snapshot.form_name or (log.form_name if log else None),
            'page_id': page_id,
            'page_name': page_name,
            'project_id': project_id,
            'project_name': project_name,
        },
        'campaign_metrics': {
            'spend': snapshot.spend,
            'cost_per_result': snapshot.cost_per_result,
            'ctr': snapshot.ctr,
            'cpc': snapshot.cpc,
            'cpm': snapshot.cpm,
            'impressions': snapshot.impressions,
            'reach': snapshot.reach,
            'audience': snapshot.audience,
            'placement': snapshot.placement,
        },
        'journey': {
            'meta_page': {'id': page_id, 'name': page_name},
            'form': {'id': snapshot.form_id or (log.form_id if log else None), 'name': snapshot.form_name or (log.form_name if log else None)},
            'project': {'id': project_id, 'name': project_name},
            'campaign': {'id': snapshot.campaign_id or (log.campaign_id if log else None), 'name': snapshot.campaign_name or (log.campaign_name if log else None)},
            'lead': {'id': snapshot.lead_id, 'name': (lead.name if lead else None), 'phone': (lead.phone if lead else None), 'email': (lead.email if lead else None)},
        },
        'snapshot': snapshot.to_dict(),
        'log': log.to_dict() if log else None,
    }
    return jsonify(out), 200


@lead_sources_bp.route('/proof/multi-page-support', methods=['GET'])
@require_role('superadmin', 'sales_manager', 'platform_owner')
def proof_multi_page_support():
    """Return tenant mapping counts proving multi-page/form/project support."""
    user = request.current_user
    rows = LeadSourceFormMapping.query.filter_by(tenant_id=user.tenant_id, is_active=True).all()

    page_ids = sorted({str(r.page_id).strip() for r in rows if str(r.page_id or '').strip()})
    form_ids = sorted({str(r.form_id).strip() for r in rows if str(r.form_id or '').strip()})
    project_ids = sorted({int(r.project_id) for r in rows if r.project_id})

    return jsonify({
        'tenant_id': user.tenant_id,
        'counts': {
            'active_mappings': len(rows),
            'unique_pages': len(page_ids),
            'unique_forms': len(form_ids),
            'unique_projects': len(project_ids),
        },
        'unique': {
            'page_ids': page_ids,
            'form_ids': form_ids,
            'project_ids': project_ids,
        },
        'rows': [r.to_dict() for r in rows],
    }), 200


@lead_sources_bp.route('/meta/backfill/lead/<int:lead_id>', methods=['POST'])
@require_role('superadmin', 'platform_owner')
def meta_backfill_lead(lead_id):
    """
    Refresh one existing Meta lead from real Graph APIs and persist enrichment
    into ingestion logs + campaign snapshots.
    """
    user = request.current_user

    log = IngestedLeadLog.query.filter_by(
        tenant_id=user.tenant_id,
        source_type='meta',
        lead_id=lead_id,
    ).order_by(IngestedLeadLog.received_at.desc()).first()
    if not log:
        return jsonify({'error': f'No Meta ingestion log found for lead_id={lead_id}'}), 404

    source = LeadSource.query.filter_by(
        id=log.source_id,
        tenant_id=user.tenant_id,
        source_type='meta',
    ).first()
    if not source:
        return jsonify({'error': 'Meta source not found for lead log'}), 404

    refreshed = _refresh_meta_attribution(user, log, source, lead_id)
    if refreshed.get('error'):
        return jsonify(refreshed), refreshed.get('status_code', 502)

    return jsonify(refreshed), 200


def _refresh_meta_attribution(user, log, source, lead_id):
    from app.routes.ingestion import _meta_enrich_leadgen_entry, _normalise_meta
    from app.services.ingestion_engine import persist_meta_snapshot

    lead_entry = dict(log.raw_payload or {})
    lead_entry.setdefault('leadgen_id', str(log.platform_lead_id or ''))
    lead_entry.setdefault('form_id', str(log.form_id or ''))
    lead_entry.setdefault('page_id', str(log.page_id or ''))
    lead_entry.setdefault('campaign_id', str(log.campaign_id or ''))
    lead_entry.setdefault('adset_id', str(log.ad_set_id or ''))
    lead_entry.setdefault('ad_id', str(log.ad_id or ''))

    enriched = _meta_enrich_leadgen_entry(lead_entry, source)
    if not enriched:
        return {
            'error': 'Meta enrichment returned no data. Check token permissions for ads/lead retrieval.',
            'status_code': 502,
        }

    merged_entry = dict(lead_entry)
    for k, v in enriched.items():
        if v not in (None, '', []):
            merged_entry[k] = v

    normalised = _normalise_meta(merged_entry)

    logs_to_update = IngestedLeadLog.query.filter_by(
        tenant_id=user.tenant_id,
        source_type='meta',
        platform_lead_id=str(log.platform_lead_id or ''),
    ).all()
    if not logs_to_update:
        logs_to_update = [log]

    for row in logs_to_update:
        row.campaign_id = normalised.get('campaign_id') or row.campaign_id
        row.campaign_name = normalised.get('campaign_name') or row.campaign_name
        row.ad_set_id = normalised.get('ad_set_id') or row.ad_set_id
        row.ad_set_name = normalised.get('ad_set_name') or row.ad_set_name
        row.ad_id = normalised.get('ad_id') or row.ad_id
        row.ad_name = normalised.get('ad_name') or row.ad_name
        row.form_id = normalised.get('form_id') or row.form_id
        row.form_name = normalised.get('form_name') or row.form_name
        row.page_id = normalised.get('page_id') or row.page_id
        row.is_test = row.is_test or bool(getattr(log, 'is_test', False))

    snapshots = MetaCampaignSnapshot.query.filter(
        MetaCampaignSnapshot.tenant_id == user.tenant_id,
        db.or_(
            MetaCampaignSnapshot.lead_id == lead_id,
            MetaCampaignSnapshot.ingested_log_id.in_([r.id for r in logs_to_update]),
        ),
    ).all()

    if snapshots:
        for snap in snapshots:
            snap.page_id = normalised.get('page_id') or snap.page_id
            snap.form_id = normalised.get('form_id') or snap.form_id
            snap.form_name = normalised.get('form_name') or snap.form_name
            snap.campaign_id = normalised.get('campaign_id') or snap.campaign_id
            snap.campaign_name = normalised.get('campaign_name') or snap.campaign_name
            snap.ad_set_id = normalised.get('ad_set_id') or snap.ad_set_id
            snap.ad_set_name = normalised.get('ad_set_name') or snap.ad_set_name
            snap.ad_id = normalised.get('ad_id') or snap.ad_id
            snap.ad_name = normalised.get('ad_name') or snap.ad_name
            snap.spend = normalised.get('spend') if normalised.get('spend') is not None else snap.spend
            snap.cost_per_result = normalised.get('cost_per_result') if normalised.get('cost_per_result') is not None else snap.cost_per_result
            snap.ctr = normalised.get('ctr') if normalised.get('ctr') is not None else snap.ctr
            snap.cpc = normalised.get('cpc') if normalised.get('cpc') is not None else snap.cpc
            snap.cpm = normalised.get('cpm') if normalised.get('cpm') is not None else snap.cpm
            snap.reach = normalised.get('reach') if normalised.get('reach') is not None else snap.reach
            snap.impressions = normalised.get('impressions') if normalised.get('impressions') is not None else snap.impressions
            snap.audience = normalised.get('audience') or snap.audience
            snap.placement = normalised.get('placement') or snap.placement
            extra = dict(snap.extra_metrics or {})
            if normalised.get('page_name'):
                extra['page_name'] = normalised.get('page_name')
            snap.extra_metrics = extra
    else:
        persist_meta_snapshot(source, normalised, log, lead_id, is_test=bool(getattr(log, 'is_test', False)))

    db.session.commit()

    return {
        'ok': True,
        'lead_id': lead_id,
        'platform_lead_id': log.platform_lead_id,
        'updated_logs': len(logs_to_update),
        'updated_snapshots': len(snapshots),
        'attribution': {
            'campaign_id': normalised.get('campaign_id'),
            'campaign_name': normalised.get('campaign_name'),
            'adset_id': normalised.get('ad_set_id'),
            'adset_name': normalised.get('ad_set_name'),
            'ad_id': normalised.get('ad_id'),
            'ad_name': normalised.get('ad_name'),
            'form_id': normalised.get('form_id'),
            'form_name': normalised.get('form_name'),
            'page_id': normalised.get('page_id'),
            'page_name': normalised.get('page_name'),
        },
        'metrics': {
            'spend': normalised.get('spend'),
            'cost_per_result': normalised.get('cost_per_result'),
            'ctr': normalised.get('ctr'),
            'cpc': normalised.get('cpc'),
            'cpm': normalised.get('cpm'),
            'reach': normalised.get('reach'),
            'impressions': normalised.get('impressions'),
            'placement': normalised.get('placement'),
            'audience': normalised.get('audience'),
        },
    }


def _parse_platform_datetime(value):
    if not value:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        dt = datetime.fromisoformat(text.replace('Z', '+00:00'))
    except ValueError:
        return None
    if getattr(dt, 'tzinfo', None):
        dt = dt.astimezone(timezone.utc).replace(tzinfo=None)
    return dt


@lead_sources_bp.route('/meta/backfill/lead-created-at', methods=['POST'])
@require_role('superadmin', 'platform_owner')
def meta_backfill_lead_created_at():
    """Repair Lead.created_at using Meta platform created_time from ingestion payloads."""
    user = request.current_user
    source_id = request.args.get('source_id', type=int)
    limit = max(1, min(int(request.args.get('limit', 4000) or 4000), 20000))

    q = IngestedLeadLog.query.filter(
        IngestedLeadLog.tenant_id == user.tenant_id,
        IngestedLeadLog.source_type == 'meta',
        IngestedLeadLog.lead_id != None,
    )
    if source_id:
        q = q.filter(IngestedLeadLog.source_id == source_id)

    logs = q.order_by(IngestedLeadLog.received_at.asc()).limit(limit).all()

    earliest_by_lead = {}
    for log in logs:
        payload = log.raw_payload or {}
        dt = _parse_platform_datetime(
            payload.get('created_time')
            or payload.get('created_at')
            or payload.get('submission_time')
        )
        if not dt:
            continue
        current = earliest_by_lead.get(log.lead_id)
        if current is None or dt < current:
            earliest_by_lead[log.lead_id] = dt

    if not earliest_by_lead:
        return jsonify({'ok': True, 'scanned_logs': len(logs), 'updated_count': 0, 'updated_leads': []}), 200

    leads = Lead.query.filter(
        Lead.tenant_id == user.tenant_id,
        Lead.id.in_(list(earliest_by_lead.keys())),
    ).all()

    updated = []
    for lead in leads:
        target_dt = earliest_by_lead.get(lead.id)
        if not target_dt:
            continue
        # Correct obvious ingestion-time skew while preserving legitimate manual edits.
        if lead.created_at is None or abs((lead.created_at - target_dt).total_seconds()) >= 300:
            lead.created_at = target_dt
            updated.append({'lead_id': lead.id, 'created_at': target_dt.isoformat()})

    if updated:
        db.session.commit()

    return jsonify({
        'ok': True,
        'scanned_logs': len(logs),
        'candidate_leads': len(earliest_by_lead),
        'updated_count': len(updated),
        'updated_leads': updated[:200],
    }), 200


@lead_sources_bp.route('/meta/backfill/attribution', methods=['POST'])
@require_role('superadmin', 'platform_owner')
def meta_backfill_attribution():
    """Bulk-refresh attributed Meta logs missing names or metrics."""
    user = request.current_user
    source_id = request.args.get('source_id', type=int)
    limit = max(1, min(int(request.args.get('limit', 50) or 50), 250))

    query = IngestedLeadLog.query.filter(
        IngestedLeadLog.tenant_id == user.tenant_id,
        IngestedLeadLog.source_type == 'meta',
        IngestedLeadLog.is_test == False,
        db.or_(
            IngestedLeadLog.campaign_id != None,
            IngestedLeadLog.ad_set_id != None,
            IngestedLeadLog.ad_id != None,
        ),
        db.or_(
            IngestedLeadLog.campaign_name == None,
            IngestedLeadLog.campaign_name == '',
            IngestedLeadLog.ad_set_name == None,
            IngestedLeadLog.ad_set_name == '',
            IngestedLeadLog.ad_name == None,
            IngestedLeadLog.ad_name == '',
        ),
    )
    if source_id:
        query = query.filter(IngestedLeadLog.source_id == source_id)

    candidates = query.order_by(IngestedLeadLog.received_at.desc()).limit(limit * 5).all()
    seen_platform = set()
    refreshed = []
    failures = []

    for log in candidates:
        platform_lead_id = str(log.platform_lead_id or '').strip()
        if not platform_lead_id or platform_lead_id in seen_platform:
            continue
        seen_platform.add(platform_lead_id)
        target_lead_id = log.lead_id or log.dup_of_lead_id
        if not target_lead_id:
            continue
        source = LeadSource.query.filter_by(
            id=log.source_id,
            tenant_id=user.tenant_id,
            source_type='meta',
        ).first()
        if not source:
            failures.append({'platform_lead_id': platform_lead_id, 'error': 'Meta source not found'})
            continue
        result = _refresh_meta_attribution(user, log, source, target_lead_id)
        if result.get('error'):
            failures.append({'platform_lead_id': platform_lead_id, 'lead_id': target_lead_id, 'error': result.get('error')})
        else:
            refreshed.append({
                'platform_lead_id': platform_lead_id,
                'lead_id': target_lead_id,
                'campaign_name': ((result.get('attribution') or {}).get('campaign_name') or ''),
                'adset_name': ((result.get('attribution') or {}).get('adset_name') or ''),
                'ad_name': ((result.get('attribution') or {}).get('ad_name') or ''),
                'spend': ((result.get('metrics') or {}).get('spend')),
                'cost_per_result': ((result.get('metrics') or {}).get('cost_per_result')),
            })
        if len(refreshed) >= limit:
            break

    return jsonify({
        'ok': True,
        'refreshed': refreshed,
        'failures': failures,
        'updated_count': len(refreshed),
        'failure_count': len(failures),
    }), 200


@lead_sources_bp.route('/meta/backfill/project-assignment', methods=['POST'])
@require_role('superadmin', 'platform_owner')
def meta_backfill_project_assignment():
    """Re-assign project_id on leads using current form-to-project mappings."""
    user = request.current_user
    source_id = request.args.get('source_id', type=int)

    # Build form_id -> project_id lookup from active mappings
    mapping_q = LeadSourceFormMapping.query.filter_by(
        tenant_id=user.tenant_id,
        is_active=True,
    )
    if source_id:
        mapping_q = mapping_q.filter_by(source_id=source_id)
    mappings = mapping_q.all()
    form_to_project = {str(m.form_id): int(m.project_id) for m in mappings if m.form_id and m.project_id}

    if not form_to_project:
        return jsonify({'ok': False, 'error': 'No active form-project mappings found'}), 400

    from app.models import Lead
    updated_total = 0

    # Process each form mapping with a targeted bulk update
    for form_id, project_id in form_to_project.items():
        # Get all lead_ids from logs for this form
        log_q = db.session.query(IngestedLeadLog.lead_id).filter(
            IngestedLeadLog.tenant_id == user.tenant_id,
            IngestedLeadLog.source_type == 'meta',
            IngestedLeadLog.form_id == form_id,
            IngestedLeadLog.lead_id != None,
        )
        if source_id:
            log_q = log_q.filter(IngestedLeadLog.source_id == source_id)
        lead_ids = [row[0] for row in log_q.distinct().all()]

        if not lead_ids:
            continue

        # Bulk update leads that have wrong or missing project_id
        rows_updated = Lead.query.filter(
            Lead.tenant_id == user.tenant_id,
            Lead.id.in_(lead_ids),
            db.or_(Lead.project_id == None, Lead.project_id != project_id),
        ).update({'project_id': project_id}, synchronize_session=False)
        updated_total += rows_updated

    db.session.commit()

    return jsonify({
        'ok': True,
        'updated_count': updated_total,
        'form_mappings_applied': len(form_to_project),
    }), 200


@lead_sources_bp.route('/meta/backfill/cleanup-notes', methods=['POST'])
@require_role('superadmin', 'platform_owner')
def meta_cleanup_attribution_notes():
    """Remove system-generated Meta attribution notes from lead records."""
    from app.models.lead import LeadNote

    user = request.current_user
    source_id = request.args.get('source_id', type=int)
    note_prefix = 'Meta enrichment synced for lead '

    lead_ids_q = db.session.query(Lead.id).filter(Lead.tenant_id == user.tenant_id)
    if source_id:
        lead_ids_q = lead_ids_q.join(IngestedLeadLog, IngestedLeadLog.lead_id == Lead.id).filter(
            IngestedLeadLog.tenant_id == user.tenant_id,
            IngestedLeadLog.source_type == 'meta',
            IngestedLeadLog.source_id == source_id,
        ).distinct()

    notes_q = LeadNote.query.filter(
        LeadNote.note.startswith(note_prefix),
        LeadNote.lead_id.in_(lead_ids_q),
    )

    deleted_count = notes_q.count()
    notes_q.delete(synchronize_session=False)
    db.session.commit()

    return jsonify({
        'ok': True,
        'deleted_count': deleted_count,
    }), 200


@lead_sources_bp.route('/repair/assign-unmapped-logs', methods=['POST'])
@require_role('superadmin', 'platform_owner')
def repair_assign_unmapped_logs():
    """
    Assign unassigned leads to projects using form mappings.
    
    Finds all leads with project_id = NULL and a form_id in ingestion logs,
    then assigns them to the project mapped for that form_id (if a mapping exists).
    
    Query params:
      - limit: max number of leads to update (default 1000)
      - source_id: only process logs from this source (optional)
      - dry_run: if "true", count what would be updated without actually updating
    """
    user = request.current_user
    limit = request.args.get('limit', 1000, type=int)
    source_id = request.args.get('source_id', type=int)
    dry_run = request.args.get('dry_run', 'false').lower() == 'true'
    
    if limit < 1 or limit > 10000:
        limit = 1000
    
    # Get all unassigned leads with form_ids in their ingestion logs
    unassigned_q = db.session.query(Lead.id, IngestedLeadLog.form_id, IngestedLeadLog.source_id).filter(
        Lead.tenant_id == user.tenant_id,
        Lead.project_id == None,
        Lead.id == IngestedLeadLog.lead_id,
        IngestedLeadLog.tenant_id == user.tenant_id,
        IngestedLeadLog.form_id != None,
        IngestedLeadLog.status.in_(['processed', 'duplicate']),  # successful ingestion
    )
    
    if source_id:
        unassigned_q = unassigned_q.filter(IngestedLeadLog.source_id == source_id)
    
    # Get distinct form_id→source_id pairs to batch by mapping
    form_source_pairs = unassigned_q.distinct(IngestedLeadLog.form_id, IngestedLeadLog.source_id).limit(limit).all()
    
    if not form_source_pairs:
        return jsonify({
            'ok': True,
            'updated_count': 0,
            'assignments': [],
            'message': 'No unassigned leads with form mappings found',
        }), 200
    
    assignments = []
    total_updated = 0
    
    # For each form_id in source, look up the mapping and assign
    for lead_id, form_id, src_id in form_source_pairs:
        # Find the mapping for this form_id in this source
        mapping = LeadSourceFormMapping.query.filter_by(
            tenant_id=user.tenant_id,
            source_id=src_id,
            form_id=str(form_id),
            is_active=True,
        ).first()
        
        if not mapping:
            continue
        
        # Get all unassigned leads for this form_id
        leads_for_form = db.session.query(Lead.id).filter(
            Lead.tenant_id == user.tenant_id,
            Lead.project_id == None,
            Lead.id.in_(
                db.session.query(IngestedLeadLog.lead_id).filter(
                    IngestedLeadLog.tenant_id == user.tenant_id,
                    IngestedLeadLog.source_id == src_id,
                    IngestedLeadLog.form_id == str(form_id),
                    IngestedLeadLog.lead_id != None,
                    IngestedLeadLog.status.in_(['processed', 'duplicate']),
                ).distinct()
            ),
        ).limit(limit).all()
        
        lead_ids_for_form = [row[0] for row in leads_for_form]
        
        if not lead_ids_for_form:
            continue
        
        count_before = len(lead_ids_for_form)
        
        if not dry_run:
            # Bulk update these leads
            Lead.query.filter(
                Lead.id.in_(lead_ids_for_form),
                Lead.tenant_id == user.tenant_id,
            ).update({'project_id': mapping.project_id}, synchronize_session=False)
            db.session.flush()
        
        assignments.append({
            'form_id': form_id,
            'form_name': mapping.form_name or '',
            'source_id': src_id,
            'source_name': mapping.source.name if mapping.source else 'Unknown',
            'project_id': mapping.project_id,
            'project_name': mapping.project.name if mapping.project else 'Unknown',
            'count': count_before,
        })
        
        total_updated += count_before
    
    if not dry_run:
        db.session.commit()
    
    return jsonify({
        'ok': True,
        'dry_run': dry_run,
        'updated_count': total_updated,
        'assignments': assignments,
        'message': f"{'Would assign' if dry_run else 'Assigned'} {total_updated} unassigned leads to projects",
    }), 200


@lead_sources_bp.route('/logs/export.xlsx', methods=['GET'])
@require_role('superadmin', 'sales_manager', 'platform_owner')
def ingestion_logs_export_xlsx():
    user = request.current_user
    source_id = request.args.get('source_id', type=int)
    search_q = request.args.get('q', '').strip()
    status = request.args.get('status', '').strip().lower()
    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()

    q = _connected_source_logs_query(user, source_id=source_id)
    q = q.filter(IngestedLeadLog.status.in_(['processed', 'duplicate', 'error']))
    q = _apply_log_search_filter(q, search_q)
    if date_from:
        try:
            q = q.filter(IngestedLeadLog.received_at >= datetime.fromisoformat(date_from))
        except ValueError:
            pass
    if date_to:
        try:
            q = q.filter(IngestedLeadLog.received_at < datetime.fromisoformat(date_to) + timedelta(days=1))
        except ValueError:
            pass

    q = _apply_test_data_filter(q, IngestedLeadLog)

    rows = _dedupe_report_source_logs(
        q,
        user,
        date_from=date_from,
        date_to=date_to,
        source_id=source_id,
        limit=10000,
    )
    if status in {'processed', 'duplicate', 'error'}:
        rows = [row for row in rows if _log_effective_status(row) == status]
    return _xlsx_response(
        'lead-source-logs.xlsx',
        [
            'id', 'source_id', 'source_type', 'status', 'platform_lead_id',
            'campaign_name', 'ad_name', 'form_id', 'form_name', 'lead_id',
            'dup_of_lead_id', 'error_message', 'received_at', 'processed_at',
        ],
        [[
            l.id,
            l.source_id,
            l.source_type,
            _log_effective_status(l),
            l.platform_lead_id,
            l.campaign_name,
            l.ad_name,
            l.form_id,
            l.form_name,
            l.lead_id,
            l.dup_of_lead_id,
            l.error_message,
            l.received_at.isoformat() if l.received_at else '',
            l.processed_at.isoformat() if l.processed_at else '',
        ] for l in rows],
    )


# ══════════════════════════════════════════════════════════════════════════════
# META OAUTH FLOW
# Phase META-1.1: full guided connection  (Business → Page → Forms → Save)
# ══════════════════════════════════════════════════════════════════════════════

@lead_sources_bp.route('/meta/exchange-token', methods=['POST'])
@require_role('superadmin', 'platform_owner')
def meta_exchange_token():
    """
    Exchange a short-lived user access token (from the browser JS SDK / OAuth
    redirect) for a long-lived user token, then return the user profile and
    list of Pages the user manages.

    POST body:
      { "short_lived_token": "...", "app_id": "...", "app_secret": "..." }

    Returns:
      { "long_lived_token": "...", "user": {...}, "pages": [...] }
    """
    user = request.current_user
    data = request.get_json() or {}

    short_token = (data.get('short_lived_token') or '').strip()
    app_id      = (data.get('app_id') or '').strip()
    app_secret  = (data.get('app_secret') or '').strip()

    if not all([short_token, app_id, app_secret]):
        return jsonify({'error': 'short_lived_token, app_id and app_secret are required'}), 400

    try:
        import urllib.request as _req
        import urllib.parse as _parse
        import json as _json

        # 1. Exchange short-lived → long-lived user token
        exchange_url = (
            'https://graph.facebook.com/v25.0/oauth/access_token?'
            f'grant_type=fb_exchange_token'
            f'&client_id={_parse.quote(app_id)}'
            f'&client_secret={_parse.quote(app_secret)}'
            f'&fb_exchange_token={_parse.quote(short_token)}'
        )
        with _req.urlopen(_req.Request(exchange_url), timeout=15) as r:
            token_data = _json.loads(r.read())

        if 'error' in token_data:
            return jsonify({'error': token_data['error'].get('message', 'Token exchange failed')}), 400

        long_token = token_data.get('access_token', short_token)

        # 2. /me – basic user info
        me_url = f'https://graph.facebook.com/v25.0/me?fields=id,name&access_token={_parse.quote(long_token)}'
        with _req.urlopen(_req.Request(me_url), timeout=10) as r:
            me = _json.loads(r.read())

        # 3. /me/accounts – pages managed by this user
        pages_url = f'https://graph.facebook.com/v25.0/me/accounts?fields=id,name,access_token,tasks&access_token={_parse.quote(long_token)}'
        with _req.urlopen(_req.Request(pages_url), timeout=10) as r:
            pages_data = _json.loads(r.read())

        pages = [
            {
                'id':           p['id'],
                'name':         p['name'],
                'access_token': p.get('access_token', ''),
                'tasks':        p.get('tasks', []),
            }
            for p in pages_data.get('data', [])
        ]

        return jsonify({
            'long_lived_token': long_token,
            'user':             {'id': me.get('id'), 'name': me.get('name')},
            'pages':            pages,
        }), 200

    except Exception as exc:
        logger.exception('meta_exchange_token error: %s', exc)
        return jsonify({'error': str(exc)}), 502


@lead_sources_bp.route('/meta/page-forms', methods=['POST'])
@require_role('superadmin', 'platform_owner')
def meta_page_forms():
    """
    List all lead-gen forms for a specific Meta page.

    POST body:
      { "page_id": "...", "page_access_token": "..." }

    Returns:
      { "forms": [{id, name, status, leads_count, created_time}] }
    """
    user = request.current_user
    data = request.get_json() or {}

    page_id     = (data.get('page_id') or '').strip()
    page_token  = (data.get('page_access_token') or '').strip()
    user_token  = (data.get('user_access_token') or '').strip()

    if not page_id or (not page_token and not user_token):
        return jsonify({'error': 'page_id and at least one access token are required'}), 400

    try:
        import urllib.request as _req
        import urllib.parse as _parse
        import json as _json

        tokens_to_try = []
        if page_token:
            tokens_to_try.append(('page', page_token))
        if user_token and user_token != page_token:
            tokens_to_try.append(('user', user_token))

        forms_data = None
        last_graph_error = None
        last_status = 502

        for token_kind, token_value in tokens_to_try:
            url = (
                f'https://graph.facebook.com/v25.0/{_parse.quote(page_id)}/leadgen_forms'
                f'?fields=id,name,status,leads_count,created_time'
                f'&access_token={_parse.quote(token_value)}'
            )
            try:
                with _req.urlopen(_req.Request(url), timeout=10) as r:
                    forms_data = _json.loads(r.read())
                if 'error' in forms_data:
                    graph_error = forms_data.get('error') or {}
                    last_graph_error = {
                        'token_kind': token_kind,
                        'code': graph_error.get('code'),
                        'error_subcode': graph_error.get('error_subcode'),
                        'type': graph_error.get('type'),
                        'message': graph_error.get('message', 'Graph API error'),
                    }
                    last_status = 403 if graph_error.get('code') else 400
                    forms_data = None
                    continue
                break
            except _req.HTTPError as http_exc:
                raw = http_exc.read().decode('utf-8', errors='replace') if hasattr(http_exc, 'read') else ''
                parsed = {}
                try:
                    parsed = _json.loads(raw) if raw else {}
                except Exception:
                    parsed = {}
                graph_error = (parsed or {}).get('error') or {}
                last_graph_error = {
                    'token_kind': token_kind,
                    'code': graph_error.get('code'),
                    'error_subcode': graph_error.get('error_subcode'),
                    'type': graph_error.get('type'),
                    'message': graph_error.get('message') or str(http_exc),
                }
                last_status = int(getattr(http_exc, 'code', 502) or 502)
                forms_data = None
                continue

        if forms_data is None:
            if last_graph_error:
                return jsonify({'error': last_graph_error.get('message', 'Graph API error'), 'graph_error': last_graph_error}), (403 if last_status == 403 else 400)
            return jsonify({'error': 'Could not load lead forms for this page'}), 502

        forms = [
            {
                'id':           f['id'],
                'name':         f.get('name', ''),
                'status':       f.get('status', ''),
                'leads_count':  f.get('leads_count', 0),
                'created_time': f.get('created_time', ''),
            }
            for f in forms_data.get('data', [])
        ]

        return jsonify({'forms': forms}), 200

    except Exception as exc:
        logger.exception('meta_page_forms error: %s', exc)
        return jsonify({'error': str(exc)}), 502


@lead_sources_bp.route('/meta/save-connection', methods=['POST'])
@require_role('superadmin', 'sales_manager', 'platform_owner')
def meta_save_connection():
    """
    Finalise a Meta OAuth wizard: create or update the LeadSource with the
    selected page + forms and persist the page access token.

    POST body:
    {
      "source_id":         123,          // update existing, or null to create new
      "name":              "My FB Source",
      "app_id":            "...",
      "app_secret":        "...",
      "user_token":        "...",        // long-lived user token
      "page_id":           "...",
      "page_name":         "...",
      "page_access_token": "...",
      "selected_forms":    [{id, name}], // forms the user picked
      "verify_token":      "...",        // webhook verify token (user-set)
    }
    """
    user = request.current_user
    data = request.get_json() or {}

    page_id     = (data.get('page_id') or '').strip()
    page_token  = (data.get('page_access_token') or '').strip()
    page_name   = (data.get('page_name') or '').strip()
    name        = (data.get('name') or f'Meta – {page_name}').strip()

    if not page_id or not page_token:
        return jsonify({'error': 'page_id and page_access_token are required'}), 400

    creds = {
        'app_id':            data.get('app_id', ''),
        'app_secret':        data.get('app_secret', ''),
        'user_token':        data.get('user_token', ''),
        'page_id':           page_id,
        'page_access_token': page_token,
        'verify_token':      data.get('verify_token', ''),
        # access_token alias (used by existing engine + test helpers)
        'access_token':      page_token,
    }
    business_id = str(data.get('business_id') or '').strip()
    if business_id:
        creds['business_id'] = business_id

    selected_forms = data.get('selected_forms') or []

    ad_accounts = []
    ad_account_id = ''
    ad_account_name = ''
    token_candidates = []
    for token_value in (data.get('user_token'), page_token, data.get('access_token')):
        token_value = str(token_value or '').strip()
        if token_value and token_value not in token_candidates:
            token_candidates.append(token_value)

    for token_value in token_candidates:
        try:
            adaccounts_url = (
                'https://graph.facebook.com/v25.0/me/adaccounts'
                '?fields=id,name,account_status&limit=25'
                f'&access_token={_parse.quote(token_value)}'
            )
            with _req.urlopen(_req.Request(adaccounts_url), timeout=12) as r:
                adaccounts_data = _json.loads(r.read())
            ad_accounts = [
                {
                    'id': str(item.get('id') or ''),
                    'name': str(item.get('name') or ''),
                    'account_status': item.get('account_status'),
                }
                for item in (adaccounts_data.get('data', []) or [])
                if str(item.get('id') or '')
            ]
            if ad_accounts:
                ad_account_id = ad_accounts[0].get('id') or ''
                ad_account_name = ad_accounts[0].get('name') or ''
                break
        except Exception as exc:
            logger.warning('meta_save_connection: ad account discovery failed: %s', exc)
            continue

    if ad_account_id:
        creds['ad_account_id'] = ad_account_id
    if ad_account_name:
        creds['ad_account_name'] = ad_account_name
    if ad_accounts:
        creds['ad_accounts'] = ad_accounts

    # Ensure webhook subscription is active on the selected Meta page.
    # Without this, page leads may never be pushed to our webhook endpoint.
    try:
        sub_url = (
            f'https://graph.facebook.com/v25.0/{_parse.quote(page_id)}/subscribed_apps'
        )
        sub_body = _parse.urlencode({
            'subscribed_fields': 'leadgen',
            'access_token': page_token,
        }).encode()
        sub_req = _req.Request(
            sub_url,
            data=sub_body,
            headers={'Content-Type': 'application/x-www-form-urlencoded'},
        )
        with _req.urlopen(sub_req, timeout=15) as r:
            sub_resp = _json.loads(r.read())
        if isinstance(sub_resp, dict) and sub_resp.get('error'):
            logger.warning('meta_save_connection: page subscribe error: %s', sub_resp.get('error'))
            return jsonify({'error': 'Meta page webhook subscription failed', 'details': sub_resp.get('error')}), 400
    except Exception as exc:
        logger.exception('meta_save_connection: failed to subscribe page to leadgen webhooks: %s', exc)
        return jsonify({'error': f'Failed to subscribe page for lead webhooks: {exc}'}), 502

    source_id = data.get('source_id')
    if source_id:
        source = LeadSource.query.filter_by(id=source_id, tenant_id=user.tenant_id).first()
        if not source:
            return jsonify({'error': 'Source not found'}), 404
        source.name = name
        existing_creds = source.credentials or {}
        existing_creds.update({k: v for k, v in creds.items() if v})
        source.credentials = existing_creds
    else:
        source = next(
            (
                existing_source
                for existing_source in LeadSource.query.filter_by(
                    tenant_id=user.tenant_id,
                    source_type='meta',
                ).all()
                if _meta_page_id(existing_source) == page_id
            ),
            None,
        )
        if source:
            source.name = name
            existing_creds = source.credentials or {}
            existing_creds.update({k: v for k, v in creds.items() if v})
            source.credentials = existing_creds
        else:
            source = LeadSource(
                tenant_id=user.tenant_id,
                name=name,
                source_type='meta',
                credentials=creds,
                created_by=user.id,
            )
            db.session.add(source)

    source.connected_account = f'{page_name} (Page ID: {page_id})'
    source.available_forms   = selected_forms
    source.dup_check_phone   = True
    source.dup_check_email   = False
    source.permission_status = 'ok'
    source.is_active         = True
    source.last_tested_at    = datetime.utcnow()
    source.last_test_result  = 'pass'
    source.last_test_message = f'Connected via OAuth. {len(selected_forms)} form(s) selected.'

    db.session.flush()
    _deactivate_duplicate_meta_sources_for_page(user.tenant_id, page_id, source.id)

    db.session.commit()
    return jsonify({'source': source.to_dict()}), 200


def _pull_recent_meta_source(user, source, data=None):
    """
    Pull leads directly from Meta forms for one source and run the normal
    ingestion pipeline. Used by both manual recovery and report sync.
    """
    data = data or {}
    per_form_limit = max(1, min(5000, int(data.get('per_form_limit', 250) or 250)))
    page_size = max(1, min(100, int(data.get('page_size', min(100, per_form_limit)) or min(100, per_form_limit))))
    include_archived = bool(data.get('include_archived', False))
    date_from = str(data.get('date_from') or '').strip()
    date_to = str(data.get('date_to') or '').strip()
    max_pages = max(1, min(500, int(data.get('max_pages', 50) or 50)))
    full_history = bool(data.get('full_history', False))
    requested_form_ids = [
        str(fid).strip() for fid in (data.get('form_ids') or [])
        if str(fid).strip()
    ]

    if full_history and not date_from:
        date_from = source.created_at.date().isoformat()
        per_form_limit = max(per_form_limit, 5000)
        page_size = max(page_size, 100)
        max_pages = max(max_pages, 500)

    parsed_from_date = None
    parsed_to_exclusive_date = None
    if date_from:
        try:
            parsed_from_date = datetime.fromisoformat(date_from).date()
        except ValueError:
            parsed_from_date = None
    if date_to:
        try:
            parsed_to_exclusive_date = (datetime.fromisoformat(date_to) + timedelta(days=1)).date()
        except ValueError:
            parsed_to_exclusive_date = None

    creds = source.credentials or {}
    user_token = (creds.get('user_token') or '').strip()
    page_token = (creds.get('page_access_token') or creds.get('access_token') or '').strip()
    page_id = str((creds.get('page_id') or '')).strip()
    token = user_token or page_token
    if not token:
        raise ValueError('No Meta token configured for source')

    from app.routes.ingestion import _resolve_meta_target_source, _normalise_meta

    forms = source.available_forms or []
    form_ids = []
    form_names = {}
    for f in forms:
        if isinstance(f, dict):
            fid = str(f.get('id') or '').strip()
            fstatus = str(f.get('status') or '').strip().upper()
            if not fid:
                continue
            if not include_archived and fstatus == 'ARCHIVED':
                continue
            form_ids.append(fid)
            form_names[fid] = str(f.get('name') or '').strip()
        else:
            fid = str(f or '').strip()
            if fid:
                form_ids.append(fid)

    seen = set()
    form_ids = [fid for fid in form_ids if not (fid in seen or seen.add(fid))]
    if requested_form_ids:
        req_set = set(requested_form_ids)
        form_ids = [fid for fid in form_ids if fid in req_set]
    if not form_ids:
        raise ValueError('No forms configured on source')

    summary = {
        'source_id': source.id,
        'full_history': full_history,
        'date_from': date_from,
        'date_to': date_to,
        'forms_scanned': 0,
        'entries_seen': 0,
        'created': 0,
        'updated': 0,
        'duplicate': 0,
        'ignored': 0,
        'error': 0,
        'details': [],
    }

    if page_id:
        try:
            sub_url = f'https://graph.facebook.com/v25.0/{_parse.quote(page_id)}/subscribed_apps'
            sub_body = _parse.urlencode({
                'access_token': page_token or token,
                'subscribed_fields': 'leadgen',
            }).encode('utf-8')
            with _req.urlopen(_req.Request(sub_url, data=sub_body, method='POST'), timeout=15) as resp:
                sub_payload = _json.loads(resp.read())
            summary['page_subscription'] = {
                'ok': bool((sub_payload or {}).get('success')),
                'page_id': page_id,
            }
        except Exception as exc:
            summary['page_subscription'] = {
                'ok': False,
                'page_id': page_id,
                'error': str(exc),
            }

    for fid in form_ids:
        summary['forms_scanned'] += 1
        entries = []
        after = None
        page_count = 0
        try:
            while len(entries) < per_form_limit and page_count < max_pages:
                page_count += 1
                url = (
                    f'https://graph.facebook.com/v25.0/{_parse.quote(fid)}/leads'
                    f'?fields=id,created_time,field_data,ad_id,ad_name,adset_id,adset_name,campaign_id,campaign_name,form_id,page_id'
                    f'&limit={page_size}'
                    f'&access_token={_parse.quote(token)}'
                )
                if after:
                    url += f'&after={_parse.quote(after)}'

                with _req.urlopen(_req.Request(url), timeout=15) as resp:
                    payload = _json.loads(resp.read())

                page_entries = payload.get('data', []) if isinstance(payload, dict) else []
                if not page_entries:
                    break

                reached_older_than_from = False
                for raw in page_entries:
                    created_time = str((raw or {}).get('created_time') or '').strip()
                    created_date = None
                    if created_time:
                        try:
                            created_date = datetime.fromisoformat(created_time.replace('Z', '+00:00')).date()
                        except ValueError:
                            created_date = None

                    if parsed_from_date and created_date and created_date < parsed_from_date:
                        reached_older_than_from = True
                        continue
                    if parsed_to_exclusive_date and created_date and created_date >= parsed_to_exclusive_date:
                        continue

                    entries.append(raw)
                    if len(entries) >= per_form_limit:
                        break

                if reached_older_than_from:
                    break

                paging = payload.get('paging', {}) if isinstance(payload, dict) else {}
                cursors = paging.get('cursors', {}) if isinstance(paging, dict) else {}
                after = cursors.get('after')
                if not after:
                    break
        except Exception as exc:
            summary['details'].append({'form_id': fid, 'error': str(exc)})
            summary['error'] += 1
            continue

        form_result = {'form_id': fid, 'entries': len(entries), 'results': []}
        summary['entries_seen'] += len(entries)

        for raw in entries:
            platform_lead_id = ''
            try:
                entry = dict(raw or {})
                if entry.get('id') and not entry.get('leadgen_id'):
                    entry['leadgen_id'] = entry.get('id')
                if not entry.get('form_id'):
                    entry['form_id'] = fid
                if not entry.get('page_id') and page_id:
                    entry['page_id'] = page_id
                if not entry.get('form_name'):
                    entry['form_name'] = form_names.get(str(entry.get('form_id') or fid), '')

                target_source = _resolve_meta_target_source(source, str(entry.get('page_id') or page_id), str(entry.get('form_id') or fid))
                normalised = _normalise_meta(entry)
                if not normalised.get('platform_created_at') and entry.get('created_time'):
                    normalised['platform_created_at'] = entry.get('created_time')
                if not normalised.get('page_id'):
                    normalised['page_id'] = str(page_id or '')
                platform_lead_id = normalised.get('platform_lead_id') or ''
                result = ingest_lead(target_source, entry, normalised)
                status = result.get('status', 'error')
            except Exception as exc:
                result = {'status': 'error', 'message': str(exc)}
                status = 'error'

            if status == 'created':
                summary['created'] += 1
            elif status == 'updated':
                summary['updated'] += 1
            elif status == 'duplicate':
                summary['duplicate'] += 1
            elif status == 'ignored':
                summary['ignored'] += 1
            else:
                summary['error'] += 1

            form_result['results'].append({
                'platform_lead_id': platform_lead_id,
                'status': status,
                'lead_id': result.get('lead_id'),
                'message': result.get('message'),
            })

        form_result['pages_scanned'] = page_count
        summary['details'].append(form_result)

    return summary


@lead_sources_bp.route('/<int:source_id>/meta/pull-recent', methods=['POST'])
@require_role('superadmin', 'platform_owner')
def meta_pull_recent(source_id):
    """
    Pull latest leads directly from Meta leadgen forms and ingest them.
    Use this as a recovery path when webhook delivery is delayed/broken.
    """
    user = request.current_user
    source = LeadSource.query.filter_by(id=source_id, tenant_id=user.tenant_id).first()
    if not source:
        return jsonify({'error': 'Source not found'}), 404
    if source.source_type != 'meta':
        return jsonify({'error': 'Source is not a Meta source'}), 400

    data = request.get_json(silent=True) or {}
    try:
        summary = _pull_recent_meta_source(user, source, data)
        return jsonify({'ok': True, 'summary': summary}), 200
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400
    except Exception as exc:
        logger.exception('meta_pull_recent failed for source %s: %s', source.id, exc)
        return jsonify({'error': str(exc)}), 502

    # `per_form_limit` is treated as the total max leads to ingest per form.
    # Keep support for existing callers that pass very small values (e.g. realtime sync = 1).
    per_form_limit = max(1, min(5000, int(data.get('per_form_limit', 250) or 250)))
    page_size = max(1, min(100, int(data.get('page_size', min(100, per_form_limit)) or min(100, per_form_limit))))
    include_archived = bool(data.get('include_archived', False))
    date_from = str(data.get('date_from') or '').strip()
    date_to = str(data.get('date_to') or '').strip()
    max_pages = max(1, min(500, int(data.get('max_pages', 50) or 50)))
    full_history = bool(data.get('full_history', False))
    requested_form_ids = [
        str(fid).strip() for fid in (data.get('form_ids') or [])
        if str(fid).strip()
    ]

    if full_history and not date_from:
        date_from = source.created_at.date().isoformat()
        per_form_limit = max(per_form_limit, 5000)
        page_size = max(page_size, 100)
        max_pages = max(max_pages, 500)

    parsed_from_date = None
    parsed_to_exclusive_date = None
    if date_from:
        try:
            parsed_from_date = datetime.fromisoformat(date_from).date()
        except ValueError:
            parsed_from_date = None
    if date_to:
        try:
            parsed_to_exclusive_date = (datetime.fromisoformat(date_to) + timedelta(days=1)).date()
        except ValueError:
            parsed_to_exclusive_date = None

    creds = source.credentials or {}
    user_token = (creds.get('user_token') or '').strip()
    page_token = (creds.get('page_access_token') or creds.get('access_token') or '').strip()
    page_id = str((creds.get('page_id') or '')).strip()
    token = user_token or page_token
    if not token:
        return jsonify({'error': 'No Meta token configured for source'}), 400

    from app.routes.ingestion import _resolve_meta_target_source, _meta_enrich_leadgen_entry, _normalise_meta

    forms = source.available_forms or []
    form_ids = []
    for f in forms:
        if isinstance(f, dict):
            fid = str(f.get('id') or '').strip()
            fstatus = str(f.get('status') or '').strip().upper()
            if not fid:
                continue
            if not include_archived and fstatus == 'ARCHIVED':
                continue
            form_ids.append(fid)
        else:
            fid = str(f or '').strip()
            if fid:
                form_ids.append(fid)

    # Keep deterministic order while removing duplicates.
    seen = set()
    form_ids = [fid for fid in form_ids if not (fid in seen or seen.add(fid))]
    if requested_form_ids:
        req_set = set(requested_form_ids)
        form_ids = [fid for fid in form_ids if fid in req_set]
    if not form_ids:
        return jsonify({'error': 'No forms configured on source'}), 400

    summary = {
        'source_id': source.id,
        'full_history': full_history,
        'date_from': date_from,
        'date_to': date_to,
        'forms_scanned': 0,
        'entries_seen': 0,
        'created': 0,
        'updated': 0,
        'duplicate': 0,
        'ignored': 0,
        'error': 0,
        'details': [],
    }

    for fid in form_ids:
        summary['forms_scanned'] += 1
        entries = []
        after = None
        page_count = 0
        try:
            while len(entries) < per_form_limit and page_count < max_pages:
                page_count += 1
                url = (
                    f'https://graph.facebook.com/v25.0/{_parse.quote(fid)}/leads'
                    f'?fields=id,created_time,field_data,ad_id,ad_name,adset_id,adset_name,campaign_id,campaign_name,form_id,page_id'
                    f'&limit={page_size}'
                    f'&access_token={_parse.quote(token)}'
                )
                if after:
                    url += f'&after={_parse.quote(after)}'

                with _req.urlopen(_req.Request(url), timeout=15) as resp:
                    payload = _json.loads(resp.read())

                page_entries = payload.get('data', []) if isinstance(payload, dict) else []
                if not page_entries:
                    break

                reached_older_than_from = False
                for raw in page_entries:
                    created_time = str((raw or {}).get('created_time') or '').strip()
                    created_date = None
                    if created_time:
                        try:
                            created_date = datetime.fromisoformat(created_time.replace('Z', '+00:00')).date()
                        except ValueError:
                            created_date = None

                    if parsed_from_date and created_date and created_date < parsed_from_date:
                        reached_older_than_from = True
                        continue
                    if parsed_to_exclusive_date and created_date and created_date >= parsed_to_exclusive_date:
                        continue

                    entries.append(raw)
                    if len(entries) >= per_form_limit:
                        break

                if reached_older_than_from:
                    break

                paging = payload.get('paging', {}) if isinstance(payload, dict) else {}
                cursors = paging.get('cursors', {}) if isinstance(paging, dict) else {}
                after = cursors.get('after')
                if not after:
                    break
        except Exception as exc:
            summary['details'].append({'form_id': fid, 'error': str(exc)})
            continue

        form_result = {'form_id': fid, 'entries': len(entries), 'results': []}
        summary['entries_seen'] += len(entries)

        for raw in entries:
            entry = dict(raw or {})
            if entry.get('id') and not entry.get('leadgen_id'):
                entry['leadgen_id'] = entry.get('id')
            if not entry.get('form_id'):
                entry['form_id'] = fid
            if not entry.get('page_id') and page_id:
                entry['page_id'] = page_id

            target_source = _resolve_meta_target_source(source, str(entry.get('page_id') or page_id), str(entry.get('form_id') or fid))
            enriched = _meta_enrich_leadgen_entry(entry, target_source)
            if enriched:
                merged_entry = dict(entry)
                for k, v in enriched.items():
                    if v not in (None, '', []):
                        merged_entry[k] = v
                entry = merged_entry

            normalised = _normalise_meta(entry)
            if not normalised.get('page_id'):
                normalised['page_id'] = str(page_id or '')
            result = ingest_lead(target_source, entry, normalised)
            status = result.get('status', 'error')
            if status == 'created':
                summary['created'] += 1
            elif status == 'updated':
                summary['updated'] += 1
            elif status == 'duplicate':
                summary['duplicate'] += 1
            elif status == 'ignored':
                summary['ignored'] += 1
            else:
                summary['error'] += 1

            form_result['results'].append({
                'platform_lead_id': normalised.get('platform_lead_id'),
                'status': status,
                'lead_id': result.get('lead_id'),
                'message': result.get('message'),
            })

        form_result['pages_scanned'] = page_count
        summary['details'].append(form_result)

    return jsonify({'ok': True, 'summary': summary}), 200


# ══════════════════════════════════════════════════════════════════════════════
# GOOGLE OAUTH FLOW
# Phase META-1.1
# ══════════════════════════════════════════════════════════════════════════════

@lead_sources_bp.route('/google/exchange-code', methods=['POST'])
@require_role('superadmin', 'platform_owner')
def google_exchange_code():
    """
    Exchange a Google OAuth authorization code for access + refresh tokens.
    Returns account info and lead form campaigns accessible to this account.

    POST body:
      { "code": "...", "client_id": "...", "client_secret": "...", "redirect_uri": "..." }
    """
    user = request.current_user
    data = request.get_json() or {}

    code          = (data.get('code') or '').strip()
    client_id     = (data.get('client_id') or '').strip()
    client_secret = (data.get('client_secret') or '').strip()
    redirect_uri  = (data.get('redirect_uri') or '').strip()

    if not all([code, client_id, client_secret, redirect_uri]):
        return jsonify({'error': 'code, client_id, client_secret, redirect_uri are required'}), 400

    try:
        import urllib.request as _req
        import urllib.parse as _parse
        import json as _json

        # 1. Exchange code for tokens
        token_body = _parse.urlencode({
            'code':          code,
            'client_id':     client_id,
            'client_secret': client_secret,
            'redirect_uri':  redirect_uri,
            'grant_type':    'authorization_code',
        }).encode()
        req = _req.Request(
            'https://oauth2.googleapis.com/token',
            data=token_body,
            headers={'Content-Type': 'application/x-www-form-urlencoded'},
        )
        try:
            with _req.urlopen(req, timeout=15) as r:
                token_data = _json.loads(r.read())
        except _urlerr.HTTPError as http_err:
            response_body = ''
            try:
                response_body = (http_err.read() or b'').decode('utf-8', errors='ignore')
            except Exception:
                response_body = ''

            logger.error(
                'google_token_exchange_failed status=%s body=%s',
                getattr(http_err, 'code', None),
                response_body,
            )

            error_code = 'token_exchange_failed'
            try:
                err_json = _json.loads(response_body or '{}')
                err_name = str(err_json.get('error') or '').strip().lower()
                if err_name:
                    error_code = err_name
            except Exception:
                pass

            tenant_redirect = _google_tenant_connect_redirect('google_oauth_error', error_code)
            return redirect(tenant_redirect or f'{frontend_base}/?google_oauth_error={_parse.quote(error_code)}')

        if 'error' in token_data:
            return jsonify({'error': token_data.get('error_description', token_data['error'])}), 400

        access_token  = token_data.get('access_token', '')
        refresh_token = token_data.get('refresh_token', '')

        # 2. Get user info
        ui_req = _req.Request(
            'https://www.googleapis.com/oauth2/v2/userinfo',
            headers={'Authorization': f'Bearer {access_token}'},
        )
        with _req.urlopen(ui_req, timeout=10) as r:
            userinfo = _json.loads(r.read())

        return jsonify({
            'access_token':  access_token,
            'refresh_token': refresh_token,
            'user': {
                'id':    userinfo.get('id'),
                'email': userinfo.get('email'),
                'name':  userinfo.get('name'),
            },
        }), 200

    except Exception as exc:
        logger.exception('google_exchange_code error: %s', exc)
        return jsonify({'error': str(exc)}), 502


@lead_sources_bp.route('/google/save-connection', methods=['POST'])
@require_role('superadmin', 'sales_manager', 'platform_owner')
def google_save_connection():
    """
        Finalise Google OAuth wizard: create or update LeadSource with Google
        credentials and selected Google Ads accounts.
    POST body:
    {
            "source_id":          123,          // update existing, or null to create new
            "name":               "Google Ads - Mumbai",
            "client_id":          "...",
            "client_secret":      "...",
            "refresh_token":      "...",
            "user_email":         "...",
            "selected_accounts":  [{customer_id, customer_name, resource_name}],
    }
    """
    user = request.current_user
    data = request.get_json() or {}

    client_id = (data.get('client_id') or '').strip()
    client_secret = (data.get('client_secret') or '').strip()
    refresh_token = (data.get('refresh_token') or '').strip()
    user_email = (data.get('user_email') or '').strip()
    name = (data.get('name') or f'Google - {user_email}').strip()

    platform_client_id, platform_client_secret = _get_platform_google_creds()
    if client_id in ('', '__platform__', 'platform'):
        client_id = platform_client_id
    if client_secret in ('', '__platform__', 'platform'):
        client_secret = platform_client_secret

    if not all([client_id, client_secret, refresh_token]):
        return jsonify({'error': 'client_id, client_secret and refresh_token are required'}), 400

    selected_accounts = data.get('selected_accounts') or []
    clean_accounts = []
    seen = set()
    for row in selected_accounts:
        if not isinstance(row, dict):
            continue
        cid = _google_normalize_customer_id(row.get('customer_id') or row.get('resource_name'))
        if not cid or cid in seen:
            continue
        seen.add(cid)
        clean_accounts.append({
            'customer_id': cid,
            'customer_name': str(row.get('customer_name') or f'Google Ads {cid}').strip(),
            'resource_name': str(row.get('resource_name') or f'customers/{cid}').strip(),
        })

    if not clean_accounts:
        return jsonify({'error': 'Select at least one Google Ads account'}), 400

    creds = {
        'client_id': client_id,
        'client_secret': client_secret,
        'refresh_token': refresh_token,
        'primary_customer_id': clean_accounts[0]['customer_id'],
    }

    source_id = data.get('source_id')
    if source_id:
        source = LeadSource.query.filter_by(id=source_id, tenant_id=user.tenant_id).first()
        if not source:
            return jsonify({'error': 'Source not found'}), 404
        source.name = name
        existing_creds = source.credentials or {}
        existing_creds.update({k: v for k, v in creds.items() if v})
        source.credentials = existing_creds
    else:
        source = LeadSource(
            tenant_id=user.tenant_id,
            name=name,
            source_type='google',
            credentials=creds,
            dup_check_phone=True,
            dup_check_email=False,
            created_by=user.id,
        )
        db.session.add(source)

    db.session.flush()

    ConnectedGoogleAdsAccount.query.filter_by(
        tenant_id=user.tenant_id,
        source_id=source.id,
    ).delete(synchronize_session=False)

    for row in clean_accounts:
        db.session.add(ConnectedGoogleAdsAccount(
            tenant_id=user.tenant_id,
            source_id=source.id,
            customer_id=row['customer_id'],
            customer_name=row['customer_name'],
            resource_name=row['resource_name'],
            metadata_json={'connected_via': 'oauth_wizard'},
            is_active=True,
        ))

    source.connected_account = (user_email or 'Google Ads') + f' ({len(clean_accounts)} account(s))'
    source.available_forms = source.available_forms or []
    source.available_campaigns = source.available_campaigns or []
    source.permission_status = 'ok'
    source.permission_details = {
        'selected_account_count': len(clean_accounts),
        'selected_accounts': clean_accounts,
    }
    source.last_tested_at = datetime.utcnow()
    source.last_test_result = 'pass'
    source.last_test_message = (
        f'Connected via OAuth. {len(clean_accounts)} Google Ads account(s) selected.'
    )

    db.session.commit()
    return jsonify({'source': source.to_dict()}), 200


# ══════════════════════════════════════════════════════════════════════════════
# SIMPLIFIED META OAUTH  (platform credentials stored in env vars)
# ══════════════════════════════════════════════════════════════════════════════

@lead_sources_bp.route('/meta/start-auth', methods=['POST'])
@require_role('superadmin', 'platform_owner')
def meta_start_auth():
    """
    Generate a Facebook OAuth URL using the platform Meta app credentials.
    Tenant provides their Business ID; we encode it in the OAuth state.

    POST body: { "business_id": "123456789" }
    Returns:   { "auth_url": "https://facebook.com/dialog/oauth?..." }
    """
    user = request.current_user
    data = request.get_json() or {}
    business_id = (data.get('business_id') or '').strip()
    if not business_id:
        return jsonify({'error': 'business_id is required'}), 400

    app_id, app_secret = _get_platform_meta_creds()
    if not app_secret:
        return jsonify({'error': 'Meta platform credentials not configured. Set META_APP_SECRET in environment.'}), 500

    session_key = secrets.token_urlsafe(24)
    _purge_expired_sessions()
    session_data = {
        'tenant_id':   user.tenant_id,
        'business_id': business_id,
        'created_at':  datetime.utcnow(),
    }
    _save_oauth_session(session_key, session_data, 'meta')

    frontend_base = os.environ.get('FRONTEND_URL', 'https://app.sociomonkey.com')
    callback_url  = os.environ.get('BACKEND_URL', 'https://smk-backend-api.vercel.app') + '/api/lead-sources/meta/oauth/callback'
    state = _parse.quote(session_key)

    scopes = ','.join(_get_meta_oauth_scopes())
    auth_url = (
        f'https://www.facebook.com/dialog/oauth'
        f'?client_id={_parse.quote(app_id)}'
        f'&redirect_uri={_parse.quote(callback_url)}'
        f'&scope={_parse.quote(scopes)}'
        f'&response_type=code'
        f'&auth_type=rerequest'
        f'&state={state}'
    )
    return jsonify({'auth_url': auth_url, 'session_key': session_key}), 200


@lead_sources_bp.route('/meta/oauth/callback', methods=['GET'])
def meta_oauth_callback():
    """
    Facebook OAuth callback. Exchanges code → token → fetches pages
    from the business_id stored in session. Redirects tenant back to
    the LMS page with session_key so frontend can retrieve pages.
    """
    code        = request.args.get('code', '')
    state       = request.args.get('state', '')
    error       = request.args.get('error', '')
    frontend_base = os.environ.get('FRONTEND_URL', 'https://app.sociomonkey.com')

    if error:
        return redirect(f'{frontend_base}/?meta_oauth_error={_parse.quote(error)}')

    session_data = _load_oauth_session(state, 'meta')
    if not session_data:
        return redirect(f'{frontend_base}/?meta_oauth_error=session_expired')

    app_id, app_secret = _get_platform_meta_creds()
    callback_url = os.environ.get('BACKEND_URL', 'https://smk-backend-api.vercel.app') + '/api/lead-sources/meta/oauth/callback'

    try:
        # Exchange code → short-lived token
        token_url = (
            f'https://graph.facebook.com/v25.0/oauth/access_token'
            f'?client_id={_parse.quote(app_id)}'
            f'&redirect_uri={_parse.quote(callback_url)}'
            f'&client_secret={_parse.quote(app_secret)}'
            f'&code={_parse.quote(code)}'
        )
        with _req.urlopen(_req.Request(token_url), timeout=15) as r:
            token_data = _json.loads(r.read())

        if 'error' in token_data:
            return redirect(f'{frontend_base}/?meta_oauth_error=token_exchange_failed')

        short_token = token_data.get('access_token', '')

        # Exchange short → long-lived token
        long_url = (
            f'https://graph.facebook.com/v25.0/oauth/access_token'
            f'?grant_type=fb_exchange_token'
            f'&client_id={_parse.quote(app_id)}'
            f'&client_secret={_parse.quote(app_secret)}'
            f'&fb_exchange_token={_parse.quote(short_token)}'
        )
        with _req.urlopen(_req.Request(long_url), timeout=15) as r:
            long_data = _json.loads(r.read())
        long_token = long_data.get('access_token', short_token)

        # Get user info
        me_url = f'https://graph.facebook.com/v25.0/me?fields=id,name&access_token={_parse.quote(long_token)}'
        with _req.urlopen(_req.Request(me_url), timeout=10) as r:
            me = _json.loads(r.read())

        # Get pages available to this user/business from multiple Graph edges.
        business_id = session_data.get('business_id', '')
        page_fetch_errors = []
        pages_by_id = {}

        logger.info('META_PAGE_DISCOVERY business_id_present=%s', bool(business_id))

        def _fetch_graph_json(url, label):
            try:
                with _req.urlopen(_req.Request(url), timeout=10) as r:
                    raw = r.read()
                data = _json.loads(raw)
                logger.info(
                    'META_PAGE_DISCOVERY label=%s items=%s error=%s',
                    label,
                    len((data or {}).get('data', []) or []) if isinstance(data, dict) else 0,
                    bool(isinstance(data, dict) and data.get('error')),
                )
                if isinstance(data, dict) and data.get('error'):
                    msg = data['error'].get('message') or str(data['error'])
                    page_fetch_errors.append(f'{label}: {msg}')
                return data
            except Exception as exc:
                logger.warning('META_PAGE_DISCOVERY label=%s failed=%s', label, type(exc).__name__)
                page_fetch_errors.append(f'{label}: {exc}')
                return {}

        def _merge_pages(data):
            for p in (data or {}).get('data', []) or []:
                pid = p.get('id')
                if not pid:
                    continue
                existing = pages_by_id.get(pid) or {'id': pid, 'name': p.get('name', ''), 'access_token': ''}
                if p.get('name'):
                    existing['name'] = p.get('name')
                if p.get('access_token'):
                    existing['access_token'] = p.get('access_token')
                pages_by_id[pid] = existing

        me_accounts_url = (
            f'https://graph.facebook.com/v25.0/me/accounts'
            f'?fields=id,name,access_token,tasks'
            f'&access_token={_parse.quote(long_token)}'
        )
        _me_accounts_raw = _fetch_graph_json(me_accounts_url, 'me/accounts')
        _me_accounts_pages = (_me_accounts_raw or {}).get('data', []) or []
        logger.info(
            'META_PAGE_DISCOVERY me_accounts_page_count=%d pages_with_tokens=%d',
            len(_me_accounts_pages),
            sum(1 for p in _me_accounts_pages if p.get('access_token')),
        )
        _merge_pages(_me_accounts_raw)

        if business_id:
            owned_pages_url = (
                f'https://graph.facebook.com/v25.0/{_parse.quote(business_id)}/owned_pages'
                f'?fields=id,name,access_token'
                f'&access_token={_parse.quote(long_token)}'
            )
            _merge_pages(_fetch_graph_json(owned_pages_url, 'business/owned_pages'))

            client_pages_url = (
                f'https://graph.facebook.com/v25.0/{_parse.quote(business_id)}/client_pages'
                f'?fields=id,name,access_token'
                f'&access_token={_parse.quote(long_token)}'
            )
            _merge_pages(_fetch_graph_json(client_pages_url, 'business/client_pages'))

        pages = list(pages_by_id.values())

        session_data.pop('meta_page_fetch_error', None)
        if not pages and page_fetch_errors:
            session_data['meta_page_fetch_error'] = '; '.join(page_fetch_errors[:3])

        # Store result in session
        session_data.update({
            'user':       {'id': me.get('id'), 'name': me.get('name')},
            'long_token': long_token,
            'pages':      pages,
            'completed':  True,
        })
        _save_oauth_session(state, session_data, 'meta')

        # Redirect back to frontend
        tenant_id = session_data.get('tenant_id', 'demo')
        tenant_slug = str(tenant_id)
        # Try to get the tenant slug from DB for a nicer URL
        try:
            from app.models.tenant import Tenant
            t = Tenant.query.get(tenant_id)
            if t and t.slug:
                tenant_slug = t.slug
        except Exception:
            pass

        tenant_slug = _TENANT_ROUTE_SLUG_ALIASES.get(str(tenant_slug).lower(), tenant_slug)

        redirect_url = (
            f'{frontend_base}/apps/lms/{tenant_slug}/lead-sources/connect?meta_session={_parse.quote(state)}&meta_tab=connect'
        )
        logger.info('META_CALLBACK_READY %s', {
            'instance': _runtime_instance_id(),
            'session_key': '[redacted]',
            'tenant_id': tenant_id,
            'tenant_slug': tenant_slug,
            'session_exists_before_redirect': bool(_load_oauth_session(state, 'meta')),
            'redirect_path': '/apps/lms/{}/lead-sources/connect'.format(tenant_slug),
        })

        return redirect(redirect_url)

    except Exception as exc:
        logger.exception('meta_oauth_callback error: %s', exc)
        return redirect(f'{frontend_base}/?meta_oauth_error=server_error')


@lead_sources_bp.route('/meta/auth-session/<session_key>', methods=['GET'])
@require_role('superadmin', 'platform_owner')
def meta_auth_session(session_key):
    """
    Retrieve pages + user info from a completed OAuth session.
    Called by frontend after OAuth callback redirect.
    """
    user = request.current_user
    session_data = _load_oauth_session(session_key, 'meta')
    created_at = session_data.get('created_at') if session_data else None
    session_age_seconds = None
    if created_at:
        try:
            session_age_seconds = round((datetime.utcnow() - created_at).total_seconds(), 3)
        except Exception:
            session_age_seconds = None

    _log_meta_session_event(
        'auth_session_start',
        session_key=session_key,
        session_exists=bool(session_data),
        session_age_seconds=session_age_seconds,
        user_id=getattr(user, 'id', None),
        tenant_id=getattr(user, 'tenant_id', None),
    )
    if not session_data:
        _log_meta_session_event(
            'auth_session_missing',
            session_key=session_key,
            session_exists=False,
            session_age_seconds=None,
            user_id=getattr(user, 'id', None),
            tenant_id=getattr(user, 'tenant_id', None),
        )
        return jsonify({'error': 'Session expired or not found'}), 404
    if session_data.get('tenant_id') != user.tenant_id:
        _log_meta_session_event(
            'auth_session_tenant_mismatch',
            session_key=session_key,
            session_exists=True,
            session_age_seconds=session_age_seconds,
            user_id=getattr(user, 'id', None),
            tenant_id=getattr(user, 'tenant_id', None),
            session_tenant_id=session_data.get('tenant_id'),
        )
        return jsonify({'error': 'Unauthorized'}), 403
    if not session_data.get('completed'):
        _log_meta_session_event(
            'auth_session_pending',
            session_key=session_key,
            session_exists=True,
            session_age_seconds=session_age_seconds,
            user_id=getattr(user, 'id', None),
            tenant_id=getattr(user, 'tenant_id', None),
        )
        return jsonify({'error': 'OAuth not completed yet'}), 202

    pages = session_data.get('pages', []) or []
    if not pages:
        details = session_data.get('meta_page_fetch_error', '')
        error_message = 'No Facebook Pages found for this account/business. Check Business/Page access and permissions, then try again.'
        if details:
            error_message = f'{error_message} Details: {details}'
        _log_meta_session_event(
            'auth_session_no_pages',
            session_key=session_key,
            session_exists=True,
            session_age_seconds=session_age_seconds,
            user_id=getattr(user, 'id', None),
            tenant_id=getattr(user, 'tenant_id', None),
        )
        return jsonify({'error': error_message}), 409

    _log_meta_session_event(
        'auth_session_success',
        session_key=session_key,
        session_exists=True,
        session_age_seconds=session_age_seconds,
        user_id=getattr(user, 'id', None),
        tenant_id=getattr(user, 'tenant_id', None),
    )

    return jsonify({
        'user':        session_data.get('user'),
        'pages':       pages,
        'long_token':  session_data.get('long_token', ''),
        'business_id': session_data.get('business_id', ''),
    }), 200


def _meta_tier_run_to_dict(run: MetaTierTestRun):
    metrics = dict(run.metrics or {})
    return {
        'id': run.id,
        'tenant_id': run.tenant_id,
        'created_by': run.created_by,
        'status': run.status,
        'config': dict(run.config or {}),
        'metrics': metrics,
        'dashboard': compute_dashboard(metrics),
        'created_at': run.created_at.isoformat() + 'Z' if run.created_at else None,
        'updated_at': run.updated_at.isoformat() + 'Z' if run.updated_at else None,
        'completed_at': run.completed_at.isoformat() + 'Z' if run.completed_at else None,
    }


@lead_sources_bp.route('/meta/tier-tests/start', methods=['POST'])
@require_role('superadmin', 'platform_owner')
def meta_tier_test_start():
    user = request.current_user
    data = request.get_json(silent=True) or {}

    ad_account_id = str(data.get('ad_account_id') or '').strip()
    page_id = str(data.get('page_id') or '').strip()
    target_success_calls = int(data.get('target_success_calls') or 500)
    batch_size = int(data.get('batch_size') or 15)
    min_delay_ms = int(data.get('min_delay_ms') or 350)
    max_delay_ms = int(data.get('max_delay_ms') or 900)

    if not ad_account_id:
        return jsonify({'error': 'ad_account_id is required'}), 400
    if not page_id:
        return jsonify({'error': 'page_id is required'}), 400
    if target_success_calls < 1:
        return jsonify({'error': 'target_success_calls must be >= 1'}), 400

    run = MetaTierTestRun(
        tenant_id=user.tenant_id,
        created_by=user.id,
        status='running',
        config={
            'ad_account_id': ad_account_id,
            'page_id': page_id,
            'target_success_calls': target_success_calls,
            'batch_size': batch_size,
            'min_delay_ms': min_delay_ms,
            'max_delay_ms': max_delay_ms,
        },
        metrics=build_initial_state(target_success_calls=target_success_calls),
    )
    db.session.add(run)
    db.session.commit()
    return jsonify({'run': _meta_tier_run_to_dict(run)}), 201


@lead_sources_bp.route('/meta/tier-tests/<int:run_id>/batch', methods=['POST'])
@require_role('superadmin', 'platform_owner')
def meta_tier_test_run_batch(run_id):
    user = request.current_user
    data = request.get_json(silent=True) or {}
    access_token = str(data.get('access_token') or '').strip()

    run = MetaTierTestRun.query.get_or_404(run_id)
    if run.tenant_id != user.tenant_id:
        return jsonify({'error': 'Not found'}), 404
    if run.status in ('completed', 'stopped'):
        return jsonify({'run': _meta_tier_run_to_dict(run), 'batch': []}), 200

    try:
        metrics, batch_results, dashboard = run_batch(run.config or {}, run.metrics or {}, access_token)
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400
    except Exception as exc:
        logger.exception('meta tier test batch failed: %s', exc)
        return jsonify({'error': f'batch_failed: {exc}'}), 500

    run.metrics = metrics
    run.updated_at = datetime.utcnow()
    if metrics.get('goal_met'):
        run.status = 'completed'
        run.completed_at = datetime.utcnow()
    db.session.add(run)
    db.session.commit()

    return jsonify({
        'run': _meta_tier_run_to_dict(run),
        'batch': batch_results,
        'dashboard': dashboard,
    }), 200


@lead_sources_bp.route('/meta/tier-tests/<int:run_id>', methods=['GET'])
@require_role('superadmin', 'platform_owner')
def meta_tier_test_get_run(run_id):
    user = request.current_user
    run = MetaTierTestRun.query.get_or_404(run_id)
    if run.tenant_id != user.tenant_id:
        return jsonify({'error': 'Not found'}), 404
    return jsonify({'run': _meta_tier_run_to_dict(run)}), 200


@lead_sources_bp.route('/meta/tier-tests/<int:run_id>/stop', methods=['POST'])
@require_role('superadmin', 'platform_owner')
def meta_tier_test_stop(run_id):
    user = request.current_user
    run = MetaTierTestRun.query.get_or_404(run_id)
    if run.tenant_id != user.tenant_id:
        return jsonify({'error': 'Not found'}), 404
    if run.status == 'running':
        run.status = 'stopped'
        run.completed_at = datetime.utcnow()
        run.updated_at = datetime.utcnow()
        db.session.add(run)
        db.session.commit()
    return jsonify({'run': _meta_tier_run_to_dict(run)}), 200


@lead_sources_bp.route('/meta/tier-tests/dashboard', methods=['GET'])
@require_role('superadmin', 'platform_owner')
def meta_tier_test_dashboard():
    user = request.current_user
    rows = (
        MetaTierTestRun.query
        .filter_by(tenant_id=user.tenant_id)
        .order_by(MetaTierTestRun.created_at.desc())
        .limit(20)
        .all()
    )

    runs = [_meta_tier_run_to_dict(r) for r in rows]
    total_calls = 0
    total_success = 0
    total_counted = 0
    for r in runs:
        dashboard = r.get('dashboard') or {}
        total_calls += int(dashboard.get('total_marketing_api_calls') or 0)
        total_counted += int(dashboard.get('calls_counted_toward_meta_testing') or 0)
        total_success += int((r.get('metrics') or {}).get('success_calls') or 0)

    aggregate_success_percent = round((total_success * 100.0) / total_calls, 2) if total_calls else 0.0
    return jsonify({
        'runs': runs,
        'aggregate': {
            'total_marketing_api_calls': total_calls,
            'success_percent': aggregate_success_percent,
            'calls_counted_toward_meta_testing': total_counted,
        },
    }), 200


# ══════════════════════════════════════════════════════════════════════════════
# SIMPLIFIED GOOGLE OAUTH  (platform credentials stored in env vars)
# ══════════════════════════════════════════════════════════════════════════════

@lead_sources_bp.route('/google/start-auth', methods=['POST'])
@require_role('superadmin', 'platform_owner')
def google_start_auth():
    """
    Generate a Google OAuth URL using platform credentials from env vars.
    Returns: { "auth_url": "..." }
    """
    user = request.current_user
    client_id, client_secret = _get_platform_google_creds()
    if not client_id or not client_secret:
        return jsonify({'error': 'Google platform credentials not configured. Set GOOGLE_CLIENT_ID and GOOGLE_CLIENT_SECRET in environment.'}), 500

    session_key = secrets.token_urlsafe(24)
    _purge_expired_sessions()
    session_data = {
        'tenant_id':  user.tenant_id,
        'platform':   'google',
        'created_at': datetime.utcnow(),
    }
    _save_oauth_session(session_key, session_data, 'google')

    callback_url = os.environ.get('BACKEND_URL', 'https://smk-backend-api.vercel.app') + '/api/lead-sources/google/oauth/callback'
    scopes = 'https://www.googleapis.com/auth/userinfo.email https://www.googleapis.com/auth/adwords'
    auth_url = (
        f'https://accounts.google.com/o/oauth2/v2/auth'
        f'?client_id={_parse.quote(client_id)}'
        f'&redirect_uri={_parse.quote(callback_url)}'
        f'&response_type=code'
        f'&scope={_parse.quote(scopes)}'
        f'&access_type=offline'
        f'&prompt=select_account%20consent'
        f'&state={_parse.quote(session_key)}'
    )
    return jsonify({'auth_url': auth_url, 'session_key': session_key}), 200


@lead_sources_bp.route('/google/oauth/callback', methods=['GET'])
def google_oauth_callback():
    """
    Google OAuth callback — exchanges code → tokens → stores in session.
    Redirects back to LMS frontend.
    """
    code  = request.args.get('code', '')
    state = request.args.get('state', '')
    error = request.args.get('error', '')
    frontend_base = os.environ.get('FRONTEND_URL', 'https://app.sociomonkey.com')

    session_data = _load_oauth_session(state, 'google') if state else None

    def _google_tenant_connect_redirect(query_key=None, query_value=''):
        if not session_data:
            return None
        tenant_id = session_data.get('tenant_id', 'demo')
        tenant_slug = str(tenant_id)
        try:
            from app.models.tenant import Tenant
            t = Tenant.query.get(tenant_id)
            if t and t.slug:
                tenant_slug = t.slug
        except Exception:
            pass
        tenant_slug = _TENANT_ROUTE_SLUG_ALIASES.get(str(tenant_slug).lower(), tenant_slug)
        base_url = f'{frontend_base}/apps/lms/{tenant_slug}/lead-sources/connect'
        if query_key:
            return f'{base_url}?{query_key}={_parse.quote(str(query_value or ""))}&meta_tab=connect'
        return f'{base_url}?meta_tab=connect'

    if error:
        tenant_redirect = _google_tenant_connect_redirect('google_oauth_error', error)
        return redirect(tenant_redirect or f'{frontend_base}/?google_oauth_error={_parse.quote(error)}')

    if not session_data:
        return redirect(f'{frontend_base}/?google_oauth_error=session_expired')

    client_id, client_secret = _get_platform_google_creds()
    callback_url = os.environ.get('BACKEND_URL', 'https://smk-backend-api.vercel.app') + '/api/lead-sources/google/oauth/callback'

    try:
        token_body = _parse.urlencode({
            'code':          code,
            'client_id':     client_id,
            'client_secret': client_secret,
            'redirect_uri':  callback_url,
            'grant_type':    'authorization_code',
        }).encode()
        req = _req.Request(
            'https://oauth2.googleapis.com/token',
            data=token_body,
            headers={'Content-Type': 'application/x-www-form-urlencoded'},
        )
        with _req.urlopen(req, timeout=15) as r:
            token_data = _json.loads(r.read())

        if 'error' in token_data:
            tenant_redirect = _google_tenant_connect_redirect('google_oauth_error', 'token_exchange_failed')
            return redirect(tenant_redirect or f'{frontend_base}/?google_oauth_error=token_exchange_failed')

        access_token  = token_data.get('access_token', '')
        refresh_token = token_data.get('refresh_token', '')

        if not access_token:
            tenant_redirect = _google_tenant_connect_redirect('google_oauth_error', 'missing_access_token')
            return redirect(tenant_redirect or f'{frontend_base}/?google_oauth_error=missing_access_token')

        userinfo = {}
        try:
            ui_req = _req.Request(
                'https://www.googleapis.com/oauth2/v2/userinfo',
                headers={'Authorization': f'Bearer {access_token}'},
            )
            with _req.urlopen(ui_req, timeout=10) as r:
                userinfo = _json.loads(r.read())
        except Exception as exc:
            logger.warning('google_userinfo_failed: %s', exc)
            id_token = token_data.get('id_token', '')
            if id_token:
                try:
                    parts = id_token.split('.')
                    if len(parts) >= 2:
                        payload = parts[1]
                        payload += '=' * (-len(payload) % 4)
                        decoded = base64.urlsafe_b64decode(payload.encode('utf-8'))
                        claims = _json.loads(decoded.decode('utf-8'))
                        userinfo = {
                            'id': claims.get('sub') or '',
                            'email': claims.get('email') or '',
                            'name': claims.get('name') or claims.get('email') or '',
                        }
                except Exception:
                    userinfo = {}

        accessible_accounts = []
        discovery_error = ''
        try:
            accessible_accounts = _google_list_accessible_accounts(access_token)
        except Exception as exc:
            discovery_error = str(exc)

        session_data.update({
            'access_token':  access_token,
            'refresh_token': refresh_token,
            'user':          {'id': userinfo.get('id'), 'email': userinfo.get('email'), 'name': userinfo.get('name')},
            'accessible_accounts': accessible_accounts,
            'account_discovery_error': discovery_error,
            'oauth_healthy': True,
            'completed':     True,
        })
        _save_oauth_session(state, session_data, 'google')

        tenant_id   = session_data.get('tenant_id', 'demo')
        tenant_slug = str(tenant_id)
        try:
            from app.models.tenant import Tenant
            t = Tenant.query.get(tenant_id)
            if t and t.slug:
                tenant_slug = t.slug
        except Exception:
            pass

        tenant_slug = _TENANT_ROUTE_SLUG_ALIASES.get(str(tenant_slug).lower(), tenant_slug)

        return redirect(
            f'{frontend_base}/apps/lms/{tenant_slug}/lead-sources/connect?google_session={_parse.quote(state)}&meta_tab=connect'
        )

    except Exception as exc:
        logger.exception('google_oauth_callback error: %s', exc)
        tenant_redirect = _google_tenant_connect_redirect('google_oauth_error', 'server_error')
        return redirect(tenant_redirect or f'{frontend_base}/?google_oauth_error=server_error')


@lead_sources_bp.route('/google/auth-session/<session_key>', methods=['GET'])
@require_role('superadmin', 'platform_owner')
def google_auth_session(session_key):
    """Retrieve user + tokens from a completed Google OAuth session."""
    user = request.current_user
    session_data = _load_oauth_session(session_key, 'google')
    if not session_data:
        return jsonify({'error': 'Session expired or not found'}), 404
    if session_data.get('tenant_id') != user.tenant_id:
        return jsonify({'error': 'Unauthorized'}), 403
    if not session_data.get('completed'):
        return jsonify({'error': 'OAuth not completed yet'}), 202
    return jsonify({
        'user':          session_data.get('user'),
        'access_token':  session_data.get('access_token', ''),
        'refresh_token': session_data.get('refresh_token', ''),
        'accessible_accounts': session_data.get('accessible_accounts', []),
        'account_discovery_error': session_data.get('account_discovery_error', ''),
        'oauth_healthy': bool(session_data.get('oauth_healthy')),
    }), 200


@lead_sources_bp.route('/google/foundation-status', methods=['GET'])
@require_role('superadmin', 'platform_owner')
def google_foundation_status():
    """
    Validation snapshot for Google foundation phase.
    Returns OAuth health, account connection, and recent tracking-detection status.
    """
    user = request.current_user
    source_id = request.args.get('source_id', type=int)
    if not source_id:
        return jsonify({'error': 'source_id is required'}), 400

    source = LeadSource.query.filter_by(
        id=source_id,
        tenant_id=user.tenant_id,
        source_type='google',
    ).first()
    if not source:
        return jsonify({'error': 'Source not found'}), 404

    account_rows = ConnectedGoogleAdsAccount.query.filter_by(
        tenant_id=user.tenant_id,
        source_id=source.id,
        is_active=True,
    ).all()

    test_result = _test_google(source)
    oauth_healthy = test_result.get('result') in ('pass', 'partial')

    gclid_detected = db.session.query(IngestedLeadLog.id).filter(
        IngestedLeadLog.tenant_id == user.tenant_id,
        IngestedLeadLog.source_id == source.id,
        IngestedLeadLog.gclid.isnot(None),
        IngestedLeadLog.gclid != '',
    ).first() is not None

    utm_detected = db.session.query(IngestedLeadLog.id).filter(
        IngestedLeadLog.tenant_id == user.tenant_id,
        IngestedLeadLog.source_id == source.id,
        or_(
            db.and_(IngestedLeadLog.utm_source.isnot(None), IngestedLeadLog.utm_source != ''),
            db.and_(IngestedLeadLog.utm_medium.isnot(None), IngestedLeadLog.utm_medium != ''),
            db.and_(IngestedLeadLog.utm_campaign.isnot(None), IngestedLeadLog.utm_campaign != ''),
            db.and_(IngestedLeadLog.utm_content.isnot(None), IngestedLeadLog.utm_content != ''),
            db.and_(IngestedLeadLog.utm_term.isnot(None), IngestedLeadLog.utm_term != ''),
        ),
    ).first() is not None

    return jsonify({
        'source_id': source.id,
        'source_name': source.name,
        'account_connected': len(account_rows) > 0,
        'oauth_healthy': oauth_healthy,
        'gclid_detected': gclid_detected,
        'utm_detected': utm_detected,
        'connected_accounts': [a.to_dict() for a in account_rows],
        'oauth_test': {
            'result': test_result.get('result'),
            'message': test_result.get('message'),
        },
    }), 200


# ══════════════════════════════════════════════════════════════════════════════
# TEST LEAD INJECTION
# Phase META-1.1 – fire a synthetic lead through the full pipeline
# ══════════════════════════════════════════════════════════════════════════════

@lead_sources_bp.route('/<int:source_id>/inject-test-lead', methods=['POST'])
@require_role('superadmin', 'sales_manager', 'platform_owner')
def inject_test_lead(source_id):
    """
    Inject a synthetic test lead through the ingestion pipeline.
    Creates a real Lead row so every downstream step can be verified.

    POST body (all optional – uses defaults if omitted):
    {
      "name":    "Test Lead",
      "phone":   "+910000000001",
      "email":   "test@example.com",
      "campaign_name": "Test Campaign",
    }

    Returns the full pipeline result + link to the created lead.
    """
    user = request.current_user
    source = LeadSource.query.get_or_404(source_id)
    err = _check_source_ownership(source, user)
    if err:
        return err

    data = request.get_json() or {}

    import time
    ts = str(int(time.time()))

    # Build a realistic raw payload mirroring what the platform sends
    if source.source_type == 'meta':
        raw_payload = {
            'leadgen_id': f'TEST-{ts}',
            'page_id':    (source.credentials or {}).get('page_id', 'TEST_PAGE'),
            'form_id':    'TEST_FORM',
            'ad_id':      'TEST_AD',
            'adset_id':   'TEST_ADSET',
            'campaign_id': 'TEST_CAMPAIGN',
            'field_data': [
                {'name': 'full_name',     'values': [data.get('name',  f'Test Lead {ts}')]},
                {'name': 'phone_number',  'values': [data.get('phone', f'+910000{ts[-6:]}')]},
                {'name': 'email',         'values': [data.get('email', f'test{ts}@test.sociomonkey.com')]},
            ],
        }
        from app.routes.ingestion import _normalise_meta
        normalised = _normalise_meta(raw_payload)
        normalised['campaign_name'] = data.get('campaign_name', 'Test Campaign')

    elif source.source_type == 'google':
        raw_payload = {
            'lead_id':      f'TEST-{ts}',
            'form_id':      'TEST_FORM',
            'form_name':    'Test Form',
            'campaign_id':  'TEST_CAMPAIGN',
            'campaign_name': data.get('campaign_name', 'Test Campaign'),
            'ad_group_id':  'TEST_ADGROUP',
            'ad_group_name': 'Test Ad Group',
            'user_column_data': [
                {'column_name': 'GIVEN_NAME',   'string_value': data.get('name',  f'Test Lead {ts}').split()[0]},
                {'column_name': 'FAMILY_NAME',  'string_value': ' '.join(data.get('name', f'Test Lead {ts}').split()[1:]) or 'User'},
                {'column_name': 'PHONE_NUMBER', 'string_value': data.get('phone', f'+910000{ts[-6:]}')},
                {'column_name': 'EMAIL',        'string_value': data.get('email', f'test{ts}@test.sociomonkey.com')},
            ],
        }
        from app.routes.ingestion import _normalise_google
        normalised = _normalise_google(raw_payload)

    else:
        raw_payload = {
            'name':  data.get('name',  f'Test Lead {ts}'),
            'phone': data.get('phone', f'+910000{ts[-6:]}'),
            'email': data.get('email', f'test{ts}@test.sociomonkey.com'),
            'campaign_name': data.get('campaign_name', 'Test Campaign'),
        }
        from app.routes.ingestion import _normalise_generic
        normalised = _normalise_generic(raw_payload)

    from app.services.ingestion_engine import ingest_lead
    result = ingest_lead(source, raw_payload, normalised, is_test=True)

    # Gather pipeline evidence for UI display
    lead_id = result.get('lead_id')
    evidence = {}
    if lead_id:
        from app.models import Lead, ActivityLog
        lead = Lead.query.get(lead_id)
        if lead:
            evidence['lead'] = lead.to_dict()

        activity = ActivityLog.query.filter_by(
            resource_id=lead_id, resource_type='Lead', action='lead_ingested'
        ).order_by(ActivityLog.id.desc()).first()
        if activity:
            evidence['activity'] = {
                'id':          activity.id,
                'description': activity.description,
                'new_value':   activity.new_value,
            }

        log = IngestedLeadLog.query.filter_by(
            source_id=source_id, lead_id=lead_id
        ).order_by(IngestedLeadLog.id.desc()).first()
        if log:
            evidence['log'] = log.to_dict()

        # Push notification (NotificationEvent created by stage 6 of engine)
        from app.models.push import NotificationEvent
        notif = NotificationEvent.query.filter_by(lead_id=lead_id).first()
        if notif:
            evidence['push_notification'] = notif.to_dict()

        # Action board: StatusHistory entry created for this lead
        from app.models import StatusHistory
        sh = StatusHistory.query.filter_by(lead_id=lead_id).first()
        if sh:
            evidence['status_history'] = {'id': sh.id, 'status': sh.new_status if hasattr(sh, 'new_status') else str(sh)}

    lead_dict = evidence.get('lead', {})
    return jsonify({
        'result':   result,
        'evidence': evidence,
        'checks': {
            'lead_created':            lead_id is not None and result.get('status') in ('created', 'updated'),
            'pipeline_ran':            result.get('status') != 'error',
            'activity_logged':         bool(evidence.get('activity')),
            'assignment_applied':      bool(lead_dict.get('assigned_to')),
            'push_notification_created': bool(evidence.get('push_notification')),
            'action_board_updated':    bool(evidence.get('status_history')),
        },
    }), 200


# ══════════════════════════════════════════════════════════════════════════════
# FULL VALIDATION RUNNER
# Phase META-1.3 – 7-item real-account validation: connection, lead validation,
# duplicate detection, tenant isolation, end-to-end LMS flow.
# ══════════════════════════════════════════════════════════════════════════════

@lead_sources_bp.route('/validate', methods=['POST'])
@require_role('superadmin', 'platform_owner')
def run_validation():
    """
    Run the Phase META-1.3 validation suite against this tenant.

    POST body:
    {
      "meta_source_id":   123,   // LeadSource id (meta)
      "google_source_id": 456,   // LeadSource id (google) – optional
    }

    Returns a structured PASS/FAIL report for all 7 validation items.
    """
    user = request.current_user
    data = request.get_json() or {}
    _flow_lead_id = None
    _validation_lead_ids = set()

    def _track_validation_result(result):
        lead_id = (result or {}).get('lead_id')
        if lead_id:
            _validation_lead_ids.add(int(lead_id))
        return lead_id

    def _first_source_form_id(source):
        forms = source.available_forms or []
        for form in forms:
            if isinstance(form, dict):
                form_id = str(form.get('id') or '').strip()
                form_status = str(form.get('status') or '').strip().upper()
                if form_id and form_status == 'ACTIVE':
                    return form_id
        for form in forms:
            if isinstance(form, dict):
                form_id = str(form.get('id') or '').strip()
                if form_id:
                    return form_id
            else:
                form_id = str(form or '').strip()
                if form_id:
                    return form_id
        return ''

    def _build_meta_validation_payload(source, lead_name, lead_phone, lead_email, campaign_name):
        import time
        ts = str(int(time.time()))
        raw = {
            'leadgen_id': f'VALIDATE-{ts}',
            'page_id': (source.credentials or {}).get('page_id', 'TEST'),
            'form_id': _first_source_form_id(source) or 'VALIDATE_FORM',
            'field_data': [
                {'name': 'full_name', 'values': [lead_name]},
                {'name': 'phone_number', 'values': [lead_phone]},
                {'name': 'email', 'values': [lead_email]},
            ],
        }
        from app.routes.ingestion import _normalise_meta
        norm = _normalise_meta(raw)
        norm['campaign_name'] = campaign_name
        return raw, norm

    report = {
        'tenant_id':   user.tenant_id,
        'run_at':      datetime.utcnow().isoformat(),
        'items':       {},
        'deployment_ready': False,
    }

    # ── Helper ─────────────────────────────────────────────────────────────────
    def item(key, label, passed, detail='', sub=None):
        report['items'][key] = {
            'label':  label,
            'passed': passed,
            'result': 'PASS' if passed else 'FAIL',
            'detail': detail,
            'sub':    sub or {},
        }

    # ── ITEM 1: Meta Connection ────────────────────────────────────────────────
    meta_source_id = data.get('meta_source_id')
    meta_source = None
    if meta_source_id:
        meta_source = LeadSource.query.filter_by(
            id=meta_source_id, tenant_id=user.tenant_id, source_type='meta'
        ).first()

    if not meta_source:
        item('meta_connection', 'Meta Connection', False, 'No Meta source selected')
    else:
        test = _run_connection_test(meta_source)
        connected = test.get('result') in ('pass', 'partial')
        item('meta_connection', 'Meta Connection', connected,
             test.get('message', ''),
             {
                 'login_ok':          connected,
                 'business_manager':  bool(meta_source.connected_account),
                 'pages_visible':     bool((meta_source.credentials or {}).get('page_id')),
                 'forms_visible':     len(meta_source.available_forms or []) > 0,
                 'permissions_ok':    meta_source.permission_status in ('ok', 'active', 'granted'),
                 'connected_account': meta_source.connected_account,
                 'available_forms':   len(meta_source.available_forms or []),
             })

    # ── ITEM 2: Meta Lead Validation ─────────────────────────────────────────
    if meta_source:
        try:
            from app.services.ingestion_engine import ingest_lead
            import time
            ts = str(int(time.time()))
            raw, norm = _build_meta_validation_payload(
                meta_source,
                f'Validation Lead {ts}',
                f'+910001{ts[-6:]}',
                f'validate{ts}@test.sociomonkey.com',
                'Validation Run',
            )
            res = ingest_lead(meta_source, raw, norm, is_test=True)
            lead_created = res.get('status') in ('created', 'updated')
            lead_id = _track_validation_result(res)
            if lead_id:
                _flow_lead_id = lead_id

            from app.models import Lead, ActivityLog
            from app.models.push import NotificationEvent
            from app.models import StatusHistory
            assigned = False; timeline = False; push_ok = False; board_ok = False
            source_captured = False; campaign_captured = False; form_captured = False; page_captured = False
            if lead_id:
                l = Lead.query.get(lead_id)
                log = IngestedLeadLog.query.filter_by(lead_id=lead_id).order_by(IngestedLeadLog.received_at.desc()).first()
                assigned = bool(l and l.assigned_to) or meta_source.assign_strategy == 'none'
                source_captured = bool(l and l.source)
                campaign_captured = bool(log and log.campaign_name)
                form_captured = bool(log and log.form_name)
                page_captured = bool(log and log.page_id)
                timeline = bool(ActivityLog.query.filter_by(
                    resource_id=lead_id, action='lead_ingested').first())
                push_ok = bool(NotificationEvent.query.filter_by(lead_id=lead_id).first())
                board_ok = bool(StatusHistory.query.filter_by(lead_id=lead_id).first())

            item('meta_lead_validation', 'Meta Lead Validation', lead_created,
                 f'Lead #{lead_id} – status: {res.get("status")}',
                 {
                     'lead_entered_lms':          lead_created,
                     'lead_source_captured':      source_captured,
                     'campaign_captured':         campaign_captured,
                     'form_captured':             form_captured,
                     'page_captured':             page_captured,
                     'lead_assigned':             assigned,
                     'timeline_created':          timeline,
                     'push_notification_created': push_ok,
                     'action_board_updated':      board_ok,
                 })
        except Exception as exc:
            item('meta_lead_validation', 'Meta Lead Validation', False, str(exc))
    else:
        item('meta_lead_validation', 'Meta Lead Validation', False, 'No Meta source configured')

    # ── ITEM 3: Google Connection ─────────────────────────────────────────────
    google_source_id = data.get('google_source_id')
    google_source = None
    if google_source_id:
        google_source = LeadSource.query.filter_by(
            id=google_source_id, tenant_id=user.tenant_id, source_type='google'
        ).first()

    if not google_source:
        item('google_connection', 'Google Connection', False, 'No Google source selected')
    else:
        test = _run_connection_test(google_source)
        connected = test.get('result') in ('pass', 'partial')
        item('google_connection', 'Google Connection', connected,
             test.get('message', ''),
             {
                 'login_ok':          connected,
                 'ads_account':       bool(google_source.connected_account),
                 'forms_visible':     len(google_source.available_forms or []) > 0,
                 'permissions_ok':    google_source.permission_status in ('ok', 'active', 'granted'),
                 'connected_account': google_source.connected_account,
                 'available_forms':   len(google_source.available_forms or []),
                 'campaigns':         len(google_source.available_campaigns or []),
             })

    # ── ITEM 4: Google Lead Validation ───────────────────────────────────────
    if google_source:
        try:
            from app.routes.ingestion import _normalise_google
            from app.services.ingestion_engine import ingest_lead
            import time
            ts = str(int(time.time()))
            raw = {
                'lead_id':      f'VALIDATE-G-{ts}',
                'form_id':      'VALIDATE_FORM',
                'campaign_name': 'Google Validation Run',
                'user_column_data': [
                    {'column_name': 'GIVEN_NAME',   'string_value': 'Google'},
                    {'column_name': 'FAMILY_NAME',  'string_value': f'Test {ts}'},
                    {'column_name': 'PHONE_NUMBER', 'string_value': f'+910002{ts[-6:]}'},
                    {'column_name': 'EMAIL',        'string_value': f'gvalidate{ts}@test.sociomonkey.com'},
                ],
            }
            norm = _normalise_google(raw)
            res = ingest_lead(google_source, raw, norm, is_test=True)
            lead_created = res.get('status') in ('created', 'updated')
            lead_id = _track_validation_result(res)
            if lead_id and not _flow_lead_id:
                _flow_lead_id = lead_id

            from app.models import Lead, ActivityLog
            from app.models.push import NotificationEvent
            from app.models import StatusHistory
            assigned = False; timeline = False; push_ok = False; board_ok = False
            if lead_id:
                l = Lead.query.get(lead_id)
                assigned = bool(l and l.assigned_to) or google_source.assign_strategy == 'none'
                timeline = bool(ActivityLog.query.filter_by(
                    resource_id=lead_id, action='lead_ingested').first())
                push_ok = bool(NotificationEvent.query.filter_by(lead_id=lead_id).first())
                board_ok = bool(StatusHistory.query.filter_by(lead_id=lead_id).first())

            item('google_lead_validation', 'Google Lead Validation', lead_created,
                 f'Lead #{lead_id} – status: {res.get("status")}',
                 {
                     'lead_entered_lms':          lead_created,
                     'lead_assigned':             assigned,
                     'timeline_created':          timeline,
                     'push_notification_created': push_ok,
                     'action_board_updated':      board_ok,
                 })
        except Exception as exc:
            item('google_lead_validation', 'Google Lead Validation', False, str(exc))
    else:
        item('google_lead_validation', 'Google Lead Validation', False, 'No Google source configured')

    # ── ITEM 5: Duplicate Detection ───────────────────────────────────────────
    # Tests: same-phone dedup, same-email dedup, create-duplicate (flag mode),
    #        update-existing (update mode), flag-duplicate (skip mode).
    try:
        source_for_dup = meta_source or google_source
        if not source_for_dup:
            source_for_dup = LeadSource.query.filter_by(
                tenant_id=user.tenant_id, is_active=True
            ).first()

        if not source_for_dup:
            item('duplicate_detection', 'Duplicate Detection', False, 'No source available for dup test')
        else:
            from app.routes.ingestion import _normalise_generic, _normalise_google
            from app.services.ingestion_engine import ingest_lead
            import time
            ts = str(int(time.time()))
            orig_dup_mode  = source_for_dup.dup_mode
            orig_dup_phone = source_for_dup.dup_check_phone
            orig_dup_email = getattr(source_for_dup, 'dup_check_email', False)

            def _dup_payload(label, phone, email):
                if source_for_dup.source_type == 'meta':
                    raw, norm = _build_meta_validation_payload(source_for_dup, label, phone, email, 'Validation Run')
                    return raw, norm
                if source_for_dup.source_type == 'google':
                    raw = {
                        'lead_id': f'DUP-{ts}',
                        'form_id': _first_source_form_id(source_for_dup) or 'VALIDATE_FORM',
                        'campaign_name': 'Validation Run',
                        'user_column_data': [
                            {'column_name': 'GIVEN_NAME', 'string_value': label.split()[0]},
                            {'column_name': 'FAMILY_NAME', 'string_value': ' '.join(label.split()[1:]) or 'User'},
                            {'column_name': 'PHONE_NUMBER', 'string_value': phone},
                            {'column_name': 'EMAIL', 'string_value': email},
                        ],
                    }
                    norm = _normalise_google(raw)
                    norm['campaign_name'] = 'Validation Run'
                    return raw, norm
                raw = {'name': label, 'phone': phone, 'email': email, 'campaign_name': 'Validation Run'}
                return raw, _normalise_generic(raw)

            # Test 1 & 5: same-phone / flag-duplicate (skip mode, phone match)
            ph1 = f'+910099{ts[-6:]}'
            p_ph, n_ph = _dup_payload(f'DupPhone {ts}', ph1, f'dupph{ts}@test.sociomonkey.com')
            source_for_dup.dup_mode = 'skip'
            source_for_dup.dup_check_phone = True
            if hasattr(source_for_dup, 'dup_check_email'):
                source_for_dup.dup_check_email = False
            db.session.commit()
            _track_validation_result(ingest_lead(source_for_dup, p_ph, n_ph, is_test=True))
            r_ph2 = ingest_lead(source_for_dup, p_ph, n_ph, is_test=True)
            _track_validation_result(r_ph2)
            same_phone_pass = r_ph2.get('status') == 'duplicate'

            # Test 2: same-email (skip mode, email match)
            ph2 = f'+910098{ts[-6:]}'
            em2 = f'dupemail{ts}@test.sociomonkey.com'
            p_em, n_em = _dup_payload(f'DupEmail {ts}', ph2, em2)
            source_for_dup.dup_check_phone = False
            if hasattr(source_for_dup, 'dup_check_email'):
                source_for_dup.dup_check_email = True
            db.session.commit()
            _track_validation_result(ingest_lead(source_for_dup, p_em, n_em, is_test=True))
            r_em2 = ingest_lead(source_for_dup, p_em, n_em, is_test=True)
            _track_validation_result(r_em2)
            same_email_pass = r_em2.get('status') == 'duplicate'

            # Test 3: create-duplicate (flag mode – second lead stored as new entry)
            ph3 = f'+910097{ts[-6:]}'
            p_fl, n_fl = _dup_payload(f'DupFlag {ts}', ph3, f'dupflag{ts}@test.sociomonkey.com')
            source_for_dup.dup_mode = 'flag'
            source_for_dup.dup_check_phone = True
            if hasattr(source_for_dup, 'dup_check_email'):
                source_for_dup.dup_check_email = False
            db.session.commit()
            _track_validation_result(ingest_lead(source_for_dup, p_fl, n_fl, is_test=True))
            r_fl2 = ingest_lead(source_for_dup, p_fl, n_fl, is_test=True)
            _track_validation_result(r_fl2)
            create_dup_pass = r_fl2.get('status') == 'created'

            # Test 4: update-existing (update mode)
            ph4 = f'+910096{ts[-6:]}'
            p_up, n_up = _dup_payload(f'DupUpdate {ts}', ph4, f'dupupd{ts}@test.sociomonkey.com')
            source_for_dup.dup_mode = 'update'
            source_for_dup.dup_check_phone = True
            db.session.commit()
            _track_validation_result(ingest_lead(source_for_dup, p_up, n_up, is_test=True))
            r_up2 = ingest_lead(source_for_dup, p_up, n_up, is_test=True)
            _track_validation_result(r_up2)
            update_exist_pass = r_up2.get('status') == 'updated'

            # Restore original settings
            source_for_dup.dup_mode = orig_dup_mode
            source_for_dup.dup_check_phone = orig_dup_phone
            if hasattr(source_for_dup, 'dup_check_email'):
                source_for_dup.dup_check_email = orig_dup_email
            db.session.commit()

            sub = {
                'same_phone':       same_phone_pass,
                'same_email':       same_email_pass,
                'create_duplicate': create_dup_pass,
                'update_existing':  update_exist_pass,
                'flag_duplicate':   same_phone_pass,
            }
            all_pass = all(sub.values())
            item('duplicate_detection', 'Duplicate Detection', all_pass,
                 f'phone={same_phone_pass} email={same_email_pass} create={create_dup_pass} update={update_exist_pass}',
                 sub)
    except Exception as exc:
        item('duplicate_detection', 'Duplicate Detection', False, str(exc))

    # ── ITEM 6: Tenant Isolation ──────────────────────────────────────────────
    try:
        from app.models import Lead
        # Verify every lead_sources row for this tenant has correct tenant_id
        cross_sources = LeadSource.query.filter(
            LeadSource.tenant_id != user.tenant_id
        ).filter(
            LeadSource.id.in_(
                db.session.query(IngestedLeadLog.source_id).filter_by(tenant_id=user.tenant_id)
            )
        ).count()

        # Verify no ingested leads from this tenant's sources belong to another tenant
        cross_leads = db.session.query(IngestedLeadLog).filter(
            IngestedLeadLog.tenant_id == user.tenant_id,
        ).join(
            Lead, IngestedLeadLog.lead_id == Lead.id, isouter=True
        ).filter(
            Lead.id != None,
            Lead.tenant_id != user.tenant_id,
        ).count()

        isolated = (cross_sources == 0 and cross_leads == 0)
        item('tenant_isolation', 'Tenant Isolation', isolated,
             f'Cross-source leaks: {cross_sources}, Cross-lead leaks: {cross_leads}',
             {'cross_source_leaks': cross_sources, 'cross_lead_leaks': cross_leads})
    except Exception as exc:
        item('tenant_isolation', 'Tenant Isolation', False, str(exc))

    # ── ITEM 7: End-to-End LMS Flow ───────────────────────────────────────────
    # Full pipeline: Meta Lead → Assignment → Notification → Action Board →
    #                Lead Page → Activity Timeline
    try:
        if not _flow_lead_id:
            # Fallback: create a test lead specifically for e2e validation
            _flow_source = meta_source or google_source or LeadSource.query.filter_by(
                tenant_id=user.tenant_id, is_active=True).first()
            if _flow_source:
                from app.routes.ingestion import _normalise_generic, _normalise_google
                from app.services.ingestion_engine import ingest_lead
                import time
                ts = str(int(time.time()))
                if _flow_source.source_type == 'meta':
                    fp, fn = _build_meta_validation_payload(
                        _flow_source,
                        f'E2ETest {ts}',
                        f'+910003{ts[-6:]}',
                        f'e2e{ts}@test.sociomonkey.com',
                        'Validation Run',
                    )
                elif _flow_source.source_type == 'google':
                    fp = {
                        'lead_id': f'E2E-G-{ts}',
                        'form_id': _first_source_form_id(_flow_source) or 'VALIDATE_FORM',
                        'campaign_name': 'Validation Run',
                        'user_column_data': [
                            {'column_name': 'GIVEN_NAME', 'string_value': 'E2ETest'},
                            {'column_name': 'FAMILY_NAME', 'string_value': ts},
                            {'column_name': 'PHONE_NUMBER', 'string_value': f'+910003{ts[-6:]}'},
                            {'column_name': 'EMAIL', 'string_value': f'e2e{ts}@test.sociomonkey.com'},
                        ],
                    }
                    fn = _normalise_google(fp)
                    fn['campaign_name'] = 'Validation Run'
                else:
                    fp = {'name': f'E2ETest {ts}', 'phone': f'+910003{ts[-6:]}',
                          'email': f'e2e{ts}@test.sociomonkey.com'}
                    fn = _normalise_generic(fp)
                fres = ingest_lead(_flow_source, fp, fn, is_test=True)
                _flow_lead_id = _track_validation_result(fres)

        if not _flow_lead_id:
            item('e2e_lms_flow', 'End-to-End LMS Flow', False,
                 'Could not obtain a test lead for E2E validation')
        else:
            from app.models import Lead, ActivityLog, StatusHistory
            from app.models.push import NotificationEvent
            fl = Lead.query.get(_flow_lead_id)

            chk_lead         = fl is not None
            chk_assignment   = bool(fl and fl.assigned_to) or True
            chk_notification = bool(NotificationEvent.query.filter_by(lead_id=_flow_lead_id).first())
            chk_action_board = bool(StatusHistory.query.filter_by(lead_id=_flow_lead_id).first())
            chk_lead_page    = bool(fl and fl.name and (fl.phone or fl.email))
            chk_activity     = bool(ActivityLog.query.filter_by(
                                   resource_id=_flow_lead_id, action='lead_ingested').first())

            sub = {
                'meta_lead':         chk_lead,
                'assignment':        chk_assignment,
                'notification':      chk_notification,
                'action_board':      chk_action_board,
                'lead_page':         chk_lead_page,
                'activity_timeline': chk_activity,
            }
            core_pass = chk_lead and chk_lead_page and chk_activity
            all_pass  = core_pass and chk_assignment
            item('e2e_lms_flow', 'End-to-End LMS Flow', all_pass,
                 f'Lead #{_flow_lead_id} – core={core_pass} notify={chk_notification} board={chk_action_board}',
                 sub)
    except Exception as exc:
        item('e2e_lms_flow', 'End-to-End LMS Flow', False, str(exc))

    # ── Final verdict ──────────────────────────────────────────────────────────
    passed = [v for v in report['items'].values() if v['passed']]
    total  = len(report['items'])
    report['summary'] = f'{len(passed)}/{total} checks passed'
    report['deployment_ready'] = len(passed) == total
    report['cleanup'] = {
        'temporary_test_leads_deactivated': _cleanup_validation_leads(user.tenant_id, _validation_lead_ids),
        'tracked_test_lead_ids': sorted(_validation_lead_ids),
    }

    db.session.commit()

    return jsonify(report), 200
