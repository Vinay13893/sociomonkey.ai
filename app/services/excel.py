"""
ExcelService – handles import and export of lead data via openpyxl / pandas.
"""
from __future__ import annotations

import os
from io import BytesIO
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

from app.models.base import db
from app.models.lead import Lead
from app.models.user import User
from app.models.project import Project
from app.utils.activity import log_activity
from app.utils.leads import VALID_STATUSES


# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill('solid', fgColor='1E3A5F')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
ALT_ROW_FILL = PatternFill('solid', fgColor='EEF2F7')
TITLE_FONT = Font(bold=True, size=14, color='1E3A5F')
THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)

# Columns for the import template / export sheet
LEAD_COLUMNS = [
    'ID', 'Name', 'Phone', 'Email', 'Source',
    'Status', 'Budget Min', 'Budget Max',
    'Project ID', 'Project Name',
    'Assigned To (Email)', 'Assigned To (Name)',
    'Created At', 'Updated At',
]

TEMPLATE_COLUMNS = [
    'name', 'phone', 'email', 'source',
    'status', 'budget_min', 'budget_max',
    'project', 'project_id', 'assigned_to_email', 'sales_manager_email',
]

SAMPLE_DATA = [
    ['Amit Sharma', '9876543210', 'amit@example.com', 'Website',
     'new', '5000000', '8000000', 'Ganga Residency Phase 1', '', '', ''],
    ['Sunita Reddy', '9123456789', 'sunita@example.com', 'Referral',
     'interested', '3000000', '6000000', 'Ganga Heights', '', '', ''],
]


class ExcelService:

    # ------------------------------------------------------------------
    # Template
    # ------------------------------------------------------------------

    @staticmethod
    def build_import_template() -> BytesIO:
        """Return an in-memory .xlsx file users can fill to bulk-import leads."""
        wb = openpyxl.Workbook()

        # ---------- Instructions sheet ----------
        ws_info = wb.active
        ws_info.title = 'Instructions'

        ws_info['A1'] = 'Ganga Realty LMS – Lead Import Template'
        ws_info['A1'].font = TITLE_FONT
        ws_info.merge_cells('A1:F1')

        instructions = [
            ('Column', 'Required?', 'Allowed Values / Notes'),
            ('name', 'YES', 'Full name of the lead'),
            ('phone', 'YES', 'Mobile / contact number'),
            ('email', 'no', 'Email address'),
            ('source', 'YES', 'Website / Referral / Walk-in / Meta / Google / Email Campaign / Direct / Other / G1 / G2 / G3 / TP'),
            ('status', 'YES',
             'new / attempted / connected / interested / site_visit_planned / '
             'site_visit_done / negotiation / booking_done / lost / junk'),
            ('budget_min', 'no', 'Minimum budget in INR (numbers only)'),
            ('budget_max', 'no', 'Maximum budget in INR (numbers only)'),
            ('project', 'YES', 'Project name exactly as listed in the Projects page (e.g. Ganga Residency Phase 1)'),
            ('project_id', 'no', 'Numeric Project ID — use instead of project name if preferred'),
            ('assigned_to_email', 'no', 'Email address of the team member to assign this lead to'),
            ('sales_manager_email', 'no', 'Email address of the sales manager to assign to this lead'),
        ]

        for row_idx, row_data in enumerate(instructions, start=3):
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws_info.cell(row=row_idx, column=col_idx, value=val)
                if row_idx == 3:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill('solid', fgColor='D9E1F2')
                cell.border = THIN_BORDER
                cell.alignment = Alignment(wrap_text=True)

        ws_info.column_dimensions['A'].width = 22
        ws_info.column_dimensions['B'].width = 12
        ws_info.column_dimensions['C'].width = 70

        # ---------- Data entry sheet ----------
        ws = wb.create_sheet('Leads Import')

        for col_idx, col_name in enumerate(TEMPLATE_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = THIN_BORDER

        for row_idx, row_data in enumerate(SAMPLE_DATA, start=2):
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.fill = ALT_ROW_FILL
                cell.border = THIN_BORDER

        ws.freeze_panes = 'A2'
        for col_idx in range(1, len(TEMPLATE_COLUMNS) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 22

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # ------------------------------------------------------------------
    # Export
    # ------------------------------------------------------------------

    @staticmethod
    def export_leads(leads: list[Lead], title: str = 'Leads Export') -> BytesIO:
        """Convert a list of Lead objects into a formatted .xlsx workbook."""
        wb = openpyxl.Workbook()

        # ---------- Summary sheet ----------
        ws_summary = wb.active
        ws_summary.title = 'Summary'

        ws_summary['A1'] = f'Ganga Realty LMS – {title}'
        ws_summary['A1'].font = TITLE_FONT
        ws_summary.merge_cells('A1:D1')

        ws_summary['A3'] = 'Generated At'
        ws_summary['B3'] = datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')
        ws_summary['A4'] = 'Total Leads'
        ws_summary['B4'] = len(leads)

        status_counts: dict[str, int] = {}
        for lead in leads:
            status_counts[lead.status or 'unknown'] = (
                status_counts.get(lead.status or 'unknown', 0) + 1
            )

        ws_summary['A6'] = 'Status'
        ws_summary['B6'] = 'Count'
        ws_summary['A6'].font = Font(bold=True)
        ws_summary['B6'].font = Font(bold=True)

        for row_idx, (status, count) in enumerate(
            sorted(status_counts.items()), start=7
        ):
            ws_summary.cell(row=row_idx, column=1, value=status)
            ws_summary.cell(row=row_idx, column=2, value=count)

        ws_summary.column_dimensions['A'].width = 28
        ws_summary.column_dimensions['B'].width = 16

        # ---------- Leads data sheet ----------
        ws = wb.create_sheet('Leads')

        for col_idx, col_name in enumerate(LEAD_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = THIN_BORDER

        for row_idx, lead in enumerate(leads, start=2):
            fill = ALT_ROW_FILL if row_idx % 2 == 0 else PatternFill()
            row_values = [
                lead.id,
                lead.name,
                lead.phone or '',
                lead.email or '',
                lead.source or '',
                lead.status or '',
                lead.budget_min,
                lead.budget_max,
                lead.project_id,
                lead.project.name if lead.project else '',
                lead.assigned_user.email if lead.assigned_user else '',
                lead.assigned_user.name if lead.assigned_user else '',
                lead.created_at.strftime('%Y-%m-%d %H:%M') if lead.created_at else '',
                lead.updated_at.strftime('%Y-%m-%d %H:%M') if lead.updated_at else '',
            ]
            for col_idx, val in enumerate(row_values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.fill = fill
                cell.border = THIN_BORDER

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:{get_column_letter(len(LEAD_COLUMNS))}1'

        col_widths = [6, 25, 16, 28, 16, 22, 14, 14, 12, 25, 30, 25, 20, 20]
        for col_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # ------------------------------------------------------------------
    # Import Report
    # ------------------------------------------------------------------

    @staticmethod
    def build_import_report(imported: list, errors: list) -> BytesIO:
        """Generate an Excel report with two sheets: Imported and Failed."""
        wb = openpyxl.Workbook()

        SUCCESS_FILL = PatternFill('solid', fgColor='D1FAE5')
        FAIL_FILL    = PatternFill('solid', fgColor='FEE2E2')
        SUCCESS_FONT = Font(bold=True, color='065F46', size=11)
        FAIL_FONT    = Font(bold=True, color='991B1B', size=11)

        def _style_header(ws, headers, fill, font):
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = fill
                cell.font = font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = THIN_BORDER
            ws.row_dimensions[1].height = 22

        def _style_row(ws, row_idx):
            row_fill = ALT_ROW_FILL if row_idx % 2 == 0 else PatternFill()
            for cell in ws[row_idx]:
                cell.border = THIN_BORDER
                if row_fill.fill_type:
                    cell.fill = row_fill

        # ---- Sheet 1: Imported ----
        ws_ok = wb.active
        ws_ok.title = 'Imported Leads'
        ok_headers = ['#', 'Name', 'Phone', 'Email', 'Source', 'Status',
                      'Project', 'Assigned To', 'Sales Manager', 'Lead ID']
        _style_header(ws_ok, ok_headers, SUCCESS_FILL, SUCCESS_FONT)

        for i, lead in enumerate(imported, start=1):
            ws_ok.append([
                i,
                lead.get('name', ''),
                lead.get('phone', ''),
                lead.get('email', ''),
                lead.get('source', ''),
                lead.get('status', ''),
                lead.get('project_name', ''),
                lead.get('assigned_to_name', ''),
                lead.get('sales_manager_name', ''),
                lead.get('id', ''),
            ])
            _style_row(ws_ok, i + 1)

        col_widths_ok = [5, 28, 16, 28, 14, 20, 24, 24, 24, 10]
        for col_idx, width in enumerate(col_widths_ok, 1):
            ws_ok.column_dimensions[get_column_letter(col_idx)].width = width

        # ---- Sheet 2: Failed ----
        ws_fail = wb.create_sheet('Failed Rows')
        fail_headers = ['Row #', 'Error Reason', 'Name (if available)', 'Phone (if available)']
        _style_header(ws_fail, fail_headers, FAIL_FILL, FAIL_FONT)

        for i, err in enumerate(errors, start=1):
            ws_fail.append([
                err.get('row', ''),
                err.get('error', ''),
                err.get('name', ''),
                err.get('phone', ''),
            ])
            _style_row(ws_fail, i + 1)

        col_widths_fail = [8, 60, 28, 18]
        for col_idx, width in enumerate(col_widths_fail, 1):
            ws_fail.column_dimensions[get_column_letter(col_idx)].width = width

        # ---- Summary sheet ----
        ws_sum = wb.create_sheet('Summary', 0)
        ws_sum.title = 'Summary'
        ws_sum['A1'] = 'Ganga Realty LMS – Bulk Import Report'
        ws_sum['A1'].font = TITLE_FONT
        ws_sum.merge_cells('A1:C1')
        ws_sum['A3'] = 'Generated'
        ws_sum['B3'] = datetime.utcnow().strftime('%d %b %Y, %H:%M UTC')
        ws_sum['A4'] = 'Total Rows Processed'
        ws_sum['B4'] = len(imported) + len(errors)
        ws_sum['A5'] = 'Successfully Imported'
        ws_sum['B5'] = len(imported)
        ws_sum['B5'].font = Font(bold=True, color='065F46')
        ws_sum['A6'] = 'Failed / Skipped'
        ws_sum['B6'] = len(errors)
        ws_sum['B6'].font = Font(bold=True, color='991B1B')
        ws_sum.column_dimensions['A'].width = 26
        ws_sum.column_dimensions['B'].width = 22

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # ------------------------------------------------------------------
    # Export for Bulk Update (update-friendly template)
    # ------------------------------------------------------------------

    @staticmethod
    def export_leads_for_update(
        leads: list[Lead], title: str = 'Update Existing Leads'
    ) -> BytesIO:
        """Export leads in a format ready for bulk update (lead_id locked, phone locked)."""
        wb = openpyxl.Workbook()

        LOCKED_FILL = PatternFill('solid', fgColor='FEE2E2')
        LOCKED_FONT = Font(bold=True, color='991B1B', size=11)
        NOTE_FILL   = PatternFill('solid', fgColor='ECFDF5')
        NOTE_FONT   = Font(bold=True, color='065F46', size=11)

        UPDATE_COLUMNS = [
            'lead_id [DO NOT EDIT]',
            'customer_name',
            'phone [DO NOT EDIT]',
            'email',
            'status',
            'source',
            'project',
            'budget_min',
            'budget_max',
            'assigned_to_email',
            'notes (append only)',
            'callback_date',
            'callback_time',
        ]
        COL_WIDTHS  = [18, 28, 18, 30, 22, 18, 28, 14, 14, 32, 40, 16, 14]
        LOCKED_COLS = {0, 2}   # lead_id, phone
        NOTE_COLS   = {10}     # notes

        # ---------- Instructions sheet ----------
        ws_info = wb.active
        ws_info.title = 'Instructions'
        ws_info['A1'] = f'Ganga Realty LMS – {title}'
        ws_info['A1'].font = TITLE_FONT
        ws_info.merge_cells('A1:F1')

        instructions = [
            ('Column', 'Editable?', 'Notes'),
            ('lead_id [DO NOT EDIT]', 'NO', 'Used to identify the lead — do not change this value.'),
            ('customer_name', 'YES', 'Full name of the lead'),
            ('phone [DO NOT EDIT]', 'NO', 'Phone cannot be changed via this template.'),
            ('email', 'YES (manager+)', 'Update email address'),
            ('status', 'YES',
             'Must be one of: ' + ', '.join(VALID_STATUSES)),
            ('source', 'YES (manager+)', 'Lead source (Website, Referral, Walk-in, etc.)'),
            ('project', 'YES (manager+)', 'Project name exactly as listed in Projects page'),
            ('budget_min', 'YES (manager+)', 'Minimum budget (numbers only, no commas)'),
            ('budget_max', 'YES (manager+)', 'Maximum budget (numbers only, no commas)'),
            ('assigned_to_email', 'YES (manager+)',
             'Email of team member to assign. Leave blank to keep existing.'),
            ('notes (append only)', 'YES',
             'Text here is ADDED as a new note. Existing notes are NOT removed.'),
            ('callback_date', 'YES', 'Schedule callback: date in YYYY-MM-DD format'),
            ('callback_time', 'YES', 'Callback time HH:MM (24-h). Default: 09:00'),
        ]
        for row_idx, row_data in enumerate(instructions, start=3):
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws_info.cell(row=row_idx, column=col_idx, value=val)
                if row_idx == 3:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill('solid', fgColor='D9E1F2')
                cell.border = THIN_BORDER
                cell.alignment = Alignment(wrap_text=True)
        ws_info.column_dimensions['A'].width = 26
        ws_info.column_dimensions['B'].width = 18
        ws_info.column_dimensions['C'].width = 65

        # ---------- Update data sheet ----------
        ws = wb.create_sheet('Update Leads')
        for col_idx, col_name in enumerate(UPDATE_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            i = col_idx - 1
            if i in LOCKED_COLS:
                cell.font = LOCKED_FONT
                cell.fill = LOCKED_FILL
            elif i in NOTE_COLS:
                cell.font = NOTE_FONT
                cell.fill = NOTE_FILL
            else:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = THIN_BORDER

        for row_idx, lead in enumerate(leads, start=2):
            alt_fill = ALT_ROW_FILL if row_idx % 2 == 0 else PatternFill()
            row_values = [
                lead.id,
                lead.name,
                lead.phone or '',
                lead.email or '',
                lead.status or '',
                lead.source or '',
                lead.project.name if lead.project else '',
                lead.budget_min,
                lead.budget_max,
                lead.assigned_user.email if lead.assigned_user else '',
                '',  # notes — blank; user fills in
                '',  # callback_date
                '',  # callback_time
            ]
            for col_idx, val in enumerate(row_values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                i = col_idx - 1
                if i in LOCKED_COLS:
                    cell.fill = PatternFill('solid', fgColor='FFF5F5')
                elif alt_fill.fill_type:
                    cell.fill = alt_fill
                cell.border = THIN_BORDER

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:{get_column_letter(len(UPDATE_COLUMNS))}1'
        for col_idx, width in enumerate(COL_WIDTHS, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # ------------------------------------------------------------------
    # Bulk Update Report
    # ------------------------------------------------------------------

    @staticmethod
    def build_update_report(results: list) -> BytesIO:
        """Build a 4-sheet Excel report summarising a bulk update operation."""
        wb = openpyxl.Workbook()

        SUCCESS_FILL = PatternFill('solid', fgColor='D1FAE5')
        FAIL_FILL    = PatternFill('solid', fgColor='FEE2E2')
        SKIP_FILL    = PatternFill('solid', fgColor='FEF3C7')
        SUCCESS_FONT = Font(bold=True, color='065F46', size=11)
        FAIL_FONT    = Font(bold=True, color='991B1B', size=11)
        SKIP_FONT    = Font(bold=True, color='92400E', size=11)

        updated = [r for r in results if r.get('applied')]
        failed  = [r for r in results if not r.get('applied') and r.get('errors')]
        skipped = [r for r in results if not r.get('applied') and not r.get('errors')]

        def _hdr(ws, headers, fill, font):
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = fill
                cell.font = font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = THIN_BORDER
            ws.row_dimensions[1].height = 22

        def _row(ws, row_idx, fill):
            for cell in ws[row_idx]:
                cell.border = THIN_BORDER
                if fill.fill_type:
                    cell.fill = fill

        # Summary
        ws_sum = wb.active
        ws_sum.title = 'Summary'
        ws_sum['A1'] = 'Ganga Realty LMS – Bulk Update Report'
        ws_sum['A1'].font = TITLE_FONT
        ws_sum.merge_cells('A1:C1')
        ws_sum['A3'] = 'Generated'
        ws_sum['B3'] = datetime.utcnow().strftime('%d %b %Y, %H:%M UTC')
        ws_sum['A4'] = 'Total Rows'
        ws_sum['B4'] = len(results)
        ws_sum['A5'] = 'Updated'
        ws_sum['B5'] = len(updated)
        ws_sum['B5'].font = Font(bold=True, color='065F46')
        ws_sum['A6'] = 'Failed'
        ws_sum['B6'] = len(failed)
        ws_sum['B6'].font = Font(bold=True, color='991B1B')
        ws_sum['A7'] = 'Skipped (no changes)'
        ws_sum['B7'] = len(skipped)
        ws_sum['B7'].font = Font(bold=True, color='92400E')
        ws_sum.column_dimensions['A'].width = 26
        ws_sum.column_dimensions['B'].width = 18

        # Updated sheet
        ws_ok = wb.create_sheet('Updated Leads')
        _hdr(ws_ok, ['Row', 'Lead ID', 'Lead Name', 'Phone', 'Changes Made', 'Note Added', 'Callback Set'],
             SUCCESS_FILL, SUCCESS_FONT)
        for i, r in enumerate(updated, 2):
            changes = '; '.join(
                f'{k}: {v.get("old", "")} → {v.get("new", "")}'
                for k, v in (r.get('field_updates') or {}).items()
                if not k.startswith('_')
            )
            ws_ok.append([
                r.get('row', ''), r.get('lead_id', ''), r.get('lead_name', ''), r.get('phone', ''),
                changes,
                '✓' if r.get('note_text') else '',
                '✓' if r.get('callback_dt') else '',
            ])
            _row(ws_ok, i, ALT_ROW_FILL if i % 2 == 0 else PatternFill())
        for col_idx, w in enumerate([6, 10, 28, 18, 65, 14, 14], 1):
            ws_ok.column_dimensions[get_column_letter(col_idx)].width = w

        # Failed sheet
        ws_fail = wb.create_sheet('Failed Rows')
        _hdr(ws_fail, ['Row', 'Lead ID', 'Lead Name', 'Errors', 'Warnings'],
             FAIL_FILL, FAIL_FONT)
        for i, r in enumerate(failed, 2):
            ws_fail.append([
                r.get('row', ''), r.get('lead_id', ''), r.get('lead_name', ''),
                '; '.join(r.get('errors', [])),
                '; '.join(r.get('warnings', [])),
            ])
            _row(ws_fail, i, PatternFill('solid', fgColor='FFF5F5') if i % 2 == 0 else PatternFill())
        for col_idx, w in enumerate([6, 10, 28, 65, 50], 1):
            ws_fail.column_dimensions[get_column_letter(col_idx)].width = w

        # Skipped sheet
        ws_skip = wb.create_sheet('Skipped Rows')
        _hdr(ws_skip, ['Row', 'Lead ID', 'Lead Name', 'Reason'], SKIP_FILL, SKIP_FONT)
        for i, r in enumerate(skipped, 2):
            ws_skip.append([
                r.get('row', ''), r.get('lead_id', ''), r.get('lead_name', ''),
                '; '.join(r.get('warnings', [])) or 'No changes detected',
            ])
            _row(ws_skip, i, PatternFill('solid', fgColor='FFFBEB') if i % 2 == 0 else PatternFill())
        for col_idx, w in enumerate([6, 10, 28, 60], 1):
            ws_skip.column_dimensions[get_column_letter(col_idx)].width = w

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # ------------------------------------------------------------------
    # Import
    # ------------------------------------------------------------------

    @staticmethod
    def import_leads(file_storage, current_user) -> dict:
        """
        Parse an uploaded Excel/CSV file and bulk-create leads.

        Returns a dict with keys: success, failed, total, imported_leads, errors.
        """
        filename = file_storage.filename or ''
        ext = os.path.splitext(filename)[1].lower()

        allowed = {'.xlsx', '.xls', '.csv'}
        if ext not in allowed:
            return {
                'success': 0, 'failed': 0, 'total': 0,
                'imported_leads': [],
                'errors': [{'row': 0, 'error': f'File type not supported. Use {", ".join(allowed)}'}],
            }

        try:
            if ext == '.csv':
                df = pd.read_csv(file_storage)
            else:
                # Try to find the right sheet — prefer 'Leads Import', else
                # scan all sheets and use the first one that has a 'name' column.
                xl = pd.ExcelFile(file_storage)
                df = None
                preferred = [s for s in xl.sheet_names if 'import' in s.lower() or 'lead' in s.lower()]
                sheet_order = preferred + [s for s in xl.sheet_names if s not in preferred]
                for sheet in sheet_order:
                    candidate = xl.parse(sheet)
                    candidate.columns = candidate.columns.str.lower().str.strip()
                    if 'name' in candidate.columns:
                        df = candidate
                        break
                if df is None:
                    df = xl.parse(xl.sheet_names[0])
        except Exception as exc:
            return {
                'success': 0, 'failed': 0, 'total': 0,
                'imported_leads': [],
                'errors': [{'row': 0, 'error': f'Could not parse file: {exc}'}],
            }

        df.columns = df.columns.str.lower().str.strip()

        # Normalise common column name aliases
        aliases = {
            'full name': 'name', 'lead name': 'name', 'customer name': 'name',
            'client name': 'name', 'contact name': 'name',
            'mobile': 'phone', 'phone number': 'phone', 'contact': 'phone',
            'contact number': 'phone', 'mobile number': 'phone',
            'project name': 'project',
            'assign to': 'assigned_to_email', 'assignee': 'assigned_to_email',
            'assigned to (email)': 'assigned_to_email',
            'sales manager email': 'sales_manager_email',
            'sales manager': 'sales_manager_email',
            'manager email': 'sales_manager_email',
            'budget min': 'budget_min', 'budget max': 'budget_max',
            'min budget': 'budget_min', 'max budget': 'budget_max',
        }
        df.rename(columns=aliases, inplace=True)

        if 'name' not in df.columns:
            return {
                'success': 0, 'failed': 0, 'total': len(df),
                'imported_leads': [],
                'errors': [{'row': 0, 'error': 'File must contain a "name" column'}],
            }

        # ---------------------------------------------------------------
        # Pre-batch duplicate phone detection (single DB query)
        # ---------------------------------------------------------------
        def _norm_phone(raw):
            if raw is None:
                return None
            s = str(raw).strip()
            if not s or s.lower() == 'nan':
                return None
            # Excel stores phones as floats: 9876543210.0 → 9876543210
            if s.endswith('.0') and s[:-2].isdigit():
                s = s[:-2]
            return s or None

        file_phones: set = set()
        for _, frow in df.iterrows():
            p = _norm_phone(frow.get('phone'))
            if p:
                file_phones.add(p)

        existing_phone_map: dict = {}
        if file_phones:
            for el in Lead.query.filter(Lead.phone.in_(list(file_phones))).all():
                existing_phone_map[el.phone] = el

        seen_phones: set = set()  # phones added in this batch
        # ---------------------------------------------------------------

        imported: list[dict] = []
        errors: list[dict] = []

        for idx, row in df.iterrows():
            row_num = int(idx) + 2  # 1-indexed with header
            # Pre-extract for error reporting even before validation
            _raw_name = str(row.get('name', '')).strip()
            _raw_name = '' if _raw_name.lower() == 'nan' else _raw_name
            _raw_phone = _norm_phone(row.get('phone')) or ''

            def _err(msg):
                return {'row': row_num, 'error': msg, 'name': _raw_name, 'phone': _raw_phone}

            try:
                name = _raw_name
                if not name:
                    errors.append(_err('"name" is required'))
                    continue

                def safe_str(col):
                    val = row.get(col)
                    return str(val).strip() if val is not None and str(val).lower() != 'nan' else None

                def safe_float(col):
                    val = row.get(col)
                    try:
                        return float(val) if val is not None and str(val).lower() != 'nan' else None
                    except (ValueError, TypeError):
                        return None

                phone = _norm_phone(row.get('phone'))
                if not phone:
                    errors.append({'row': row_num, 'error': '"phone" is required'})
                    continue
                if phone in seen_phones:
                    errors.append(_err(f'Duplicate phone {phone} in this file'))
                    continue
                if phone in existing_phone_map:
                    ex = existing_phone_map[phone]
                    errors.append(_err(f'Duplicate: phone {phone} already exists (lead "{ex.name}" ID #{ex.id})'))
                    continue
                seen_phones.add(phone)

                source = safe_str('source') or 'Other'

                status = safe_str('status') or 'new'
                if status not in VALID_STATUSES:
                    errors.append(_err(f'Invalid status "{status}". Must be one of: {", ".join(VALID_STATUSES)}'))
                    continue

                project_name_col = safe_str('project')
                raw_pid = row.get('project_id')

                # Resolve project — try name first, then project_id; project is optional
                project_id = None

                if project_name_col:
                    proj = Project.query.filter(
                        Project.name.ilike(project_name_col.strip()),
                        Project.tenant_id == current_user.tenant_id,
                    ).first()
                    if not proj:
                        errors.append(_err(f'Project "{project_name_col}" not found'))
                        continue
                    project_id = proj.id
                elif raw_pid is not None and str(raw_pid).lower() != 'nan':
                    try:
                        project_id = int(float(raw_pid))
                        if not Project.query.get(project_id):
                            errors.append(_err(f'Project ID {project_id} not found'))
                            continue
                    except (ValueError, TypeError):
                        errors.append(_err('Invalid project_id'))
                        continue

                # Resolve assignee
                assigned_to_id = None
                assigned_email = safe_str('assigned_to_email')
                if assigned_email:
                    assignee = User.query.filter_by(email=assigned_email).first()
                    if not assignee:
                        errors.append(_err(f'User "{assigned_email}" not found'))
                        continue
                    if current_user.role == 'sales_manager' and assignee.manager_id != current_user.id:
                        errors.append(_err(f'Cannot assign to user outside your team'))
                        continue
                    assigned_to_id = assignee.id

                # Resolve sales manager
                sales_manager_id = None
                sm_email = safe_str('sales_manager_email')
                if sm_email:
                    sm_user = User.query.filter_by(email=sm_email).first()
                    if not sm_user:
                        errors.append(_err(f'Sales manager "{sm_email}" not found'))
                        continue
                    if sm_user.role not in ('sales_manager', 'superadmin'):
                        errors.append(_err(f'"{sm_email}" is not a sales manager'))
                        continue
                    sales_manager_id = sm_user.id

                lead = Lead(
                    name=name,
                    phone=phone,
                    email=safe_str('email'),
                    source=source,
                    status=status,
                    budget_min=safe_float('budget_min'),
                    budget_max=safe_float('budget_max'),
                    project_id=project_id,
                    assigned_to=assigned_to_id,
                    assigned_by=current_user.id if assigned_to_id else None,
                    sales_manager_id=sales_manager_id,
                    tenant_id=current_user.tenant_id,
                    created_by=current_user.id,
                )
                db.session.add(lead)
                db.session.flush()

                log_activity(
                    user_id=current_user.id,
                    action='import_lead',
                    module='leads',
                    resource_id=lead.id,
                    resource_type='Lead',
                    description=f'Imported lead: {name}',
                )
                imported.append(lead.to_dict())

            except Exception as exc:
                errors.append({'row': row_num, 'error': str(exc), 'name': _raw_name, 'phone': _raw_phone})

        try:
            db.session.commit()
        except Exception as exc:
            db.session.rollback()
            return {
                'success': 0, 'failed': len(df),
                'total': len(df), 'imported_leads': [],
                'errors': [{'row': 0, 'error': f'Database error: {exc}'}],
            }

        return {
            'success': len(imported),
            'failed': len(errors),
            'total': len(df),
            'imported_leads': imported,
            'errors': errors,
        }
