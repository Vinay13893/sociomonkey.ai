"""
ReportService – generates downloadable Excel reports for leads, team, and activity.
"""
from __future__ import annotations

from datetime import datetime, timedelta
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.models.activity import ActivityLog
from app.models.lead import Lead
from app.models.user import User

HEADER_FILL = PatternFill('solid', fgColor='1E3A5F')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
ALT_ROW_FILL = PatternFill('solid', fgColor='EEF2F7')
TITLE_FONT = Font(bold=True, size=14, color='1E3A5F')
THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)


def _write_header_row(ws, columns: list[str], row: int = 1):
    for col_idx, name in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER


def _write_data_row(ws, row_idx: int, values: list):
    fill = ALT_ROW_FILL if row_idx % 2 == 0 else PatternFill()
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.fill = fill
        cell.border = THIN_BORDER


class ReportService:

    # ------------------------------------------------------------------
    # Lead Report
    # ------------------------------------------------------------------

    @staticmethod
    def lead_report_excel(leads: list[Lead]) -> BytesIO:
        wb = openpyxl.Workbook()

        # ---- Summary ----
        ws_s = wb.active
        ws_s.title = 'Summary'
        ws_s['A1'] = 'Ganga Realty LMS – Lead Report'
        ws_s['A1'].font = TITLE_FONT
        ws_s.merge_cells('A1:C1')
        ws_s['A3'] = 'Generated'
        ws_s['B3'] = datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')
        ws_s['A4'] = 'Total Leads'
        ws_s['B4'] = len(leads)

        status_map: dict[str, int] = {}
        source_map: dict[str, int] = {}
        for lead in leads:
            s = lead.status or 'unknown'
            status_map[s] = status_map.get(s, 0) + 1
            src = lead.source or 'unknown'
            source_map[src] = source_map.get(src, 0) + 1

        ws_s['A6'] = 'By Status'
        ws_s['A6'].font = Font(bold=True)
        for r, (k, v) in enumerate(sorted(status_map.items()), start=7):
            ws_s.cell(row=r, column=1, value=k)
            ws_s.cell(row=r, column=2, value=v)

        ws_s['D6'] = 'By Source'
        ws_s['D6'].font = Font(bold=True)
        for r, (k, v) in enumerate(sorted(source_map.items()), start=7):
            ws_s.cell(row=r, column=4, value=k)
            ws_s.cell(row=r, column=5, value=v)

        ws_s.column_dimensions['A'].width = 28
        ws_s.column_dimensions['B'].width = 16
        ws_s.column_dimensions['D'].width = 28
        ws_s.column_dimensions['E'].width = 16

        # ---- All Leads ----
        ws = wb.create_sheet('All Leads')
        cols = [
            'ID', 'Name', 'Phone', 'Email', 'Source', 'Status',
            'Budget Min', 'Budget Max', 'Project', 'Assigned To',
            'Created At',
        ]
        _write_header_row(ws, cols)
        for ri, lead in enumerate(leads, start=2):
            _write_data_row(ws, ri, [
                lead.id, lead.name, lead.phone or '', lead.email or '',
                lead.source or '', lead.status or '',
                lead.budget_min, lead.budget_max,
                lead.project.name if lead.project else '',
                lead.assigned_user.name if lead.assigned_user else '',
                lead.created_at.strftime('%Y-%m-%d') if lead.created_at else '',
            ])

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:{get_column_letter(len(cols))}1'
        for ci, w in enumerate([6, 25, 16, 28, 16, 22, 14, 14, 25, 25, 14], start=1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # ------------------------------------------------------------------
    # Team Performance Report
    # ------------------------------------------------------------------

    @staticmethod
    def team_report_excel(users: list[User]) -> BytesIO:
        wb = openpyxl.Workbook()

        ws = wb.active
        ws.title = 'Team Performance'
        ws['A1'] = 'Ganga Realty LMS – Team Performance Report'
        ws['A1'].font = TITLE_FONT
        ws.merge_cells('A1:G1')
        ws['A2'] = f'Generated: {datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")}'

        cols = [
            'Name', 'Email', 'Role', 'Total Leads',
            'Converted', 'In Progress', 'Lost', 'Conversion Rate (%)',
        ]
        _write_header_row(ws, cols, row=4)

        converted_statuses = {'booking_done', 'negotiation'}

        for ri, u in enumerate(users, start=5):
            user_leads = Lead.query.filter_by(assigned_to=u.id).all()
            total = len(user_leads)
            converted = sum(1 for l in user_leads if l.status in converted_statuses)
            lost = sum(1 for l in user_leads if l.status == 'lost')
            in_progress = total - converted - lost
            rate = round(converted / total * 100, 2) if total else 0

            _write_data_row(ws, ri, [
                u.name, u.email, u.role,
                total, converted, in_progress, lost, rate,
            ])

        ws.freeze_panes = 'A5'
        for ci, w in enumerate([28, 32, 16, 14, 12, 14, 10, 20], start=1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # ------------------------------------------------------------------
    # Activity Log Report
    # ------------------------------------------------------------------

    @staticmethod
    def activity_report_excel(days: int = 30) -> BytesIO:
        wb = openpyxl.Workbook()

        cutoff = datetime.utcnow() - timedelta(days=days)
        logs = (
            ActivityLog.query
            .filter(ActivityLog.created_at >= cutoff)
            .order_by(ActivityLog.created_at.desc())
            .all()
        )

        ws = wb.active
        ws.title = 'Activity Log'
        ws['A1'] = f'Ganga Realty LMS – Activity Log (Last {days} days)'
        ws['A1'].font = TITLE_FONT
        ws.merge_cells('A1:G1')
        ws['A2'] = f'Generated: {datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")}'

        cols = ['Timestamp', 'User', 'Action', 'Module', 'Resource ID', 'Description', 'IP Address']
        _write_header_row(ws, cols, row=4)

        for ri, log in enumerate(logs, start=5):
            _write_data_row(ws, ri, [
                log.created_at.strftime('%Y-%m-%d %H:%M:%S') if log.created_at else '',
                log.user.name if log.user else 'Unknown',
                log.action,
                log.module,
                log.resource_id or '',
                log.description or '',
                log.ip_address or '',
            ])

        ws.freeze_panes = 'A5'
        for ci, w in enumerate([22, 25, 20, 15, 13, 50, 16], start=1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf
