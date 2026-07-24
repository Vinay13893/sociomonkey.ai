"""Bounded SQL analytics over existing LMS transactional entities."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timedelta
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill
from sqlalchemy import and_, case, func, or_

from app import db
from app.models.action_item import ActionItem
from app.models.business_configuration import LeadStatusConfiguration
from app.models.channel_partner import ChannelPartner
from app.models.lead import Lead
from app.models.location import Location, MeetingRoom
from app.models.organisation import (
    OrganisationUnit,
    OrganisationUnitMembership,
    ReportingRelationship,
)
from app.models.pipeline import PipelineTransition
from app.models.project import Project
from app.models.user import User
from app.models.visit import (
    Visit,
    VisitParticipant,
    VisitStatusConfiguration,
    VisitTypeConfiguration,
)
from app.services.permissions import capability_decision


REPORT_KEYS = (
    'pipeline', 'leads', 'organisations', 'users', 'projects', 'locations',
    'visits', 'reception', 'meeting-rooms', 'channel-partners',
    'action-items',
)
TERMINAL_ACTION_STATUSES = ('COMPLETED', 'CANCELLED', 'EXPIRED')
MAX_INTERACTIVE_ROWS = 100
MAX_EXPORT_ROWS = 5000
MAX_DATE_SPAN_DAYS = 366


class AnalyticsValidationError(ValueError):
    pass


@dataclass
class AnalyticsFilters:
    start: datetime
    end: datetime
    project_id: int | None = None
    location_id: int | None = None
    user_id: int | None = None
    organisation_unit_id: int | None = None
    limit: int = 50

    @classmethod
    def from_args(cls, args, *, export=False):
        now = datetime.utcnow()
        start = _parse_date(args.get('date_from')) or now - timedelta(days=29)
        end_date = _parse_date(args.get('date_to'))
        end = end_date + timedelta(days=1) if end_date else now + timedelta(seconds=1)
        if start >= end:
            raise AnalyticsValidationError('date_from must be before date_to')
        if end - start > timedelta(days=MAX_DATE_SPAN_DAYS + 1):
            raise AnalyticsValidationError(
                f'Date range cannot exceed {MAX_DATE_SPAN_DAYS} days'
            )
        maximum = MAX_EXPORT_ROWS if export else MAX_INTERACTIVE_ROWS
        limit = min(maximum, max(1, _safe_int(args.get('limit')) or 50))
        return cls(
            start=start,
            end=end,
            project_id=_safe_int(args.get('project_id')),
            location_id=_safe_int(args.get('location_id')),
            user_id=_safe_int(args.get('user_id')),
            organisation_unit_id=_safe_int(args.get('organisation_unit_id')),
            limit=limit,
        )

    def payload(self):
        return {
            'date_from': self.start.date().isoformat(),
            'date_to': (self.end - timedelta(seconds=1)).date().isoformat(),
            'project_id': self.project_id,
            'location_id': self.location_id,
            'user_id': self.user_id,
            'organisation_unit_id': self.organisation_unit_id,
            'limit': self.limit,
        }


def _parse_date(value):
    if not value:
        return None
    try:
        return datetime.strptime(str(value), '%Y-%m-%d')
    except ValueError as exc:
        raise AnalyticsValidationError('Dates must use YYYY-MM-DD') from exc


def _safe_int(value):
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _count(value):
    return int(value or 0)


def _number(value, places=1):
    return round(float(value or 0), places)


class AnalyticsService:
    """Builds aggregate-only reports with tenant and capability scope."""

    def __init__(self, user, filters: AnalyticsFilters):
        self.user = user
        self.tenant_id = user.tenant_id
        self.filters = filters
        self.scope, self.user_ids, self.unit_ids = self._resolve_scope()

    def _resolve_scope(self):
        for requested in ('TENANT', 'TEAM', 'ORGANISATION_UNIT', 'OWN'):
            decision = capability_decision(
                self.user, 'reports.view', requested
            )
            if not decision['allowed']:
                continue
            if requested in {'TENANT', 'PLATFORM'}:
                return requested, None, None
            if requested == 'TEAM':
                ids = {self.user.id}
                ids.update(
                    row[0] for row in db.session.query(
                        ReportingRelationship.user_id
                    ).filter(
                        ReportingRelationship.tenant_id == self.tenant_id,
                        ReportingRelationship.manager_id == self.user.id,
                        ReportingRelationship.is_active == True,  # noqa: E712
                    ).all()
                )
                ids.update(
                    row[0] for row in db.session.query(User.id).filter(
                        User.tenant_id == self.tenant_id,
                        User.manager_id == self.user.id,
                        User.is_active == True,  # noqa: E712
                    ).all()
                )
                return requested, ids, self._own_unit_ids()
            if requested == 'ORGANISATION_UNIT':
                unit_ids = self._own_unit_ids()
                ids = {self.user.id}
                if unit_ids:
                    ids.update(
                        row[0] for row in db.session.query(
                            OrganisationUnitMembership.user_id
                        ).filter(
                            OrganisationUnitMembership.tenant_id == self.tenant_id,
                            OrganisationUnitMembership.organisation_unit_id.in_(
                                unit_ids
                            ),
                            OrganisationUnitMembership.is_active == True,  # noqa: E712
                        ).all()
                    )
                return requested, ids, unit_ids
            return 'OWN', {self.user.id}, self._own_unit_ids()
        return 'NONE', set(), set()

    def _own_unit_ids(self):
        return {
            row[0] for row in db.session.query(
                OrganisationUnitMembership.organisation_unit_id
            ).filter(
                OrganisationUnitMembership.tenant_id == self.tenant_id,
                OrganisationUnitMembership.user_id == self.user.id,
                OrganisationUnitMembership.is_active == True,  # noqa: E712
            ).all()
        }

    def _lead_query(self):
        query = Lead.query.filter(
            Lead.tenant_id == self.tenant_id,
            Lead.is_active == True,  # noqa: E712
            Lead.is_test == False,  # noqa: E712
            Lead.created_at >= self.filters.start,
            Lead.created_at < self.filters.end,
        )
        if self.user_ids is not None:
            query = query.filter(Lead.assigned_to.in_(self.user_ids))
        if self.filters.project_id:
            query = query.filter(Lead.project_id == self.filters.project_id)
        if self.filters.user_id:
            self._validate_user_filter()
            query = query.filter(Lead.assigned_to == self.filters.user_id)
        return query

    def _visit_query(self):
        query = Visit.query.filter(
            Visit.tenant_id == self.tenant_id,
            Visit.is_active == True,  # noqa: E712
            Visit.created_at >= self.filters.start,
            Visit.created_at < self.filters.end,
        )
        if self.user_ids is not None:
            query = query.filter(or_(
                Visit.assigned_user_id.in_(self.user_ids),
                Visit.reception_assigned_user_id.in_(self.user_ids),
                Visit.created_by.in_(self.user_ids),
            ))
        if self.filters.project_id:
            query = query.filter(Visit.project_id == self.filters.project_id)
        if self.filters.location_id:
            query = query.filter(Visit.location_id == self.filters.location_id)
        if self.filters.user_id:
            self._validate_user_filter()
            query = query.filter(or_(
                Visit.assigned_user_id == self.filters.user_id,
                Visit.reception_assigned_user_id == self.filters.user_id,
            ))
        return query

    def _action_query(self):
        query = ActionItem.query.filter(
            ActionItem.tenant_id == self.tenant_id,
            ActionItem.is_active == True,  # noqa: E712
            ActionItem.created_at >= self.filters.start,
            ActionItem.created_at < self.filters.end,
        )
        if self.user_ids is not None:
            clauses = [ActionItem.assigned_user_id.in_(self.user_ids)]
            if self.unit_ids:
                clauses.append(ActionItem.organisation_unit_id.in_(self.unit_ids))
            query = query.filter(or_(*clauses))
        if self.filters.project_id:
            query = query.filter(ActionItem.project_id == self.filters.project_id)
        if self.filters.location_id:
            query = query.filter(ActionItem.location_id == self.filters.location_id)
        if self.filters.user_id:
            self._validate_user_filter()
            query = query.filter(
                ActionItem.assigned_user_id == self.filters.user_id
            )
        if self.filters.organisation_unit_id:
            self._validate_unit_filter()
            query = query.filter(
                ActionItem.organisation_unit_id
                == self.filters.organisation_unit_id
            )
        return query

    def _validate_user_filter(self):
        if self.user_ids is not None and self.filters.user_id not in self.user_ids:
            raise AnalyticsValidationError('User filter is outside report scope')

    def _validate_unit_filter(self):
        if (
            self.unit_ids is not None
            and self.filters.organisation_unit_id not in self.unit_ids
        ):
            raise AnalyticsValidationError(
                'Organisation unit filter is outside report scope'
            )

    def filter_options(self):
        users = User.query.filter(
            User.tenant_id == self.tenant_id,
            User.is_active == True,  # noqa: E712
        )
        units = OrganisationUnit.query.filter(
            OrganisationUnit.tenant_id == self.tenant_id,
            OrganisationUnit.is_active == True,  # noqa: E712
        )
        if self.user_ids is not None:
            users = users.filter(User.id.in_(self.user_ids))
        if self.unit_ids is not None:
            units = units.filter(OrganisationUnit.id.in_(self.unit_ids))
        return {
            'reports': list(REPORT_KEYS),
            'projects': [
                {'id': row.id, 'name': row.name}
                for row in Project.query.filter(
                    Project.tenant_id == self.tenant_id,
                    Project.is_active == True,  # noqa: E712
                ).order_by(Project.name).limit(200).all()
            ],
            'locations': [
                {'id': row.id, 'name': row.name}
                for row in Location.query.filter(
                    Location.tenant_id == self.tenant_id,
                    Location.is_active == True,  # noqa: E712
                ).order_by(Location.name).limit(200).all()
            ],
            'users': [
                {'id': row.id, 'name': row.name}
                for row in users.order_by(User.name).limit(200).all()
            ],
            'organisation_units': [
                {'id': row.id, 'name': row.name}
                for row in units.order_by(OrganisationUnit.name).limit(200).all()
            ],
            'scope': self.scope,
            'max_date_span_days': MAX_DATE_SPAN_DAYS,
        }

    def build(self, report_key):
        if report_key not in REPORT_KEYS:
            raise AnalyticsValidationError('Unknown report')
        builder = getattr(self, f'_report_{report_key.replace("-", "_")}')
        payload = builder()
        payload.update({
            'report': report_key,
            'filters': self.filters.payload(),
            'scope': self.scope,
            'generated_at_utc': datetime.utcnow().isoformat() + 'Z',
        })
        return payload

    def _result(self, title, summary, columns, rows):
        truncated = len(rows) > self.filters.limit
        return {
            'title': title,
            'summary': summary,
            'columns': columns,
            'rows': rows[:self.filters.limit],
            'truncated': truncated,
        }

    def _status_metadata(self):
        rows = LeadStatusConfiguration.query.filter_by(
            tenant_id=self.tenant_id
        ).all()
        return {
            row.internal_key: {
                'label': row.display_name,
                'colour': row.colour,
                'success': row.is_success,
                'lost': row.is_lost,
            }
            for row in rows
        }

    def _report_pipeline(self):
        lead_sq = self._lead_query().with_entities(
            Lead.id.label('id'), Lead.status.label('status')
        ).subquery()
        current = dict(db.session.query(
            lead_sq.c.status, func.count(lead_sq.c.id)
        ).group_by(lead_sq.c.status).all())
        transitions = PipelineTransition.query.filter(
            PipelineTransition.tenant_id == self.tenant_id,
            PipelineTransition.lead_id.in_(
                db.session.query(lead_sq.c.id)
            ),
            PipelineTransition.created_at >= self.filters.start,
            PipelineTransition.created_at < self.filters.end,
        )
        entries = dict(transitions.with_entities(
            PipelineTransition.to_stage_key,
            func.count(PipelineTransition.id),
        ).group_by(PipelineTransition.to_stage_key).all())
        exits = dict(transitions.with_entities(
            PipelineTransition.from_stage_key,
            func.count(PipelineTransition.id),
        ).filter(
            PipelineTransition.from_stage_key.isnot(None)
        ).group_by(PipelineTransition.from_stage_key).all())
        lagged = db.session.query(
            PipelineTransition.from_stage_key.label('stage_key'),
            PipelineTransition.created_at.label('created_at'),
            func.lag(PipelineTransition.created_at).over(
                partition_by=PipelineTransition.lead_id,
                order_by=(
                    PipelineTransition.created_at,
                    PipelineTransition.id,
                ),
            ).label('previous_created_at'),
        ).filter(
            PipelineTransition.tenant_id == self.tenant_id,
            PipelineTransition.lead_id.in_(
                db.session.query(lead_sq.c.id)
            ),
        ).subquery()
        dialect = db.session.get_bind().dialect.name
        if dialect == 'sqlite':
            hours = (
                func.julianday(lagged.c.created_at)
                - func.julianday(lagged.c.previous_created_at)
            ) * 24
        else:
            hours = func.extract(
                'epoch',
                lagged.c.created_at - lagged.c.previous_created_at,
            ) / 3600
        duration_rows = db.session.query(
            lagged.c.stage_key, func.avg(hours)
        ).filter(
            lagged.c.previous_created_at.isnot(None),
            lagged.c.stage_key.isnot(None),
            lagged.c.created_at >= self.filters.start,
            lagged.c.created_at < self.filters.end,
        ).group_by(lagged.c.stage_key).all()
        average_hours = dict(duration_rows)
        metadata = self._status_metadata()
        keys = set(current) | set(entries) | set(exits) | set(metadata)
        rows = []
        for key in keys:
            meta = metadata.get(key, {})
            rows.append({
                'key': key or 'unknown',
                'stage': meta.get('label') or (key or 'Unknown').replace('_', ' ').title(),
                'active_leads': _count(current.get(key)),
                'entries': _count(entries.get(key)),
                'exits': _count(exits.get(key)),
                'net_movement': _count(entries.get(key)) - _count(exits.get(key)),
                'average_hours': _number(average_hours.get(key)),
                'colour': meta.get('colour') or '#64748b',
            })
        rows.sort(key=lambda row: (-row['active_leads'], row['stage']))
        success_keys = {key for key, value in metadata.items() if value['success']}
        success = sum(current.get(key, 0) for key in success_keys)
        total = sum(current.values())
        return self._result(
            'Pipeline history and movement',
            [
                _metric('active_leads', 'Active Leads', total),
                _metric('movements', 'Movements', sum(entries.values())),
                _metric('successful', 'Successful', success),
                _metric(
                    'conversion_rate', 'Conversion %',
                    _number(success * 100 / total if total else 0),
                    'percent',
                ),
            ],
            _columns(
                ('stage', 'Stage'), ('active_leads', 'Active Leads', 'number'),
                ('entries', 'Entries', 'number'), ('exits', 'Exits', 'number'),
                ('net_movement', 'Net', 'number'),
                ('average_hours', 'Avg Hours in Stage', 'number'),
            ),
            rows,
        )

    def _report_leads(self):
        query = self._lead_query()
        status_meta = self._status_metadata()
        success_keys = [key for key, value in status_meta.items() if value['success']]
        lost_keys = [key for key, value in status_meta.items() if value['lost']]
        rows = query.with_entities(
            func.coalesce(Lead.source, 'Unspecified').label('source'),
            func.count(Lead.id).label('total'),
            func.sum(case((Lead.assigned_to.is_(None), 1), else_=0)).label(
                'unassigned'
            ),
            func.sum(case((Lead.status.in_(success_keys), 1), else_=0)).label(
                'successful'
            ),
            func.sum(case((Lead.status.in_(lost_keys), 1), else_=0)).label('lost'),
        ).group_by(func.coalesce(Lead.source, 'Unspecified')).order_by(
            func.count(Lead.id).desc()
        ).limit(self.filters.limit + 1).all()
        data = [{
            'source': row.source,
            'leads': _count(row.total),
            'assigned': _count(row.total) - _count(row.unassigned),
            'unassigned': _count(row.unassigned),
            'successful': _count(row.successful),
            'lost': _count(row.lost),
            'conversion_rate': _number(
                _count(row.successful) * 100 / _count(row.total)
                if row.total else 0
            ),
        } for row in rows]
        total = query.count()
        assigned = query.filter(Lead.assigned_to.isnot(None)).count()
        return self._result(
            'Lead acquisition and outcomes',
            [
                _metric('leads', 'Leads', total),
                _metric('assigned', 'Assigned', assigned),
                _metric('unassigned', 'Unassigned', total - assigned),
                _metric('sources', 'Sources', len(data)),
            ],
            _columns(
                ('source', 'Source'), ('leads', 'Leads', 'number'),
                ('assigned', 'Assigned', 'number'),
                ('unassigned', 'Unassigned', 'number'),
                ('successful', 'Successful', 'number'),
                ('lost', 'Lost', 'number'),
                ('conversion_rate', 'Conversion %', 'percent'),
            ),
            data,
        )

    def _report_organisations(self):
        lead_sq = self._lead_query().with_entities(
            Lead.assigned_to.label('user_id'),
            func.count(Lead.id).label('lead_count'),
        ).group_by(Lead.assigned_to).subquery()
        action_sq = self._action_query().with_entities(
            ActionItem.assigned_user_id.label('user_id'),
            func.count(ActionItem.id).label('action_count'),
            func.sum(case(
                (ActionItem.status_key == 'COMPLETED', 1), else_=0
            )).label('completed_count'),
        ).group_by(ActionItem.assigned_user_id).subquery()
        query = db.session.query(
            OrganisationUnit.id, OrganisationUnit.name,
            func.count(func.distinct(OrganisationUnitMembership.user_id)),
            func.sum(func.coalesce(lead_sq.c.lead_count, 0)),
            func.sum(func.coalesce(action_sq.c.action_count, 0)),
            func.sum(func.coalesce(action_sq.c.completed_count, 0)),
        ).outerjoin(
            OrganisationUnitMembership,
            and_(
                OrganisationUnitMembership.organisation_unit_id
                == OrganisationUnit.id,
                OrganisationUnitMembership.is_active == True,  # noqa: E712
                OrganisationUnitMembership.is_primary == True,  # noqa: E712
            ),
        ).outerjoin(
            lead_sq, lead_sq.c.user_id == OrganisationUnitMembership.user_id
        ).outerjoin(
            action_sq, action_sq.c.user_id == OrganisationUnitMembership.user_id
        ).filter(
            OrganisationUnit.tenant_id == self.tenant_id,
            OrganisationUnit.is_active == True,  # noqa: E712
        )
        if self.unit_ids is not None:
            query = query.filter(OrganisationUnit.id.in_(self.unit_ids))
        if self.filters.organisation_unit_id:
            self._validate_unit_filter()
            query = query.filter(
                OrganisationUnit.id == self.filters.organisation_unit_id
            )
        rows = query.group_by(
            OrganisationUnit.id, OrganisationUnit.name
        ).order_by(
            func.sum(func.coalesce(lead_sq.c.lead_count, 0)).desc()
        ).limit(self.filters.limit + 1).all()
        data = [{
            'id': row[0], 'organisation_unit': row[1],
            'users': _count(row[2]), 'leads': _count(row[3]),
            'actions': _count(row[4]), 'completed_actions': _count(row[5]),
            'completion_rate': _number(
                _count(row[5]) * 100 / _count(row[4]) if row[4] else 0
            ),
        } for row in rows]
        return self._result(
            'Organisation unit performance',
            [
                _metric('units', 'Units', len(data)),
                _metric('users', 'Users', sum(row['users'] for row in data)),
                _metric('leads', 'Leads', sum(row['leads'] for row in data)),
                _metric(
                    'actions', 'Actions', sum(row['actions'] for row in data)
                ),
            ],
            _columns(
                ('organisation_unit', 'Organisation Unit'),
                ('users', 'Users', 'number'), ('leads', 'Leads', 'number'),
                ('actions', 'Actions', 'number'),
                ('completed_actions', 'Completed', 'number'),
                ('completion_rate', 'Completion %', 'percent'),
            ),
            data,
        )

    def _report_users(self):
        lead_sq = self._lead_query().with_entities(
            Lead.assigned_to.label('user_id'),
            func.count(Lead.id).label('lead_count'),
        ).group_by(Lead.assigned_to).subquery()
        visit_sq = self._visit_query().with_entities(
            Visit.assigned_user_id.label('user_id'),
            func.count(Visit.id).label('visit_count'),
        ).group_by(Visit.assigned_user_id).subquery()
        action_sq = self._action_query().with_entities(
            ActionItem.assigned_user_id.label('user_id'),
            func.count(ActionItem.id).label('action_count'),
            func.sum(case(
                (ActionItem.status_key == 'COMPLETED', 1), else_=0
            )).label('completed_count'),
            func.sum(case((
                and_(
                    ActionItem.due_at < datetime.utcnow(),
                    ActionItem.status_key.notin_(TERMINAL_ACTION_STATUSES),
                ),
                1,
            ), else_=0)).label('overdue_count'),
        ).group_by(ActionItem.assigned_user_id).subquery()
        query = db.session.query(
            User.id, User.name, User.role,
            func.coalesce(lead_sq.c.lead_count, 0),
            func.coalesce(visit_sq.c.visit_count, 0),
            func.coalesce(action_sq.c.action_count, 0),
            func.coalesce(action_sq.c.completed_count, 0),
            func.coalesce(action_sq.c.overdue_count, 0),
        ).outerjoin(
            lead_sq, lead_sq.c.user_id == User.id
        ).outerjoin(
            visit_sq, visit_sq.c.user_id == User.id
        ).outerjoin(
            action_sq, action_sq.c.user_id == User.id
        ).filter(
            User.tenant_id == self.tenant_id,
            User.is_active == True,  # noqa: E712
        )
        if self.user_ids is not None:
            query = query.filter(User.id.in_(self.user_ids))
        if self.filters.user_id:
            self._validate_user_filter()
            query = query.filter(User.id == self.filters.user_id)
        rows = query.order_by(
            func.coalesce(lead_sq.c.lead_count, 0).desc(), User.name
        ).limit(self.filters.limit + 1).all()
        data = [{
            'id': row[0], 'user': row[1],
            'role': (row[2] or '').replace('_', ' ').title(),
            'leads': _count(row[3]), 'visits': _count(row[4]),
            'actions': _count(row[5]), 'completed_actions': _count(row[6]),
            'overdue_actions': _count(row[7]),
        } for row in rows]
        return self._result(
            'User workload and productivity',
            [
                _metric('users', 'Users', len(data)),
                _metric('leads', 'Leads', sum(row['leads'] for row in data)),
                _metric('visits', 'Visits', sum(row['visits'] for row in data)),
                _metric(
                    'overdue', 'Overdue Actions',
                    sum(row['overdue_actions'] for row in data),
                ),
            ],
            _columns(
                ('user', 'User'), ('role', 'Role'),
                ('leads', 'Leads', 'number'), ('visits', 'Visits', 'number'),
                ('actions', 'Actions', 'number'),
                ('completed_actions', 'Completed', 'number'),
                ('overdue_actions', 'Overdue', 'number'),
            ),
            data,
        )

    def _report_projects(self):
        lead_sq = self._lead_query().with_entities(
            Lead.project_id.label('project_id'),
            func.count(Lead.id).label('lead_count'),
        ).group_by(Lead.project_id).subquery()
        visit_sq = self._visit_query().with_entities(
            Visit.project_id.label('project_id'),
            func.count(Visit.id).label('visit_count'),
            func.sum(case(
                (Visit.status_key == 'COMPLETED', 1), else_=0
            )).label('completed_count'),
        ).group_by(Visit.project_id).subquery()
        rows = db.session.query(
            Project.id, Project.name,
            func.coalesce(lead_sq.c.lead_count, 0),
            func.coalesce(visit_sq.c.visit_count, 0),
            func.coalesce(visit_sq.c.completed_count, 0),
        ).outerjoin(
            lead_sq, lead_sq.c.project_id == Project.id
        ).outerjoin(
            visit_sq, visit_sq.c.project_id == Project.id
        ).filter(
            Project.tenant_id == self.tenant_id,
            Project.is_active == True,  # noqa: E712
        ).order_by(
            func.coalesce(lead_sq.c.lead_count, 0).desc()
        ).limit(self.filters.limit + 1).all()
        data = [{
            'id': row[0], 'project': row[1], 'leads': _count(row[2]),
            'visits': _count(row[3]), 'completed_visits': _count(row[4]),
            'visit_completion_rate': _number(
                _count(row[4]) * 100 / _count(row[3]) if row[3] else 0
            ),
        } for row in rows]
        return self._result(
            'Project performance',
            [
                _metric('projects', 'Projects', len(data)),
                _metric('leads', 'Leads', sum(row['leads'] for row in data)),
                _metric('visits', 'Visits', sum(row['visits'] for row in data)),
                _metric(
                    'completed_visits', 'Completed Visits',
                    sum(row['completed_visits'] for row in data),
                ),
            ],
            _columns(
                ('project', 'Project'), ('leads', 'Leads', 'number'),
                ('visits', 'Visits', 'number'),
                ('completed_visits', 'Completed Visits', 'number'),
                ('visit_completion_rate', 'Completion %', 'percent'),
            ),
            data,
        )

    def _report_locations(self):
        visit_sq = self._visit_query().with_entities(
            Visit.location_id.label('location_id'),
            func.count(Visit.id).label('visit_count'),
            func.sum(case(
                (Visit.status_key == 'COMPLETED', 1), else_=0
            )).label('completed_count'),
        ).group_by(Visit.location_id).subquery()
        room_sq = db.session.query(
            MeetingRoom.location_id.label('location_id'),
            func.count(MeetingRoom.id).label('room_count'),
        ).filter(
            MeetingRoom.tenant_id == self.tenant_id,
            MeetingRoom.is_active == True,  # noqa: E712
        ).group_by(MeetingRoom.location_id).subquery()
        rows = db.session.query(
            Location.id, Location.name, Location.location_type,
            func.coalesce(visit_sq.c.visit_count, 0),
            func.coalesce(visit_sq.c.completed_count, 0),
            func.coalesce(room_sq.c.room_count, 0),
        ).outerjoin(
            visit_sq, visit_sq.c.location_id == Location.id
        ).outerjoin(
            room_sq, room_sq.c.location_id == Location.id,
        ).filter(
            Location.tenant_id == self.tenant_id,
            Location.is_active == True,  # noqa: E712
        )
        if self.filters.location_id:
            rows = rows.filter(Location.id == self.filters.location_id)
        rows = rows.order_by(
            func.coalesce(visit_sq.c.visit_count, 0).desc()
        ).limit(self.filters.limit + 1).all()
        data = [{
            'id': row[0], 'location': row[1],
            'type': (row[2] or '').replace('_', ' ').title(),
            'visits': _count(row[3]), 'completed_visits': _count(row[4]),
            'meeting_rooms': _count(row[5]),
        } for row in rows]
        return self._result(
            'Location and Gallery activity',
            [
                _metric('locations', 'Locations', len(data)),
                _metric('visits', 'Visits', sum(row['visits'] for row in data)),
                _metric(
                    'completed', 'Completed',
                    sum(row['completed_visits'] for row in data),
                ),
                _metric(
                    'rooms', 'Meeting Rooms',
                    sum(row['meeting_rooms'] for row in data),
                ),
            ],
            _columns(
                ('location', 'Location'), ('type', 'Type'),
                ('visits', 'Visits', 'number'),
                ('completed_visits', 'Completed', 'number'),
                ('meeting_rooms', 'Rooms', 'number'),
            ),
            data,
        )

    def _report_visits(self):
        query = self._visit_query()
        type_labels = {
            row.internal_key: row.display_name
            for row in VisitTypeConfiguration.query.filter_by(
                tenant_id=self.tenant_id
            ).all()
        }
        status_labels = {
            row.internal_key: row.display_name
            for row in VisitStatusConfiguration.query.filter_by(
                tenant_id=self.tenant_id
            ).all()
        }
        rows = query.with_entities(
            Visit.visit_type_key, Visit.status_key,
            func.count(Visit.id), func.sum(Visit.visitor_count),
            func.avg(
                _visit_duration_minutes()
            ),
        ).group_by(
            Visit.visit_type_key, Visit.status_key
        ).order_by(func.count(Visit.id).desc()).limit(
            self.filters.limit + 1
        ).all()
        data = [{
            'visit_type': type_labels.get(row[0], _display_key(row[0])),
            'status': status_labels.get(row[1], _display_key(row[1])),
            'visits': _count(row[2]), 'visitors': _count(row[3]),
            'average_minutes': _number(row[4], 0),
        } for row in rows]
        total = query.count()
        completed = query.filter(Visit.status_key == 'COMPLETED').count()
        return self._result(
            'Visit activity',
            [
                _metric('visits', 'Visits', total),
                _metric('completed', 'Completed', completed),
                _metric('open', 'Open', total - completed),
                _metric(
                    'completion_rate', 'Completion %',
                    _number(completed * 100 / total if total else 0),
                    'percent',
                ),
            ],
            _columns(
                ('visit_type', 'Visit Type'), ('status', 'Status'),
                ('visits', 'Visits', 'number'),
                ('visitors', 'Visitors', 'number'),
                ('average_minutes', 'Avg Minutes', 'number'),
            ),
            data,
        )

    def _report_reception(self):
        query = self._visit_query()
        rows = query.with_entities(
            Visit.location_id, Location.name, Visit.status_key,
            func.count(Visit.id),
            func.sum(case((Visit.visit_type_key == 'WALK_IN', 1), else_=0)),
            func.sum(Visit.visitor_count),
        ).join(
            Location, Location.id == Visit.location_id
        ).group_by(
            Visit.location_id, Location.name, Visit.status_key
        ).order_by(
            Location.name, func.count(Visit.id).desc()
        ).limit(self.filters.limit + 1).all()
        data = [{
            'location': row[1], 'status': _display_key(row[2]),
            'visits': _count(row[3]), 'walk_ins': _count(row[4]),
            'visitors': _count(row[5]),
        } for row in rows]
        checked_in = query.filter(
            Visit.status_key.in_(('CHECKED_IN', 'WAITING', 'IN_PROGRESS'))
        ).count()
        return self._result(
            'Reception operations',
            [
                _metric('arrivals', 'Arrivals', query.count()),
                _metric('inside', 'Currently Inside', checked_in),
                _metric(
                    'walk_ins', 'Walk-ins',
                    query.filter(Visit.visit_type_key == 'WALK_IN').count(),
                ),
                _metric(
                    'no_shows', 'No Shows',
                    query.filter(Visit.status_key == 'NO_SHOW').count(),
                ),
            ],
            _columns(
                ('location', 'Location'), ('status', 'Status'),
                ('visits', 'Visits', 'number'),
                ('walk_ins', 'Walk-ins', 'number'),
                ('visitors', 'Visitors', 'number'),
            ),
            data,
        )

    def _report_meeting_rooms(self):
        visit_sq = self._visit_query().with_entities(
            Visit.meeting_room_id.label('meeting_room_id'),
            func.count(Visit.id).label('visit_count'),
            func.sum(case(
                (Visit.status_key == 'COMPLETED', 1), else_=0
            )).label('completed_count'),
        ).group_by(Visit.meeting_room_id).subquery()
        rows = db.session.query(
            MeetingRoom.id, MeetingRoom.name, Location.name,
            MeetingRoom.capacity, MeetingRoom.status,
            func.coalesce(visit_sq.c.visit_count, 0),
            func.coalesce(visit_sq.c.completed_count, 0),
        ).join(
            Location, Location.id == MeetingRoom.location_id
        ).outerjoin(
            visit_sq, visit_sq.c.meeting_room_id == MeetingRoom.id
        ).filter(
            MeetingRoom.tenant_id == self.tenant_id,
            MeetingRoom.is_active == True,  # noqa: E712
        )
        if self.filters.location_id:
            rows = rows.filter(MeetingRoom.location_id == self.filters.location_id)
        rows = rows.order_by(
            func.coalesce(visit_sq.c.visit_count, 0).desc(), MeetingRoom.name
        ).limit(self.filters.limit + 1).all()
        data = [{
            'id': row[0], 'meeting_room': row[1], 'location': row[2],
            'capacity': _count(row[3]), 'current_status': _display_key(row[4]),
            'visits': _count(row[5]), 'completed_visits': _count(row[6]),
        } for row in rows]
        return self._result(
            'Meeting Room usage',
            [
                _metric('rooms', 'Rooms', len(data)),
                _metric('capacity', 'Capacity', sum(row['capacity'] for row in data)),
                _metric('visits', 'Visits', sum(row['visits'] for row in data)),
                _metric(
                    'completed', 'Completed Visits',
                    sum(row['completed_visits'] for row in data),
                ),
            ],
            _columns(
                ('meeting_room', 'Meeting Room'), ('location', 'Location'),
                ('capacity', 'Capacity', 'number'),
                ('current_status', 'Current Status'),
                ('visits', 'Visits', 'number'),
                ('completed_visits', 'Completed Visits', 'number'),
            ),
            data,
        )

    def _report_channel_partners(self):
        lead_sq = self._lead_query().with_entities(
            Lead.channel_partner_id.label('channel_partner_id'),
            func.count(Lead.id).label('lead_count'),
        ).group_by(Lead.channel_partner_id).subquery()
        visit_sq = self._visit_query().with_entities(
            Visit.id.label('id')
        ).subquery()
        participant_sq = db.session.query(
            VisitParticipant.reference_id.label('channel_partner_id'),
            func.count(func.distinct(VisitParticipant.visit_id)).label(
                'visit_count'
            ),
        ).filter(
            VisitParticipant.tenant_id == self.tenant_id,
            VisitParticipant.participant_type == 'CHANNEL_PARTNER',
            VisitParticipant.visit_id.in_(db.session.query(visit_sq.c.id)),
        ).group_by(VisitParticipant.reference_id).subquery()
        rows = db.session.query(
            ChannelPartner.id, ChannelPartner.name,
            ChannelPartner.partner_type,
            func.coalesce(lead_sq.c.lead_count, 0),
            func.coalesce(participant_sq.c.visit_count, 0),
        ).outerjoin(
            lead_sq, lead_sq.c.channel_partner_id == ChannelPartner.id
        ).outerjoin(
            participant_sq,
            participant_sq.c.channel_partner_id == ChannelPartner.id,
        ).filter(
            ChannelPartner.tenant_id == self.tenant_id,
            ChannelPartner.is_active == True,  # noqa: E712
        ).order_by(
            func.coalesce(lead_sq.c.lead_count, 0).desc(),
            ChannelPartner.name,
        ).limit(self.filters.limit + 1).all()
        data = [{
            'id': row[0], 'channel_partner': row[1],
            'type': _display_key(row[2]),
            'attributed_leads': _count(row[3]),
            'visits': _count(row[4]),
        } for row in rows]
        return self._result(
            'Channel Partner activity',
            [
                _metric('partners', 'Active Partners', len(data)),
                _metric(
                    'attributed_leads', 'Attributed Leads',
                    sum(row['attributed_leads'] for row in data),
                ),
                _metric('visits', 'Visits', sum(row['visits'] for row in data)),
                _metric(
                    'engaged', 'Engaged Partners',
                    sum(
                        1 for row in data
                        if row['attributed_leads'] or row['visits']
                    ),
                ),
            ],
            _columns(
                ('channel_partner', 'Channel Partner'), ('type', 'Type'),
                ('attributed_leads', 'Attributed Leads', 'number'),
                ('visits', 'Visits', 'number'),
            ),
            data,
        )

    def _report_action_items(self):
        query = self._action_query()
        rows = query.with_entities(
            ActionItem.action_type_key, ActionItem.priority_key,
            func.count(ActionItem.id),
            func.sum(case(
                (ActionItem.status_key == 'COMPLETED', 1), else_=0
            )),
            func.sum(case((
                and_(
                    ActionItem.due_at < datetime.utcnow(),
                    ActionItem.status_key.notin_(TERMINAL_ACTION_STATUSES),
                ),
                1,
            ), else_=0)),
        ).group_by(
            ActionItem.action_type_key, ActionItem.priority_key
        ).order_by(func.count(ActionItem.id).desc()).limit(
            self.filters.limit + 1
        ).all()
        data = [{
            'action_type': _display_key(row[0]),
            'priority': _display_key(row[1]),
            'actions': _count(row[2]), 'completed': _count(row[3]),
            'overdue': _count(row[4]),
            'completion_rate': _number(
                _count(row[3]) * 100 / _count(row[2]) if row[2] else 0
            ),
        } for row in rows]
        total = query.count()
        completed = query.filter(ActionItem.status_key == 'COMPLETED').count()
        overdue = query.filter(
            ActionItem.due_at < datetime.utcnow(),
            ActionItem.status_key.notin_(TERMINAL_ACTION_STATUSES),
        ).count()
        return self._result(
            'Action Item productivity',
            [
                _metric('actions', 'Actions', total),
                _metric('completed', 'Completed', completed),
                _metric('open', 'Open', total - completed),
                _metric('overdue', 'Overdue', overdue),
            ],
            _columns(
                ('action_type', 'Action Type'), ('priority', 'Priority'),
                ('actions', 'Actions', 'number'),
                ('completed', 'Completed', 'number'),
                ('overdue', 'Overdue', 'number'),
                ('completion_rate', 'Completion %', 'percent'),
            ),
            data,
        )


def _display_key(value):
    return str(value or 'Unspecified').replace('_', ' ').replace('-', ' ').title()


def _metric(key, label, value, value_type='number'):
    return {'key': key, 'label': label, 'value': value, 'type': value_type}


def _columns(*definitions):
    return [
        {
            'key': item[0],
            'label': item[1],
            'type': item[2] if len(item) > 2 else 'text',
        }
        for item in definitions
    ]


def _visit_duration_minutes():
    """Portable SQL expression for completed Visit duration in minutes."""
    dialect = db.session.get_bind().dialect.name
    if dialect == 'sqlite':
        return case((
            Visit.actual_check_out.isnot(None),
            (
                func.julianday(Visit.actual_check_out)
                - func.julianday(Visit.actual_check_in)
            ) * 1440,
        ), else_=None)
    return case((
        Visit.actual_check_out.isnot(None),
        func.extract(
            'epoch', Visit.actual_check_out - Visit.actual_check_in
        ) / 60,
    ), else_=None)


def analytics_workbook(payload):
    """Create a separate aggregate export without exposing transactional rows."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Analytics'
    sheet['A1'] = payload['title']
    sheet['A1'].font = Font(size=14, bold=True)
    sheet['A2'] = f"Generated UTC: {payload['generated_at_utc']}"
    sheet['A3'] = (
        f"Period: {payload['filters']['date_from']} to "
        f"{payload['filters']['date_to']}"
    )
    row_number = 5
    for metric in payload.get('summary', []):
        sheet.cell(row=row_number, column=1, value=metric['label'])
        sheet.cell(row=row_number, column=2, value=metric['value'])
        row_number += 1
    row_number += 1
    columns = payload.get('columns', [])
    for index, column in enumerate(columns, 1):
        cell = sheet.cell(row=row_number, column=index, value=column['label'])
        cell.font = Font(color='FFFFFF', bold=True)
        cell.fill = PatternFill('solid', fgColor='1E3A5F')
    for data_row in payload.get('rows', []):
        row_number += 1
        for index, column in enumerate(columns, 1):
            sheet.cell(
                row=row_number,
                column=index,
                value=data_row.get(column['key']),
            )
    sheet.freeze_panes = f'A{row_number - len(payload.get("rows", []))}'
    for index, column in enumerate(columns, 1):
        sheet.column_dimensions[
            openpyxl.utils.get_column_letter(index)
        ].width = min(36, max(14, len(column['label']) + 4))
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer
